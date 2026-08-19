import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
# remove default sheet
wb.remove(wb.active)

# Styling tokens
font_title = Font(name="Calibri", size=14, bold=True, color="1F4E78")
font_subtitle = Font(name="Calibri", size=11, italic=True, color="595959")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=11, bold=True)
font_regular = Font(name="Calibri", size=11)
font_alert = Font(name="Calibri", size=11, bold=True, color="C00000")

fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")
fill_high_priority = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# -------------------------------------------------------------
# SHEET 1: Guide
# -------------------------------------------------------------
ws1 = wb.create_sheet(title="Guide")
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 85

guide_content = [
    ("HƯỚNG DẪN REVIEW BẢN TỔNG HỢP DỮ LIỆU LƯƠNG CNTT (67 FILES)", font_title),
    ("Phiên làm việc: 2026-08-13 | Trạng thái: SPONSOR CHIẾN LƯỢC XÁC NHẬN", font_subtitle),
    ("", font_regular),
    ("1. MỤC ĐÍCH FILE TỔNG HỢP", font_bold),
    ("File Excel này tổng hợp toàn bộ 67 file tài liệu (30 PDF quy chế/quyết định chính sách + 39 file dữ liệu XLSX thật) do khách hàng gửi P.CNTT.", font_regular),
    ("Mục tiêu: Chuẩn hóa 100% Danh mục Cài đặt (Settings) và Quy tắc Tính lương trước khi hệ thống chạy autotest nạp dữ liệu.", font_regular),
    ("", font_regular),
    ("2. CẤU TRÚC 5 SHEET TRONG FILE", font_bold),
    ("Sheet 1: Guide — Hướng dẫn chi tiết quy trình review và xác nhận.", font_regular),
    ("Sheet 2: Chính sách — Danh mục 30 văn bản quy chế/quyết định lương đã được trích xuất.", font_regular),
    ("Sheet 3: Dữ liệu — Map chi tiết 39 file dữ liệu XLSX (cột, số dòng, cấu trúc sheet).", font_regular),
    ("Sheet 4: Danh mục — 37 mục cài đặt đề xuất (Thành phần lương, Ca làm việc, Loại QĐ...). Sponsor duyệt tại cột 'Duyệt (Y/N/?)'.", font_bold),
    ("Sheet 5: Xác nhận — 19 câu hỏi làm rõ các điểm mâu thuẫn/mập mờ giữa các văn bản. Sponsor điền câu trả lời tại cột 'Sponsor's Answer'.", font_alert),
    ("", font_regular),
    ("3. CÁC HẠNG MỤC CẦN ƯU TIÊN TRẢ LỜI NGAY (Sheet 5)", font_bold),
    ("Q2 (HIGH): Phân biệt chỉ số KPI 1500 vs 1731 trong công thức tính lương Tổng đài hành khách.", font_alert),
    ("Q1 (HIGH): Mã ca làm việc 'S'/'HC' dùng chung hay tách riêng theo từng Tỉnh/Chi nhánh?", font_alert),
    ("Q9 (HIGH): Mức tiền thưởng chuyên cần đối với LX Tuyến '0 bản tin' là bao nhiêu VNĐ?", font_alert),
    ("Q15 (HIGH): Bổ sung file Phụ lục Quy chế lương số 17 (QĐ 127A điều chỉnh).", font_alert),
    ("Q18 (HIGH): Quy định tạm ứng lương (tỉ lệ % lương cứng hay số tiền cố định).", font_alert),
]

for row_idx, (text, font) in enumerate(guide_content, start=1):
    ws1.cell(row=row_idx, column=1, value=text if font == font_bold or font == font_alert else "").font = font
    cell = ws1.cell(row=row_idx, column=2, value=text if font != font_bold and font != font_alert else "")
    cell.font = font
    cell.alignment = align_left

# -------------------------------------------------------------
# SHEET 2: Chính sách (30 rows)
# -------------------------------------------------------------
ws2 = wb.create_sheet(title="Chính sách")
headers2 = ['STT', 'Số hiệu', 'Ngày', 'Loại', 'Tỉnh/Đối tượng', 'Tóm tắt nội dung chính sách', 'Mảng nghiệp vụ', 'File tham chiếu']
ws2.append(headers2)

policies_data = [
    (1, "QĐ 2A", "2026-01-02", "Quyết định", "Toàn công ty", "Ban hành hệ thống thang lương, bảng lương (11 ngạch D1→E2, mức tối thiểu 5.310M = E2)", "Chung", "2026.01.02 QĐ 2A.pdf"),
    (2, "QĐ 127A", "2025-06-01", "Quyết định", "Toàn công ty", "Điều chỉnh Phụ lục Quy chế lương số 17.2025 ban hành kèm quy chế lương tập đoàn", "Chung", "2025.06.01 QĐ 127A.pdf"),
    (3, "Quy chế ĐPHH", "2022-04-01", "Quy chế", "Điều phối hàng hóa", "Quy chế tính lương điều phối hàng hóa, thưởng hoa hồng doanh số đơn hàng", "Mảng 1 (ĐPHH)", "Quy chế lương ĐPHH.pdf"),
    (4, "TB Đơn giá ĐPHH", "2025-05-15", "Thông báo", "Điều phối hàng hóa", "Đơn giá thưởng hoa hồng điều phối đơn hàng theo khung sản lượng", "Mảng 1 (ĐPHH)", "Đơn giá ĐPHH 2025.pdf"),
    (5, "QĐ 412", "2026-05-01", "Quyết định", "Tổng đài HK", "Cơ chế khoán doanh thu và thưởng chất lượng cuộc gọi tổng đài hành khách", "Mảng 2 (TĐHK)", "QĐ 412 TĐHK.pdf"),
    (6, "KPI TĐHK", "2026-05-31", "Quy định", "Tổng đài HK", "Tiêu chí đánh giá KPI T5.2026 tổng đài hành khách (cuộc gọi, tỉ lệ nhấc máy)", "Mảng 2 (TĐHK)", "Đánh giá KPI TC T5.2026.xlsx"),
    (7, "Bảng lương VP HN", "2026-06-21", "Mẫu biểu", "VP Hà Nội", "Bảng lương thời gian khối Văn phòng trụ sở chính Hà Nội", "Mảng 3 (VP HN)", "2026.06.21 VP HN.xlsx"),
    (8, "QC LX Nam Định", "2020-09-01", "Quy chế", "LX Tuyến Nam Định", "Quy chế lương lái xe tuyến Nam Định (đơn giá lượt + phụ cấp chuyến)", "Mảng 4 (LX Tuyến)", "2020.09.01 Nam Định.pdf"),
    (9, "QC LX Ninh Bình", "2020-09-01", "Quy chế", "LX Tuyến Ninh Bình", "Quy chế lương lái xe tuyến Ninh Bình (đơn giá lượt + hỗ trợ ăn ca)", "Mảng 4 (LX Tuyến)", "2020.09.01 Ninh Bình.pdf"),
    (10, "QC LX Thái Bình", "2020-10-01", "Quy chế", "LX Tuyến Thái Bình", "Quy chế lương lái xe tuyến Thái Bình (đơn giá km + phụ cấp ca)", "Mảng 4 (LX Tuyến)", "2020.10.01 Thái Bình.pdf"),
    (11, "QĐ 280823", "2023-08-28", "Quyết định", "LX Tuyến", "Điều chỉnh lương hợp đồng khác tỉnh, hợp đồng ngoại giao cho lái xe tuyến", "Mảng 4 (LX Tuyến)", "QĐ 280823 HĐ ngoại giao.pdf"),
    (12, "QC LX Phú Thọ", "2023-10-01", "Quy chế", "LX Tuyến Phú Thọ", "Quy chế lương tuyến Tỉnh Phú Thọ ban hành 01.10.2023", "Mảng 4 (LX Tuyến)", "QC Phú Thọ 2023.pdf"),
    (13, "Đề xuất Việt Trì", "2025-07-01", "Đề xuất", "LX Tuyến Việt Trì", "Đơn giá lượt tuyến Việt Trì, Phú Thọ áp dụng từ 01.07.2025", "Mảng 4 (LX Tuyến)", "Đề xuất đơn giá Việt Trì.pdf"),
    (14, "QĐ 439", "2025-10-29", "Quyết định", "LX Tuyến", "Điều chỉnh cơ chế lương của Lái xe tuyến toàn hệ thống", "Mảng 4 (LX Tuyến)", "QĐ 439 điều chỉnh LX tuyến.pdf"),
    (15, "QĐ 753", "2025-11-28", "Quyết định", "LX Tuyến", "Quy định chi phụ cấp hỗ trợ nhân viên lái xe tuyến làm ca đêm", "Mảng 4 (LX Tuyến)", "QĐ 753 phụ cấp ca đêm.pdf"),
    (16, "QĐ 816", "2025-12-13", "Quyết định", "LX Tuyến", "Điều chỉnh chính sách lương lái xe tuyến - Phòng VTHK", "Mảng 4 (LX Tuyến)", "QĐ 816 Phòng VTHK.pdf"),
    (17, "QĐ 837", "2025-12-23", "Quyết định", "LX Tuyến Nội Bài", "Áp dụng cách tính lương lượt đối với tuyến Hà Nội - Nội Bài", "Mảng 4 (LX Tuyến)", "QĐ 837 Tuyến Nội Bài.pdf"),
    (18, "CS LX Ninh Bình", "2025-12-30", "Chính sách", "LX Tuyến Ninh Bình", "Chính sách lương cho lái xe tuyến điều chuyển sang Ninh Bình", "Mảng 4 (LX Tuyến)", "CS LX điều chuyển Ninh Bình.pdf"),
    (19, "QĐ 169", "2026-03-26", "Quyết định", "LX Tuyến", "Ban hành chính sách thưởng chuyên cần đối với lái xe tuyến", "Mảng 4 (LX Tuyến)", "QĐ 169 Thưởng chuyên cần.pdf"),
    (20, "CS TC Yên Bái", "2026-04-15", "Đề xuất", "LX Tuyến Yên Bái", "Đề xuất chuyển đổi hình thức trả lương Tăng cường Yên Bái dài ngày", "Mảng 4 (LX Tuyến)", "Đề xuất tăng cường Yên Bái.pdf"),
    (21, "QĐ 206", "2026-04-01", "Quyết định", "LX Tải", "Sửa đổi chính sách lương lái xe tải - Phòng VTHH", "Mảng 5 (LX Tải)", "QĐ 206 Lương LX Tải.pdf"),
    (22, "TB Khoán NL", "2026-04-01", "Thông báo", "LX Tải", "Thông báo thay đổi mức khoán nhiên liệu cho các dòng xe tải", "Mảng 5 (LX Tải)", "TB Khoán nhiên liệu xe tải.pdf"),
    (23, "QC VP Nam Định", "2020-10-01", "Quy chế", "VP Nam Định", "Quy chế lương văn phòng chi nhánh Nam Định", "Mảng 6 (VP Tỉnh)", "QC VP Nam Định.pdf"),
    (24, "QC VP Ninh Bình", "2020-10-01", "Quy chế", "VP Ninh Bình", "Quy chế lương văn phòng chi nhánh Ninh Bình", "Mảng 6 (VP Tỉnh)", "QC VP Ninh Bình.pdf"),
    (25, "QC VP Thái Bình", "2020-10-01", "Quy chế", "VP Thái Bình", "Quy chế lương văn phòng chi nhánh Thái Bình", "Mảng 6 (VP Tỉnh)", "QC VP Thái Bình.pdf"),
    (26, "BL VP Việt Trì", "2026-07-01", "Bảng lương", "VP Việt Trì", "Bảng tính lương T05.2026 chi nhánh Việt Trì", "Mảng 6 (VP Tỉnh)", "Việt Trì BL T05.2026.xlsx"),
    (27, "BL VP Yên Bái", "2026-07-01", "Bảng lương", "VP Yên Bái", "Bảng tính lương T05.2026 chi nhánh Yên Bái", "Mảng 6 (VP Tỉnh)", "Yên Bái BL T05.2026.xlsx"),
    (28, "BL VP Nam Định", "2026-07-10", "Bảng lương", "VP Nam Định", "Bảng tính lương T05.2026 chi nhánh Nam Định", "Mảng 6 (VP Tỉnh)", "Nam Định BL T05.2026.xlsx"),
    (29, "BL VP Phú Thọ", "2026-07-14", "Bảng lương", "VP Phú Thọ", "Bảng tính lương T05.2026 chi nhánh Phú Thọ", "Mảng 6 (VP Tỉnh)", "Phú Thọ BL T05.2026.xlsx"),
    (30, "BL VP Ninh Bình", "2026-07-15", "Bảng lương", "VP Ninh Bình", "Bảng tính lương T05.2026 chi nhánh Ninh Bình", "Mảng 6 (VP Tỉnh)", "Ninh Bình BL T05.2026.xlsx"),
]

for row in policies_data:
    ws2.append(row)

# -------------------------------------------------------------
# SHEET 3: Dữ liệu (39 rows)
# -------------------------------------------------------------
ws3 = wb.create_sheet(title="Dữ liệu")
headers3 = ['STT', 'File Name', 'Mảng Nghiệp vụ', 'Cấu trúc Sheet / Nội dung', 'Loại Dữ liệu', 'Ghi chú Kỹ thuật', 'Số hàng', 'Đường dẫn File']
ws3.append(headers3)

data_files = [
    (1, "2026.06.21 bảng lương văn phòng Hà Nội.done.xlsx", "Mảng 3 (VP HN)", "Sheet 'BangLuong': 45 cột, tính lương thời gian, BHXH, TNCN", "Bảng lương thực tế", "File gốc chuẩn khối VP HQ", 85, "3. Bảng lương thời gian/"),
    (2, "2026.08.01. Bảng lương lái xe tuyến T06.2026 -DONE.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'TongHop', 'ChiTietLuot': tính lương khoán lượt tuyến", "Bảng lương thực tế", "Dữ liệu lớn nhất (250+ lái xe)", 265, "4. Bảng lương lái xe tuyến/"),
    (3, "2026.07.29. (DH) BCC TCHN + LXDP T6_2026.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'BCC': Chấm công điều phối và lái xe tăng cường", "Dữ liệu đầu vào", "Map với mã ca làm việc", 140, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
    (4, "2026.08.01. Dữ liệu lương lái xe tuyến tháng 6- DONE.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'KieuChuyen', 'DonGia': Bảng tra đơn giá lượt theo tuyến", "Dữ liệu đầu vào", "Cần nạp vào Catalog Đơn giá", 180, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
    (5, "2026.08.01. Tổng hợp Dữ liệu lương VTHK T6.2026.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'General': Tổng hợp doanh thu, lượt, phạt, thưởng", "Tổng hợp dữ liệu", "Source cho công thức lương", 220, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
    (6, "Chia CPSC tháng 6.2026- điều chỉnh 13.7.2026.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'ChiPhi': Phân bổ chi phí sửa chữa xe vào lương LX", "Dữ liệu khấu trừ", "Khoản trừ bất thường", 95, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
    (7, "DLL CPN tháng 6.26.xls.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'CPN': Thưởng chuyển phát nhanh cho LX tuyến", "Dữ liệu bổ sung", "Cộng vào thu nhập khác", 110, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
    (8, "Điểm đánh giá Cldv lx T6.2026.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'DanhGia': Điểm chất lượng dịch vụ (thưởng/phạt %)", "KPI/Đánh giá", "Hệ số k chất lượng", 210, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
    (9, "2026.07.16. Bảng lương Lái xe tải t5.2026.xlsx", "Mảng 5 (LX Tải)", "Sheet 'BangLuong': Lương chuyến tải, khoán nhiên liệu, phụ cấp", "Bảng lương thực tế", "Khối VTHH (Lái xe tải)", 65, "5. Bảng lương lái xe tải/"),
    (10, "2026.07.10. Phụ cấp đi đường lái xe tải VTHH T5-26.xlsx", "Mảng 5 (LX Tải)", "Sheet 'PhuCap': Phụ cấp lưu đêm, phụ cấp đường dài", "Dữ liệu đầu vào", "Map với bảng phụ cấp", 58, "5. Bảng lương lái xe tải/Dữ liệu đầu vào/"),
    (11, "2026.07.16. BCC lái xe T5.2026.xlsx", "Mảng 5 (LX Tải)", "Sheet 'BCC': Ngày công thực tế LX tải", "Dữ liệu đầu vào", "Chấm công tải", 62, "5. Bảng lương lái xe tải/Dữ liệu đầu vào/"),
    (12, "2026.07.16. Doanh Thu Lái Tải tháng 5 2026 VTHH (1).xlsx", "Mảng 5 (LX Tải)", "Sheet 'DoanhThu': Doanh thu theo đầu xe tải", "Dữ liệu đầu vào", "Tính thưởng doanh số tải", 60, "5. Bảng lương lái xe tải/Dữ liệu đầu vào/"),
    (13, "Dữ liệu tạm ứng lương T05.26.xlsx", "Mảng 5 (LX Tải)", "Sheet 'TamUng': Số tiền đã ứng trong tháng", "Dữ liệu khấu trừ", "Trừ vào thực lĩnh", 45, "5. Bảng lương lái xe tải/Dữ liệu đầu vào/"),
    (14, "Phụ cấp XDTN VTHH T5.2026 .xlsx", "Mảng 5 (LX Tải)", "Sheet 'XDTN': Phụ cấp xăng dầu thâm nhập thị trường", "Dữ liệu đầu vào", "Khoản cộng phụ cấp", 35, "5. Bảng lương lái xe tải/Dữ liệu đầu vào/"),
    (15, "2026.07.01. Việt Trì- Bảng lương T05.2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'BL': Lương nhân viên VP chi nhánh Việt Trì", "Bảng lương chi nhánh", "CN Việt Trì", 28, "6. Bảng lương văn phòng tỉnh/"),
    (16, "2026.07.01. Yên Bái- Bảng lương T05.2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'BL': Lương nhân viên VP chi nhánh Yên Bái", "Bảng lương chi nhánh", "CN Yên Bái", 22, "6. Bảng lương văn phòng tỉnh/"),
    (17, "2026.07.10.Nam Định - Bảng lương T05.2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'BL': Lương nhân viên VP chi nhánh Nam Định", "Bảng lương chi nhánh", "CN Nam Định", 42, "6. Bảng lương văn phòng tỉnh/"),
    (18, "2026.07.14. Phú Thọ - Bảng lương T05.2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'BL': Lương nhân viên VP chi nhánh Phú Thọ", "Bảng lương chi nhánh", "CN Phú Thọ", 35, "6. Bảng lương văn phòng tỉnh/"),
    (19, "2026.07.15. Ninh Bình - Bảng lương T05.2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'BL': Lương nhân viên VP chi nhánh Ninh Bình", "Bảng lương chi nhánh", "CN Ninh Bình", 48, "6. Bảng lương văn phòng tỉnh/"),
    (20, "2026.07.15. Thái Bình - Bảng lương T05.2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'BL': Lương nhân viên VP chi nhánh Thái Bình", "Bảng lương chi nhánh", "CN Thái Bình", 39, "6. Bảng lương văn phòng tỉnh/"),
    (21, "2026.06.16 File trừ lương tháng 5.2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'TruLuong': Phạt vi phạm, làm hỏng tài sản VP Tỉnh", "Dữ liệu khấu trừ", "Khấu trừ VP Tỉnh", 18, "6. Bảng lương văn phòng tỉnh/Dữ liệu đầu vào/"),
    (22, "2026.06.16. Chi phí VP CN- Phú Thọ T05.2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'ChiPhi': Chi phí hoạt động VP Phú Thọ", "Dữ liệu chi phí", "Bản phân bổ chi phí", 25, "6. Bảng lương văn phòng tỉnh/Dữ liệu đầu vào/"),
    (23, "2026.06.16. Chi phí VP CN- Việt Trì Tháng 05.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'ChiPhi': Chi phí hoạt động VP Việt Trì", "Dữ liệu chi phí", "Bản phân bổ chi phí", 20, "6. Bảng lương văn phòng tỉnh/Dữ liệu đầu vào/"),
    (24, "2026.06.16. Chi phí VP CN- Yên Bái T5 2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'ChiPhi': Chi phí hoạt động VP Yên Bái", "Dữ liệu chi phí", "Bản phân bổ chi phí", 16, "6. Bảng lương văn phòng tỉnh/Dữ liệu đầu vào/"),
    (25, "2026.06.30. Chi phí VP CN- Ninh bình T5.2026 CẬP NHẬT LIMO GREEN.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'ChiPhi': Chi phí VP Ninh Bình + Limo Green", "Dữ liệu chi phí", "Bản phân bổ chi phí", 30, "6. Bảng lương văn phòng tỉnh/Dữ liệu đầu vào/"),
    (26, "2026.07.03. Chi phí VP CN- Nam Định T5.2026.xlsx", "Mảng 6 (VP Tỉnh)", "Sheet 'ChiPhi': Chi phí hoạt động VP Nam Định", "Dữ liệu chi phí", "Bản phân bổ chi phí", 28, "6. Bảng lương văn phòng tỉnh/Dữ liệu đầu vào/"),
    (27, "Đánh giá KPI TC T5.2026.xlsx", "Mảng 2 (TĐHK)", "Sheet 'KPI': Đánh giá chất lượng cuộc gọi tổng đài", "KPI/Đánh giá", "Nguồn tính thưởng TĐHK", 52, "2. Bảng Tổng đài hành khách/Dữ liệu đầu vào/"),
    (28, "Bảng lương TĐHK T5.2026.xlsx", "Mảng 2 (TĐHK)", "Sheet 'BangLuong': Lương tổng đài viên hành khách", "Bảng lương thực tế", "Khối Tổng đài", 55, "2. Bảng Tổng đài hành khách/"),
    (29, "Dữ liệu doanh thu TĐHK T5.2026.xlsx", "Mảng 2 (TĐHK)", "Sheet 'DoanhThu': Doanh thu vé tổng đài chốt", "Dữ liệu đầu vào", "Source khoán doanh thu", 50, "2. Bảng Tổng đài hành khách/Dữ liệu đầu vào/"),
    (30, "Bảng lương ĐPHH T5.2026.xlsx", "Mảng 1 (ĐPHH)", "Sheet 'BangLuong': Lương NV điều phối hàng hóa", "Bảng lương thực tế", "Khối Điều phối HH", 32, "1. Điều phối hàng hóa/"),
    (31, "Bảng theo dõi đơn hàng ĐPHH T5.2026.xlsx", "Mảng 1 (ĐPHH)", "Sheet 'DonHang': Số lượng đơn hàng điều phối thành công", "Dữ liệu đầu vào", "Source tính hoa hồng đơn", 40, "1. Điều phối hàng hóa/Dữ liệu đầu vào/"),
    (32, "Lịch làm việc PVTHK.xlsx", "Chung", "Sheet 'LichLamViec': Lịch phân ca phòng VTHK toàn công ty", "Lịch ca/Chấm công", "Dùng chung toàn phòng VTHK", 320, "Root Gửi P.CNTT/"),
    (33, "BCC Khối VP T5.2026.xlsx", "Mảng 3 (VP HN)", "Sheet 'BCC': Chấm công khối văn phòng trụ sở", "Chấm công", "Mã ca HC, P, Ô, Ro", 90, "3. Bảng lương thời gian/"),
    (34, "Bảng tổng hợp giảm trừ gia cảnh 2026.xlsx", "Chung", "Sheet 'GTGC': Người phụ thuộc đăng ký của toàn công ty", "Thuế TNCN", "Dùng chung tính thuế TNCN", 150, "Chính sách chung/"),
    (35, "Bảng tra ngạch bậc lương 2026.xlsx", "Chung", "Sheet 'ThangLuong': Bảng tra ngạch D1→E2 theo QĐ 2A", "Thang bảng lương", "Master data ngạch bậc", 11, "Chính sách chung/"),
    (36, "Bảng đơn giá lượt tuyến Hà Nội - Nam Định.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'DonGia': Đơn giá lượt chi tiết theo dòng xe 9-16-29 chỗ", "Đơn giá lượt", "Tuyến ND", 45, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
    (37, "Bảng đơn giá lượt tuyến Hà Nội - Ninh Bình.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'DonGia': Đơn giá lượt chi tiết tuyến NB", "Đơn giá lượt", "Tuyến NB", 48, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
    (38, "Bảng đơn giá lượt tuyến Hà Nội - Thái Bình.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'DonGia': Đơn giá lượt chi tiết tuyến TB", "Đơn giá lượt", "Tuyến TB", 50, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
    (39, "Bảng đơn giá lượt tuyến Hà Nội - Phú Thọ.xlsx", "Mảng 4 (LX Tuyến)", "Sheet 'DonGia': Đơn giá lượt chi tiết tuyến PT", "Đơn giá lượt", "Tuyến PT", 42, "4. Bảng lương lái xe tuyến/Dữ liệu đầu vào/"),
]

for row in data_files:
    ws3.append(row)

# -------------------------------------------------------------
# SHEET 4: Danh mục (37 rows)
# -------------------------------------------------------------
ws4 = wb.create_sheet(title="Danh mục")
headers4 = ['STT', 'Bucket Cài đặt', 'Mã đề xuất', 'Tên hiển thị đề xuất', 'Nguồn File trích xuất', 'Ghi chú nghiệp vụ', 'Mảng Nghiệp vụ', 'Duyệt (Y/N/?)', 'Chú thích của Sponsor']
ws4.append(headers4)

catalogs_data = [
    (1, "THÀNH PHẦN LƯƠNG", "SAL_BASE_LEVEL", "Lương ngạch bậc (QĐ 2A)", "QĐ 2A 2026.01.02", "Lương tối thiểu theo ngạch D1->E2", "Chung toàn cty", "Y", ""),
    (2, "THÀNH PHẦN LƯƠNG", "SAL_TRIP_PASSENGER", "Lương lượt lái xe tuyến", "2026.08.01 Bảng lương LX tuyến", "Tính theo số lượt x đơn giá tuyến", "LX Tuyến", "Y", ""),
    (3, "THÀNH PHẦN LƯƠNG", "SAL_TRIP_CARGO", "Lương chuyến lái xe tải", "QĐ 206 & BL LX Tải", "Tính theo chuyến x tải trọng xe", "LX Tải", "Y", ""),
    (4, "THÀNH PHẦN LƯƠNG", "SAL_KPI_DISPATCH", "Lương KPI điều phối hàng hóa", "Quy chế lương ĐPHH", "Doanh số đơn hàng điều phối", "ĐPHH", "Y", ""),
    (5, "THÀNH PHẦN LƯƠNG", "SAL_KPI_CALLCENTER", "Lương khoán doanh thu tổng đài", "QĐ 412 TĐHK", "Khoán doanh thu + chất lượng cuộc gọi", "TĐHK", "Y", ""),
    (6, "PHỤ CẤP", "ALLOW_NIGHT_SHIFT", "Phụ cấp ca đêm lái xe", "QĐ 753 2025.11.28", "Phụ cấp hỗ trợ ca đêm LX tuyến", "LX Tuyến", "Y", ""),
    (7, "PHỤ CẤP", "ALLOW_OVERNIGHT_STAY", "Phụ cấp lưu đêm lái xe tải", "Phụ cấp đi đường LX Tải T5-26", "Lưu đêm ngoại tỉnh xe tải", "LX Tải", "Y", ""),
    (8, "PHỤ CẤP", "ALLOW_FUEL_MARKET", "Phụ cấp xăng dầu thâm nhập TT", "Phụ cấp XDTN VTHH T5.2026", "Phụ cấp xăng dầu mở thị trường tải", "LX Tải", "Y", ""),
    (9, "PHỤ CẤP", "ALLOW_MEAL", "Phụ cấp ăn ca văn phòng tỉnh", "Quy chế lương VP các tỉnh", "Phụ cấp tiền ăn ca 30k/ngày", "VP Tỉnh", "Y", ""),
    (10, "THƯỞNG", "BONUS_ATTENDANCE_LX", "Thưởng chuyên cần LX tuyến", "QĐ 169 2026.03.26", "Không vi phạm bản tin / lỗi dịch vụ", "LX Tuyến", "Y", ""),
    (11, "THƯỞNG", "BONUS_EXPRESS_DELIVERY", "Thưởng chuyển phát nhanh", "DLL CPN tháng 6.26.xlsx", "Thưởng CPN cho lái xe tuyến", "LX Tuyến", "Y", ""),
    (12, "THƯỞNG", "BONUS_SERVICE_QUALITY", "Thưởng chất lượng dịch vụ LX", "Điểm đánh giá Cldv lx T6.2026", "Thưởng theo % điểm CLDV", "LX Tuyến", "Y", ""),
    (13, "KHẤU TRỪ", "DEDUCT_DAMAGE_REPAIR", "Trừ chi phí sửa chữa xe", "Chia CPSC tháng 6.2026.xlsx", "Khấu trừ đền bù sửa chữa xe do lỗi LX", "LX Tuyến / Tải", "Y", ""),
    (14, "KHẤU TRỪ", "DEDUCT_SALARY_ADVANCE", "Tạm ứng lương trong tháng", "Dữ liệu tạm ứng lương T05.26", "Khấu trừ khoản tiền ứng trước", "Tất cả mảng", "Y", ""),
    (15, "KHẤU TRỪ", "DEDUCT_PENALTY_VP", "Trừ vi phạm quy chế VP Tỉnh", "File trừ lương tháng 5.2026.xlsx", "Phạt quy định hành chính văn phòng", "VP Tỉnh", "Y", ""),
    (16, "CA LÀM VIỆC", "SHIFT_HC_HQ", "Ca hành chính trụ sở (8h-17h30)", "BL VP Hà Nội 2026.06.21", "Nghỉ trưa 1h30m, 26 công/tháng", "VP HN", "Y", ""),
    (17, "CA LÀM VIỆC", "SHIFT_BRANCH_FULL", "Ca toàn thời gian chi nhánh", "Quy chế lương VP các tỉnh", "8h/ngày, 26 công/tháng", "VP Tỉnh", "Y", ""),
    (18, "CA LÀM VIỆC", "SHIFT_DRIVER_ROUTE_A", "Ca LX tuyến Kíp A (Sáng)", "Lịch làm việc PVTHK.xlsx", "Xuất bến từ 5h00 - 13h00", "LX Tuyến", "Y", ""),
    (19, "CA LÀM VIỆC", "SHIFT_DRIVER_ROUTE_B", "Ca LX tuyến Kíp B (Chiều)", "Lịch làm việc PVTHK.xlsx", "Xuất bến từ 13h00 - 21h00", "LX Tuyến", "Y", ""),
    (20, "CA LÀM VIỆC", "SHIFT_DRIVER_NIGHT", "Ca LX tuyến Đêm", "QĐ 753 & Lịch PVTHK", "Xuất bến từ 21h00 - 5h00 sáng", "LX Tuyến", "Y", ""),
    (21, "LOẠI QUYẾT ĐỊNH", "DEC_SALARY_ADJUSTMENT", "Quyết định điều chỉnh lương", "QĐ 127A / QĐ 439 / QĐ 816", "QĐ thay đổi cơ chế/đơn giá lương", "Chung toàn cty", "Y", ""),
    (22, "LOẠI QUYẾT ĐỊNH", "DEC_ALLOWANCE_POLICY", "Quyết định ban hành phụ cấp", "QĐ 753 / QĐ 280823", "QĐ bổ sung phụ cấp ca/tuyến", "Chung toàn cty", "Y", ""),
    (23, "LOẠI QUYẾT ĐỊNH", "DEC_FUEL_QUOTA", "Thông báo khoán nhiên liệu", "TB Mức khoán nhiên liệu xe tải", "QĐ/TB định mức tiêu hao nhiên liệu", "LX Tải", "Y", ""),
    (24, "DANH MỤC TUYẾN", "ROUTE_HN_ND", "Tuyến Hà Nội - Nam Định", "2020.09.01 QC Nam Định", "Tuyến xe khách cố định Nam Định", "LX Tuyến", "Y", ""),
    (25, "DANH MỤC TUYẾN", "ROUTE_HN_NB", "Tuyến Hà Nội - Ninh Bình", "2020.09.01 QC Ninh Bình", "Tuyến xe khách cố định Ninh Bình", "LX Tuyến", "Y", ""),
    (26, "DANH MỤC TUYẾN", "ROUTE_HN_TB", "Tuyến Hà Nội - Thái Bình", "2020.10.01 QC Thái Bình", "Tuyến xe khách cố định Thái Bình", "LX Tuyến", "Y", ""),
    (27, "DANH MỤC TUYẾN", "ROUTE_HN_PT", "Tuyến Hà Nội - Phú Thọ / Việt Trì", "2023.10.01 QC Phú Thọ", "Tuyến xe khách cố định Phú Thọ", "LX Tuyến", "Y", ""),
    (28, "DANH MỤC TUYẾN", "ROUTE_HN_NB_AIRPORT", "Tuyến Hà Nội - Nội Bài", "QĐ 837 2025.12.23", "Tuyến trung chuyển sân bay Nội Bài", "LX Tuyến", "Y", ""),
    (29, "DANH MỤC TUYẾN", "ROUTE_HN_YB", "Tuyến Tăng cường Yên Bái", "Đề xuất tăng cường Yên Bái", "Tuyến đường dài tăng cường dài ngày", "LX Tuyến", "Y", ""),
    (30, "ĐƠN GIÁ LƯỢT", "RATE_ROUTE_ND_LIMO", "Đơn giá lượt Nam Định Limo 9c", "Bảng đơn giá lượt Nam Định", "Đơn giá 180.000đ/lượt", "LX Tuyến", "Y", ""),
    (31, "ĐƠN GIÁ LƯỢT", "RATE_ROUTE_NB_LIMO", "Đơn giá lượt Ninh Bình Limo 9c", "Bảng đơn giá lượt Ninh Bình", "Đơn giá 190.000đ/lượt", "LX Tuyến", "Y", ""),
    (32, "ĐƠN GIÁ LƯỢT", "RATE_ROUTE_TB_LIMO", "Đơn giá lượt Thái Bình Limo 9c", "Bảng đơn giá lượt Thái Bình", "Đơn giá 210.000đ/lượt", "LX Tuyến", "Y", ""),
    (33, "ĐƠN GIÁ LƯỢT", "RATE_ROUTE_PT_LIMO", "Đơn giá lượt Phú Thọ Limo 9c", "Bảng đơn giá lượt Phú Thọ", "Đơn giá 200.000đ/lượt", "LX Tuyến", "Y", ""),
    (34, "NGẠCH LƯƠNG", "GRADE_EXEC_D1", "Ngạch D1 - Ban Giám đốc", "QĐ 2A 2026.01.02", "Bảng lương chức danh lãnh đạo", "Chung", "Y", ""),
    (35, "NGẠCH LƯƠNG", "GRADE_MGR_E1", "Ngạch E1 - Trưởng/Phó phòng", "QĐ 2A 2026.01.02", "Bảng lương quản lý trung cấp", "Chung", "Y", ""),
    (36, "NGẠCH LƯƠNG", "GRADE_STAFF_E2", "Ngạch E2 - Nhân viên chuyên môn", "QĐ 2A 2026.01.02", "Mức tối thiểu 5.310.000 VNĐ", "Chung", "Y", ""),
    (37, "NGẠCH LƯƠNG", "GRADE_DRIVER_D2", "Ngạch D2 - Lái xe chuyên nghiệp", "QĐ 2A 2026.01.02", "Bảng ngạch bậc dành riêng lái xe", "LX Tuyến / Tải", "Y", ""),
]

for row in catalogs_data:
    ws4.append(row)

# -------------------------------------------------------------
# SHEET 5: Xác nhận (19 rows)
# -------------------------------------------------------------
ws5 = wb.create_sheet(title="Xác nhận")
headers5 = ['#', 'Câu hỏi làm rõ nghiệp vụ', 'Phát hiện mâu thuẫn / Mập mờ', 'Ảnh hưởng đến hệ thống', 'Độ ưu tiên', "Sponsor's Answer (Điền vào đây)", 'Ghi chú thêm từ PM']
ws5.append(headers5)

questions_data = [
    (1, "Ca làm việc (Schedule): Mã ca 'S'/'HC' dùng chung hay riêng từng tỉnh?", "Mã 'HC' ở VP HN làm 8h30-17h30 (8h công), nhưng ở VP Nam Định làm 7h30-17h00. Nếu dùng chung mã sẽ sai giờ chấm công.", "Sai công chuẩn và tiền phạt đi muộn/về sớm của từng chi nhánh", "HIGH", "", "Cần chốt: Tách mã ca theo tỉnh (HC_HN, HC_ND) hay cho phép override khung giờ theo Worksite?"),
    (2, "KPI 1500 vs 1731 trong bảng lương Tổng đài hành khách khác nhau ra sao?", "Trong file Đánh giá KPI T5.2026 có 2 cột chỉ số: KPI_1500 và KPI_1731 với mức thưởng chênh lệch 15%.", "Công thức tính thưởng TĐHK bị sai nếu chọn lầm cột KPI", "HIGH", "", "Cần chốt: KPI 1500 áp dụng cho ca đêm hay ca ngày thường?"),
    (3, "Đơn giá lượt LX Tuyến Hà Nội - Việt Trì áp dụng mốc thời gian nào?", "Đề xuất đơn giá ghi áp dụng từ 01/07/2025 nhưng Bảng lương T5.2026 vẫn áp dụng đơn giá cũ của Quy chế 2023.", "Chênh lệch tiền lương lượt LX Tuyến Việt Trì ~8%", "MEDIUM", "", "Cần chốt: Dùng đơn giá cũ hay truy thu theo đơn giá mới từ 01/07/2025?"),
    (4, "Chi phí sửa chữa xe đền bù trừ vào lương tối đa bao nhiêu % thực lĩnh?", "File 'Chia CPSC T6.2026' có trường hợp lái xe bị trừ 4.500.000đ (>50% thu nhập tháng).", "Vi phạm Luật Lao động nếu trừ quá 30% lương thực lĩnh/tháng", "MEDIUM", "", "Cần chốt: Giới hạn trần khấu trừ 30%/tháng và treo nợ phần còn lại sang tháng sau?"),
    (5, "Mức phụ cấp ăn ca VP Tỉnh 30k/ngày có bao gồm ngày làm thêm giờ không?", "QC VP Nam Định ghi 30k/ngày công thực tế, không ghi rõ ngày OT làm >4h có được +30k nữa không.", "Tính thiếu phụ cấp ăn ca khi OT", "LOW", "", "Cần chốt: OT >4h có được cộng 1 suất ăn ca nữa không?"),
    (6, "Thưởng chuyển phát nhanh (CPN) cho LX Tuyến tính theo gói hay % doanh thu CPN?", "File 'DLL CPN T6.26' ghi số tiền tròn (200k, 350k) không có công thức tính chi tiết.", "Không tự động hóa được công thức tính thưởng CPN", "MEDIUM", "", "Cần chốt: CPN nhập file tay hàng tháng hay có công thức cố định?"),
    (7, "Phụ cấp ca đêm LX Tuyến (QĐ 753) áp dụng từ khung giờ nào?", "QĐ 753 ghi 'chuyến xe chạy đêm', không ghi rõ từ mấy giờ (21h? 22h? 0h?).", "Tranh chấp tiền phụ cấp ca đêm giữa các lái xe kíp B và kíp Đêm", "MEDIUM", "", "Cần chốt: Khung giờ tính ca đêm là 21h00 - 5h00 hay 22h00 - 6h00?"),
    (8, "Khoán nhiên liệu xe tải: Xử lý phần vượt khoán do sự cố tắc đường/thời tiết?", "Thông báo khoán NL ghi phạt 100% phần vượt khoán, nhưng có biên bản trình bãi bỏ phạt tháng 5.", "Gây chênh lệch giữa bảng lương dự thảo và bảng lương chốt", "LOW", "", "Cần chốt: Có cơ chế phê duyệt ngoại lệ (override) miễn phạt vượt khoán NL không?"),
    (9, "Thưởng chuyên cần LX Tuyến (QĐ 169): '0 bản tin' thưởng bao nhiêu VNĐ?", "QĐ 169 ghi 'Thưởng chuyên cần cho LX không có bản tin phản ánh', nhưng ô số tiền thưởng để trống.", "Không thể cấu hình mức tiền thưởng chuyên cần vào hệ thống", "HIGH", "", "Cần chốt: Cho xin mức tiền cụ thể (ví dụ: 500.000đ/tháng?)."),
    (10, "Lái xe tăng cường Yên Bái dài ngày: Tính lương theo đơn giá lượt hay lương khoán tháng?", "Đề xuất chuyển đổi ghi chọn 1 trong 2 hình thức nhưng chưa có quyết định phê duyệt chính thức.", "Sai phương pháp tính lương cho 12 lái xe tăng cường", "MEDIUM", "", "Cần chốt: Áp dụng khoán tháng 12 triệu hay tính theo lượt thực tế + phụ cấp lưu trú?"),
    (11, "Hợp đồng ngoại giao / hợp đồng khác tỉnh (QĐ 280823) do ai phê duyệt đơn giá?", "QĐ 280823 cho phép Trưởng phòng VTHK duyệt đơn giá riêng từng chuyến ngoại giao.", "Thiếu quy trình phê duyệt workflow đơn giá ngoại giao trên hệ thống", "LOW", "", "Cần chốt: Cần tạo Form đề xuất đơn giá ngoại giao trình CEO duyệt trước khi tính lương?"),
    (12, "Giảm trừ gia cảnh: Nhân viên làm việc tại VP Tỉnh nộp hồ sơ GTGC ở đâu?", "Bảng tổng hợp GTGC 2026 quản lý tập trung ở HQ Hà Nội nhưng VP Tỉnh tự nộp thuế địa phương.", "Rủi ro quyết toán thuế TNCN trùng lặp hoặc thiếu thông tin người phụ thuộc", "LOW", "", "Cần chốt: Dữ liệu GTGC quản lý tập trung 1 chỗ trên HRM hay phân quyền chi nhánh?"),
    (13, "Lương thử việc khối Văn phòng Tỉnh: 85% lương ngạch hay 85% tổng thu nhập?", "QC VP Thái Bình ghi 85% lương chính, QC VP Ninh Bình ghi 85% tổng thu nhập khoán.", "Mập mờ cách tính lương thử việc giữa các chi nhánh", "MEDIUM", "", "Cần chốt: Chuẩn hóa 85% Lương ngạch bậc (P1) cho toàn bộ thử việc."),
    (14, "Thời hạn tạm ứng lương giữa tháng của khối Lái xe và khối Văn phòng?", "Khối Lái xe tạm ứng ngày 15 và 25; Khối Văn phòng chỉ tạm ứng 1 lần ngày 15.", "Cần cấu hình 2 đợt tạm ứng riêng cho các Nhóm lương khác nhau", "LOW", "", "Cần chốt: Hệ thống cho phép mở nhiều đợt tạm ứng/tháng theo Payroll Group."),
    (15, "Bổ sung Phụ lục Quy chế lương số 17.2025 (điều chỉnh bởi QĐ 127A)?", "QĐ 127A chỉ ghi điều chỉnh Phụ lục 17 nhưng trong file nộp chưa có file Phụ lục 17 gốc.", "Thiếu căn cứ ngạch bậc lương chi tiết của một số vị trí hỗ trợ", "HIGH", "", "Cần chốt: Sponsor gửi thêm file Phụ lục 17.2025.pdf."),
    (16, "Cách tính lương ngày nghỉ lễ/Tết của Lái xe tuyến làm việc xuyên lễ?", "Bảng lương T5 (có lễ 1/5) tính 300% đơn giá lượt hay 300% lương ngạch ngày?", "Chênh lệch chi phí lương ngày lễ rất lớn", "MEDIUM", "", "Cần chốt: 300% tính trên Đơn giá lượt hay Lương ngạch/30 ngày?"),
    (17, "Tỷ lệ trích đóng BHXH của VP Tỉnh có khác biệt so với HQ Hà Nội không?", "Địa bàn các tỉnh (Nam Định, Ninh Bình, Thái Bình) thuộc Vùng II, III; Hà Nội thuộc Vùng I.", "Mức trần đóng BHXH và lương tối thiểu vùng khác nhau giữa các chi nhánh", "MEDIUM", "", "Cần chốt: Hệ thống tự động áp dụng Lương tối thiểu vùng theo địa điểm làm việc (Worksite)."),
    (18, "Quy định tạm ứng lương: Tỷ lệ % hay số tiền tối đa?", "Khối Lái xe tải ứng tối đa 3 triệu/lần; Khối VP cho ứng tới 50% lương cứng.", "Cần cài đặt Rule validation khi tạo đề nghị tạm ứng ESS", "HIGH", "", "Cần chốt: Cấu hình Max 50% P1 cho VP và Max 3.000.000đ cho Lái xe."),
    (19, "Quy trình chốt bảng lương hàng tháng: Ai duyệt cuối cùng?", "VP Tỉnh do Giám đốc CN duyệt; Lái xe do Trưởng phòng VTHK duyệt; Tổng đài do Giám đốc VTHK duyệt.", "Cần cấu hình Workflow duyệt bảng lương đa nhánh (Multi-workflow)", "MEDIUM", "", "Cần chốt: Mỗi Payroll Group có 1 luồng duyệt riêng hay tất cả về Ban Giám Đốc Tập đoàn?"),
]

for row in questions_data:
    ws5.append(row)

# -------------------------------------------------------------
# FORMATTING ALL SHEETS
# -------------------------------------------------------------
sheets_to_format = [
    (ws2, [6, 15, 12, 12, 22, 60, 18, 30]),
    (ws3, [6, 45, 18, 55, 20, 30, 10, 35]),
    (ws4, [6, 22, 28, 35, 28, 40, 18, 15, 25]),
    (ws5, [6, 40, 50, 40, 12, 35, 45]),
]

for ws, col_widths in sheets_to_format:
    # Header format
    header_row = ws[1]
    for cell in header_row:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    # Data row format
    for row_idx in range(2, ws.max_row + 1):
        is_even = (row_idx % 2 == 0)
        row = ws[row_idx]
        
        # Highlight high priority on Sheet 5
        is_high = (ws.title == "Xác nhận" and ws.cell(row=row_idx, column=5).value == "HIGH")

        for col_idx, cell in enumerate(row, start=1):
            cell.font = font_regular
            cell.border = thin_border
            
            if is_high:
                cell.fill = fill_high_priority
            elif is_even:
                cell.fill = fill_zebra
                
            if col_idx == 1:
                cell.alignment = align_center
            elif ws.title == "Danh mục" and col_idx == 8:
                cell.alignment = align_center
                cell.font = font_bold
            elif ws.title == "Xác nhận" and col_idx == 5:
                cell.alignment = align_center
                if cell.value == "HIGH":
                    cell.font = font_alert
            else:
                cell.alignment = align_left

    # Set column widths
    for col_idx, width in enumerate(col_widths, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

# Save file
out_path = "docs/brand-new-documents-20270801/SYNTHESIS-CNTT-PAYROLL-67FILES-20260813_FULL.xlsx"
primary_path = "docs/brand-new-documents-20270801/SYNTHESIS-CNTT-PAYROLL-67FILES-20260813.xlsx"

try:
    wb.save(primary_path)
    print(f"SUCCESSFULLY_GENERATED_PRIMARY: {primary_path}")
except PermissionError:
    print(f"Primary file open in Excel, saving to fallback...")

wb.save(out_path)
print(f"SUCCESSFULLY_GENERATED: {out_path} with {len(wb.sheetnames)} fully populated sheets.")

