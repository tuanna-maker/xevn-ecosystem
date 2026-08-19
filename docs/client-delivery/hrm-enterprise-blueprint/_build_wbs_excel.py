# -*- coding: utf-8 -*-
"""Sinh file Excel WBS + nghiệp vụ gửi khách (tiếng Việt, cột khách trước)."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from _dien_bien_steps import build_all_steps, summary_for_uc

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "WBS_HRM_ENTERPRISE_KHACH_MOI.xlsx"
OUT_FALLBACK = ROOT / "WBS_HRM_ENTERPRISE_KHACH.xlsx"

MODULE_VI = {
    "REC": "Tuyển dụng",
    "CORE": "Nhân sự",
    "ATT": "Chấm công & Nghỉ phép",
    "PAY": "Tiền lương & Phúc lợi",
}

# WBS task catalog (customer-facing) — khớp SYNTHESIS_MASTER + WBS_HRM_ENTERPRISE.md v0.4 + SRS 0.7
WBS_TASKS: list[dict[str, str]] = [
    # REC — MVP: JD + YCTD + ứng viên gắn YCTD + báo cáo KH vs thực tế; chiến dịch = giai đoạn 2
    {"wbs": "WBS-REC-00", "mod": "REC", "ten": "Thư viện mô tả công việc (bản gốc chuẩn)", "uc": "UC-BP-REC-00", "muc_dich": "Lưu và tái sử dụng mô tả công việc chuẩn theo vị trí — làm đầu vào khi lập yêu cầu tuyển, không gõ lại mỗi lần.", "ppt": "4", "req": "REQ_REC_003", "q": "", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-REC-01", "mod": "REC", "ten": "Quản trị định biên 12 tháng (chỉ số cần tuyển)", "uc": "UC-BP-REC-01", "muc_dich": "Phòng ban trình–duyệt; hành chính nhân sự tổng hợp. Lưới tháng chỉ theo dõi số cần tuyển — biết khi nào cần có người làm.", "ppt": "4", "req": "REQ_REC_003; REQ_REC_005", "q": "Q-REC-HEADCOUNT", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-REC-01b", "mod": "REC", "ten": "Tự tạo yêu cầu tuyển theo tháng «Cần tuyển»", "uc": "UC-BP-REC-01b", "muc_dich": "Sau khi định biên được duyệt, hệ thống tự tạo yêu cầu tuyển đúng tháng cần người — không quên, không trùng; gắn mô tả công việc chuẩn khi đã có.", "ppt": "4", "req": "REQ_REC_003", "q": "", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-REC-02", "mod": "REC", "ten": "Yêu cầu tuyển trong định biên (tuyển mới / thay thế)", "uc": "UC-BP-REC-02", "muc_dich": "Tuyển đúng vị trí đã có trong kế hoạch đầu năm; ghi rõ lý do tuyển mới hoặc thay thế; quy trình duyệt ngắn gọn hơn khi trong định biên.", "ppt": "4", "req": "REQ_REC_001", "q": "Q-REC-HEADCOUNT", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-REC-02b", "mod": "REC", "ten": "Yêu cầu tuyển ngoài định biên (tuyển mới / thay thế)", "uc": "UC-BP-REC-02b", "muc_dich": "Tuyển phát sinh / vượt kế hoạch; ghi rõ tuyển mới hoặc thay thế; nhánh duyệt chặt hơn (thường có Ban giám đốc).", "ppt": "4", "req": "REQ_REC_001", "q": "Q-REC-HEADCOUNT", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-REC-02c", "mod": "REC", "ten": "Chiến dịch tuyển và hub tin đăng đa kênh", "uc": "UC-BP-REC-03", "muc_dich": "Gom nhiều yêu cầu tuyển cùng nhóm kỹ năng; đăng tin đa kênh khi đã có đối tác kết nối hệ thống. Không thuộc giai đoạn hiện tại — trạng thái đăng tin / hồ sơ / phỏng vấn gắn trên yêu cầu tuyển.", "ppt": "4", "req": "REQ_REC_002; REQ_REC_005", "q": "", "uu_tien": "Giai đoạn 2 — ngoài phạm vi hiện tại"},
    {"wbs": "WBS-REC-03", "mod": "REC", "ten": "Kho ứng viên gắn yêu cầu tuyển (liên kết bắt buộc)", "uc": "UC-BP-REC-04; UC-BP-REC-05", "muc_dich": "Mỗi hồ sơ ứng viên phải gắn ít nhất một yêu cầu tuyển; ưu tiên quét kho nội bộ; giữ lịch sử nguồn, từ chối đề nghị, mức lương mong muốn và trạng thái theo yêu cầu.", "ppt": "5", "req": "REQ_REC_002", "q": "", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-REC-04", "mod": "REC", "ten": "Phỏng vấn, thư mẫu và đánh giá (trong pipeline ứng viên)", "uc": "UC-BP-REC-06", "muc_dich": "Gửi thư cảm ơn / lịch phỏng vấn / đề nghị nhận việc theo mẫu; đánh giá đạt/không đạt trên đúng yêu cầu tuyển đã gắn — không menu chiến dịch rời ở giai đoạn hiện tại.", "ppt": "5", "req": "REQ_REC_004", "q": "", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-REC-05", "mod": "REC", "ten": "Chấp nhận đề nghị → mở hồ sơ nhân sự", "uc": "UC-BP-REC-07", "muc_dich": "Khi ứng viên nhận việc, tạo hồ sơ nhân sự mới từ dữ liệu đã có — không nhập lại tay. Tuyển dụng không nối thẳng sang tiền lương.", "ppt": "13", "req": "REQ_REC_004", "q": "", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-REC-06", "mod": "REC", "ten": "Báo cáo tuyển — kế hoạch so với thực tế", "uc": "UC-BP-REC-08", "muc_dich": "Đối chiếu số cần tuyển theo kế hoạch với số đã chốt theo thời gian × phòng ban × cấp; trả lời «bao giờ đủ người».", "ppt": "4", "req": "REQ_REC_005", "q": "", "uu_tien": "Ưu tiên gửi chốt"},
    # CORE — HĐ/BH mật; quản lý công việc = module riêng (ngoài khối này)
    {"wbs": "WBS-CORE-01", "mod": "CORE", "ten": "Hồ sơ công khai — tách vòng hợp đồng & bảo hiểm mật", "uc": "UC-BP-CORE-01; UC-BP-CORE-02; UC-BP-CORE-02b", "muc_dich": "Hồ sơ hành chính/phúc lợi (kèm người phụ thuộc) tách với lương, thuế, bảo hiểm, ngân hàng — chỉ người đúng quyền xem phần mật. Quản lý công việc / dự án không thuộc khối nhân sự hành chính.", "ppt": "6", "req": "REQ_HR_001; HR-001; PAY-001", "q": "Q-XBOT-PROFILE", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-CORE-02", "mod": "CORE", "ten": "Hợp đồng lao động và phụ lục lương", "uc": "UC-BP-CORE-09", "muc_dich": "Quản lý hợp đồng và phụ lục đổi lương (vòng mật); in ấn từ mẫu có điền sẵn thông tin chuẩn — nuôi dữ liệu cho kỳ lương.", "ppt": "13", "req": "HR-002", "q": "", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-CORE-03", "mod": "CORE", "ten": "Danh mục giấy tờ và đọc giấy tờ tự động", "uc": "UC-BP-CORE-03; UC-BP-CORE-04", "muc_dich": "Theo dõi giấy tờ bắt buộc theo vị trí; hỗ trợ lấy sẵn thông tin từ bản quét để không nhập lại.", "ppt": "7", "req": "HR-003", "q": "", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-CORE-04", "mod": "CORE", "ten": "Cấp phát và thu hồi tài sản", "uc": "UC-BP-CORE-05; UC-BP-CORE-06", "muc_dich": "Gán laptop/xe/công cụ kèm biên bản; khi nghỉ việc thu hồi đủ trước khi tất toán.", "ppt": "7 · 13", "req": "HR-006", "q": "Q-ASSET-MODULE", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-CORE-05", "mod": "CORE", "ten": "Vòng đời trạng thái hồ sơ nhân sự", "uc": "UC-BP-CORE-07", "muc_dich": "Chỉ khi hồ sơ đủ giấy tờ mới chuyển «Hoạt động», mở phép năm và ca làm việc, sẵn sàng vào kỳ lương.", "ppt": "13", "req": "HR-003; HR-004; REQ_NP_005; HR-006", "q": "Q-SI-SUSPEND", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-CORE-06", "mod": "CORE", "ten": "Khen thưởng và kỷ luật vào bảng lương", "uc": "UC-BP-CORE-08", "muc_dich": "Thưởng/phạt có trạng thái thi hành và xuất hiện đúng trên phiếu lương kỳ đích — không chỉ ghi chú rồi quên.", "ppt": "7 · 11", "req": "HR-005", "q": "", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-CORE-07", "mod": "CORE", "ten": "Bảo hiểm xã hội theo vòng đời (vòng mật)", "uc": "UC-BP-CORE-10", "muc_dich": "Theo dõi đóng / ngừng / tạm hoãn bảo hiểm theo timeline; đổi hàng loạt có xem trước và lưu vết — thuộc vòng hợp đồng & bảo hiểm mật.", "ppt": "—", "req": "HR-004", "q": "Q-SI-SUSPEND", "uu_tien": "Bổ sung sau khi chốt khung"},
    # ATT — ca + loại phép; bảng công chốt = đầu vào lương
    {"wbs": "WBS-ATT-01", "mod": "ATT", "ten": "Thiết lập ca và quy tắc phạt muộn/về sớm", "uc": "UC-BP-ATT-01; UC-BP-ATT-02", "muc_dich": "Mỗi bộ phận/nhóm có ca riêng; phạt theo phút, theo khối thời gian hoặc theo bậc — chỉ khi điểm danh hợp lệ.", "ppt": "8", "req": "TIME-001; TIME-002", "q": "", "uu_tien": "Ưu tiên gửi chốt (phạt)"},
    {"wbs": "WBS-ATT-02", "mod": "ATT", "ten": "Thu nhận điểm danh từ nhiều nguồn", "uc": "UC-BP-ATT-03", "muc_dich": "Ứng dụng, địa chỉ mạng, vị trí, máy chấm → cùng một quy tắc ca → một kết quả giờ công thô; giải trình sau duyệt cập nhật công.", "ppt": "8", "req": "TIME-002; REQ_CC_002", "q": "", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-ATT-03", "mod": "ATT", "ten": "Lịch nghỉ lễ và ngày không làm việc", "uc": "UC-BP-ATT-03b", "muc_dich": "Lịch dương cố định và ngày âm lịch cấu hình theo năm/pháp nhân — dùng chung cho phép và bảng công.", "ppt": "8 · 9", "req": "REQ_CC_001", "q": "", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-ATT-04", "mod": "ATT", "ten": "Các loại nghỉ phép (năm · thâm niên · bù tăng ca · chuyển kỳ · ứng)", "uc": "UC-BP-ATT-04; UC-BP-ATT-04b; UC-BP-ATT-05; UC-BP-ATT-06; UC-BP-ATT-07", "muc_dich": "Cấu hình quỹ: phép năm, thâm niên, nghỉ bù từ tăng ca, mang sang / bảo lưu kỳ, ứng phép; nghỉ ốm xét chế độ bảo hiểm và hỗ trợ công ty nếu có.", "ppt": "9", "req": "REQ_NP_001; REQ_NP_002; REQ_NP_004; REQ_NP_005; HR-004", "q": "Q-LEAVE-ACCRUAL; Q-LEAVE-UNIT", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-ATT-05", "mod": "ATT", "ten": "Trừ phép đúng ngày làm và giữ chỗ quỹ phép", "uc": "UC-BP-ATT-08; UC-BP-ATT-09", "muc_dich": "Đơn xuyên cuối tuần/lễ chỉ trừ ngày làm việc; khi gửi đơn thì giữ chỗ quỹ; duyệt mới trừ thật; từ chối hoàn lại.", "ppt": "9", "req": "REQ_NP_003; REQ_NP_006", "q": "Q-LEAVE-UNIT", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-ATT-06", "mod": "ATT", "ten": "Tổng hợp và ký chốt bảng công (đầu vào tính lương)", "uc": "UC-BP-ATT-10; UC-BP-ATT-11; UC-BP-PAY-01", "muc_dich": "Gộp chấm công + phép + tăng ca đã hệ số thành bảng công tổng hợp; ký chốt trước khi tính lương. Tiền lương chỉ đọc bảng đã chốt — không lấy giờ thẳng từ tăng ca hay đơn phép.", "ppt": "10", "req": "REQ_L_001", "q": "", "uu_tien": "Ưu tiên gửi chốt"},
    # PAY — đã họp; nguồn giờ = bảng công chốt; C&B từ HĐ/BH
    {"wbs": "WBS-PAY-01", "mod": "PAY", "ten": "Động cơ công thức lương (cấu hình, không cố định cứng)", "uc": "UC-BP-PAY-02", "muc_dich": "Lắp biến số từ bảng công đã chốt, lương cơ bản / phụ cấp / bảo hiểm (vòng hợp đồng–bảo hiểm), thuế…; công thức có phiên bản và kiểm soát trước khi áp dụng.", "ppt": "11", "req": "REQ_L_002; PAY-001", "q": "Q-PAY-FORMULA", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-PAY-02", "mod": "PAY", "ten": "Giảm trừ gia cảnh, thuế và trần bảo hiểm", "uc": "UC-BP-PAY-03; UC-BP-PAY-05", "muc_dich": "Lấy giảm trừ từ hồ sơ có quyền; trần bảo hiểm tính một lần trên tổng hợp kỳ (kể cả khi đổi lương giữa tháng).", "ppt": "11 · 12", "req": "REQ_L_003; HR-004", "q": "", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-PAY-03", "mod": "PAY", "ten": "Gộp lương khi đổi điều kiện giữa tháng", "uc": "UC-BP-PAY-04", "muc_dich": "Hai đoạn lương trong tháng → một phiếu thực nhận; giảm trừ gia cảnh và trần bảo hiểm không tính hai lần.", "ppt": "12", "req": "REQ_L_004", "q": "", "uu_tien": "Ưu tiên gửi chốt"},
    {"wbs": "WBS-PAY-04", "mod": "PAY", "ten": "Tính lương kỳ và tất toán nghỉ việc", "uc": "UC-BP-PAY-06; UC-BP-PAY-07", "muc_dich": "Chạy kỳ khi nhân sự đang hoạt động và bảng công đã chốt; nghỉ việc: cắt bảo hiểm, trả phép còn, thu hồi tài sản, thưởng/phạt kỳ cuối.", "ppt": "11 · 13", "req": "REQ_L_001; REQ_NP_005; HR-006; HR-004; HR-005", "q": "", "uu_tien": "Bổ sung sau khi chốt khung"},
    {"wbs": "WBS-PAY-05", "mod": "PAY", "ten": "Phiếu lương và phân nhóm bảng lương", "uc": "UC-BP-PAY-08; UC-BP-PAY-09", "muc_dich": "Xem trước phiếu, bảo mật từng người; nhóm văn phòng / kinh doanh / tài xế / vận hành theo cấu hình.", "ppt": "11", "req": "REQ_L_005; REQ_L_006", "q": "", "uu_tien": "Bổ sung sau khi chốt khung"},
]


# Actors by module (plain VN)
# Vai trò nghiệp vụ — tên chức năng khách hiểu được (không dùng biệt ngữ nội bộ «Đãi ngộ/C&B» một mình)
ACTORS = {
    "REC": "Trưởng bộ phận · Nhân sự tuyển dụng · Ban giám đốc (khi ngoài kế hoạch) · Hệ thống (tự tạo yêu cầu)",
    "CORE": "Nhân sự hành chính · Chuyên viên lương thưởng & phúc lợi · Nhân viên · Quản lý trực tiếp",
    "ATT": "Nhân viên · Quản lý trực tiếp · Nhân sự chấm công · Hệ thống (giữ chỗ quỹ phép, tính ngày làm)",
    "PAY": "Chuyên viên lương thưởng & phúc lợi · Nhân viên (xem phiếu lương) · Nhân sự (ký bảng công) · Ban lãnh đạo (chính sách)",
}

FR_PRIORITY = {
    "UC-BP-REC-01", "UC-BP-REC-01b", "UC-BP-REC-02", "UC-BP-REC-02b", "UC-BP-REC-08",
    "UC-BP-CORE-01", "UC-BP-CORE-02", "UC-BP-CORE-08",
    "UC-BP-ATT-02", "UC-BP-ATT-08", "UC-BP-ATT-09", "UC-BP-ATT-10", "UC-BP-ATT-11",
    "UC-BP-PAY-01", "UC-BP-PAY-02", "UC-BP-PAY-04",
}

WBS_FOR_UC = {
    "UC-BP-REC-00": "WBS-REC-00",
    "UC-BP-REC-01": "WBS-REC-01", "UC-BP-REC-01b": "WBS-REC-01b", "UC-BP-REC-02": "WBS-REC-02",
    "UC-BP-REC-02b": "WBS-REC-02b", "UC-BP-REC-03": "WBS-REC-02c", "UC-BP-REC-04": "WBS-REC-03",
    "UC-BP-REC-05": "WBS-REC-03", "UC-BP-REC-06": "WBS-REC-04", "UC-BP-REC-07": "WBS-REC-05",
    "UC-BP-REC-08": "WBS-REC-06",
    "UC-BP-CORE-01": "WBS-CORE-01", "UC-BP-CORE-02": "WBS-CORE-01", "UC-BP-CORE-02b": "WBS-CORE-01",
    "UC-BP-CORE-03": "WBS-CORE-03", "UC-BP-CORE-04": "WBS-CORE-03", "UC-BP-CORE-05": "WBS-CORE-04",
    "UC-BP-CORE-06": "WBS-CORE-04", "UC-BP-CORE-07": "WBS-CORE-05", "UC-BP-CORE-08": "WBS-CORE-06",
    "UC-BP-CORE-09": "WBS-CORE-02", "UC-BP-CORE-10": "WBS-CORE-07",
    "UC-BP-ATT-01": "WBS-ATT-01", "UC-BP-ATT-02": "WBS-ATT-01", "UC-BP-ATT-03": "WBS-ATT-02",
    "UC-BP-ATT-03b": "WBS-ATT-03", "UC-BP-ATT-04": "WBS-ATT-04", "UC-BP-ATT-04b": "WBS-ATT-04",
    "UC-BP-ATT-05": "WBS-ATT-04", "UC-BP-ATT-06": "WBS-ATT-04", "UC-BP-ATT-07": "WBS-ATT-04",
    "UC-BP-ATT-08": "WBS-ATT-05", "UC-BP-ATT-09": "WBS-ATT-05", "UC-BP-ATT-10": "WBS-ATT-06",
    "UC-BP-ATT-11": "WBS-ATT-06", "UC-BP-ATT-12": "WBS-CORE-05",
    "UC-BP-PAY-01": "WBS-ATT-06", "UC-BP-PAY-02": "WBS-PAY-01", "UC-BP-PAY-03": "WBS-PAY-02",
    "UC-BP-PAY-04": "WBS-PAY-03", "UC-BP-PAY-05": "WBS-PAY-02", "UC-BP-PAY-06": "WBS-PAY-04",
    "UC-BP-PAY-07": "WBS-PAY-04", "UC-BP-PAY-08": "WBS-PAY-05", "UC-BP-PAY-09": "WBS-PAY-05",
}

UC_TEN = {
    "UC-BP-REC-00": "Thư viện mô tả công việc (bản gốc chuẩn)",
    "UC-BP-REC-01": "Quản trị định biên vị trí theo 12 tháng (chỉ số cần tuyển)",
    "UC-BP-REC-01b": "Tự tạo yêu cầu tuyển theo tháng «Cần tuyển»",
    "UC-BP-REC-02": "Yêu cầu tuyển trong định biên (tuyển mới / thay thế)",
    "UC-BP-REC-02b": "Yêu cầu tuyển ngoài định biên (tuyển mới / thay thế)",
    "UC-BP-REC-03": "Chiến dịch tuyển / hub đa kênh (giai đoạn 2 — ngoài phạm vi hiện tại)",
    "UC-BP-REC-04": "Quét kho ứng viên nội bộ trước kênh ngoài",
    "UC-BP-REC-05": "Lịch sử trạng thái ứng viên gắn yêu cầu tuyển",
    "UC-BP-REC-06": "Gửi thư tuyển theo mẫu và đánh giá phỏng vấn",
    "UC-BP-REC-07": "Chấp nhận đề nghị nhận việc → tạo hồ sơ nhân sự",
    "UC-BP-REC-08": "Báo cáo tuyển — kế hoạch so với thực tế",
    "UC-BP-CORE-01": "Hồ sơ vòng công khai (hành chính / phúc lợi)",
    "UC-BP-CORE-02": "Hồ sơ vòng hợp đồng & bảo hiểm mật (lương, bảo hiểm, thuế, ngân hàng)",
    "UC-BP-CORE-02b": "Cấu hình nhóm thông tin trên hồ sơ",
    "UC-BP-CORE-03": "Danh mục giấy tờ động (bắt buộc / tùy chọn)",
    "UC-BP-CORE-04": "Đọc giấy tờ tự động — điền sẵn, không nhập lại",
    "UC-BP-CORE-05": "Cấp phát tài sản và biên bản bàn giao",
    "UC-BP-CORE-06": "Thu hồi tài sản khi nghỉ việc",
    "UC-BP-CORE-07": "Chuyển hồ sơ sang «Hoạt động» khi đủ giấy tờ",
    "UC-BP-CORE-08": "Khen thưởng & kỷ luật — thi hành vào bảng lương",
    "UC-BP-CORE-09": "Hợp đồng lao động — mẫu in điền sẵn thông tin",
    "UC-BP-CORE-10": "Bảo hiểm xã hội theo vòng đời (đóng / ngừng / tạm hoãn)",
    "UC-BP-ATT-01": "Thiết lập quy tắc ca theo bộ phận / nhóm",
    "UC-BP-ATT-02": "Phạt muộn / về sớm (phút / khối / bậc + nguồn hợp lệ)",
    "UC-BP-ATT-03": "Thu nhận điểm danh nhiều nguồn → giờ công thô",
    "UC-BP-ATT-03b": "Lịch lễ / Tết (dương và âm cấu hình theo năm)",
    "UC-BP-ATT-04": "Cấp phát phép năm theo thành phần cấu hình",
    "UC-BP-ATT-04b": "Ứng phép và thời điểm cấp / nghỉ không lương rồi bù trừ",
    "UC-BP-ATT-05": "Bảo lưu phép năm cũ đến hết quý 1",
    "UC-BP-ATT-06": "Phép nghỉ bù từ tăng ca (khi công ty bật)",
    "UC-BP-ATT-07": "Nghỉ ốm — bảo hiểm xã hội hoặc công ty hỗ trợ đủ",
    "UC-BP-ATT-08": "Tính ngày trừ phép xuyên cuối tuần và lễ",
    "UC-BP-ATT-09": "Nộp & duyệt phép — giữ chỗ quỹ khi gửi đơn",
    "UC-BP-ATT-10": "Tổng hợp bảng công (đầu vào tính lương)",
    "UC-BP-ATT-11": "Ký chốt bảng công trước khi tính lương",
    "UC-BP-ATT-12": "Mở quỹ phép và ca mặc định khi hồ sơ Hoạt động",
    "UC-BP-PAY-01": "Lương chỉ lấy số từ bảng công đã chốt",
    "UC-BP-PAY-02": "Lắp ráp và chạy công thức lương cấu hình được",
    "UC-BP-PAY-03": "Giảm trừ gia cảnh từ hồ sơ (đủ quyền)",
    "UC-BP-PAY-04": "Gộp lương khi đổi điều kiện giữa tháng (không trừ gia cảnh hai lần)",
    "UC-BP-PAY-05": "Trần bảo hiểm trên tổng hợp kỳ (kể cả gộp giữa tháng)",
    "UC-BP-PAY-06": "Tính lương kỳ khi đã Hoạt động và bảng công chốt",
    "UC-BP-PAY-07": "Tất toán nghỉ việc — bảo hiểm, phép, tài sản, thưởng/phạt kỳ cuối",
    "UC-BP-PAY-08": "Phiếu lương — xem trước, bảo mật, trạng thái thanh toán",
    "UC-BP-PAY-09": "Phân nhóm bảng lương (văn phòng / kinh doanh / tài xế / vận hành)",
}


def soft_vn(text: str) -> str:
    """Bỏ jargon kỹ thuật, giữ nghĩa nghiệp vụ tiếng Việt."""
    if not text:
        return ""
    t = text
    reps = [
        (r"`[^`]+`", ""),
        (r"\bAPI\b", "kết nối hệ thống"),
        (r"\bUI\b", "màn hình"),
        (r"\bSoT\b", "nguồn gốc chuẩn"),
        (r"\bXOR\b", "hoặc (chỉ một chế độ)"),
        (r"\bhardcode\b", "cố định cứng trong bản phát hành"),
        (r"\bHardcode\b", "Cố định cứng"),
        (r"\bdeploy\b", "phát hành phần mềm"),
        (r"\bmetadata\b", "cấu hình nhóm thông tin"),
        (r"\blegal entity\b", "pháp nhân"),
        (r"\bcompany_id\b", "pháp nhân"),
        (r"\bemployee_id\b", "mã nhân viên"),
        (r"\bheadcount_mode=in_plan\b", "chế độ trong kế hoạch định biên"),
        (r"\bheadcount_mode=out_of_plan\b", "chế độ ngoài kế hoạch định biên"),
        (r"\bplan_month\b", "tháng kế hoạch"),
        (r"\bYCTD\b", "yêu cầu tuyển dụng"),
        (r"\bĐB\b", "định biên"),
        (r"\bHC\b", "định biên nhân sự"),
        (r"\bBOD\b", "Ban giám đốc"),
        (r"\bBGĐ\b", "Ban giám đốc"),
        (r"\bTPB\b", "trưởng phòng ban"),
        (r"\bHCNS\b", "hành chính nhân sự"),
        (r"\bC&B\b", "chuyên viên lương thưởng và phúc lợi"),
        (r"\bđãi ngộ\b", "chuyên viên lương thưởng và phúc lợi"),
        (r"\bĐãi ngộ\b", "Chuyên viên lương thưởng và phúc lợi"),
        (r"\bnhóm đãi ngộ\b", "nhóm lương thưởng và phúc lợi"),
        (r"\bhồ sơ đãi ngộ\b", "hồ sơ lương thưởng"),
        (r"\bvòng đãi ngộ\b", "phần lương thưởng mật"),
        (r"\bPayroll\b", "tiền lương"),
        (r"\bpayroll\b", "tiền lương"),
        (r"\bpayslip\b", "phiếu lương"),
        (r"\btimesheet\b", "bảng công"),
        (r"\bgeofence\b", "vùng vị trí cho phép"),
        (r"\bGPS\b", "vị trí"),
        (r"\bIP\b", "địa chỉ mạng nội bộ"),
        (r"\bOU\b", "đơn vị / bộ phận"),
        (r"\bOT\b", "tăng ca"),
        (r"\bGTCG\b", "giảm trừ gia cảnh"),
        (r"\bNPT\b", "người phụ thuộc"),
        (r"\bTNCN\b", "thuế thu nhập cá nhân"),
        (r"\bBHXH\b", "bảo hiểm xã hội"),
        (r"\bBH\b", "bảo hiểm"),
        (r"\bPC\b", "phụ cấp"),
        (r"\bCB\b", "cơ bản"),
        (r"\bMST\b", "mã số thuế"),
        (r"\bNH\b", "tài khoản ngân hàng"),
        (r"\bKPI\b", "chỉ tiêu kết quả"),
        (r"\bNet\b", "thực nhận"),
        (r"\bgross\b", "tổng thu nhập trước khấu trừ"),
        (r"\bCV\b", "hồ sơ ứng viên"),
        (r"\bUV\b", "ứng viên"),
        (r"\bPV\b", "phỏng vấn"),
        (r"\bCSVC\b", "cơ sở vật chất"),
        (r"\bPDF\b", "bản quét / tệp tài liệu"),
        (r"\bOCR\b", "đọc chữ từ bản quét"),
        (r"\bWord\b", "mẫu văn bản"),
        (r"\bkeyword fill\b", "điền sẵn từ khóa"),
        (r"\bkeyword map\b", "ánh xạ trường thông tin"),
        (r"\bdual-control\b", "hai bước kiểm soát (soạn và phát hành)"),
        (r"\bDual-control\b", "Hai bước kiểm soát"),
        (r"\bversioned\b", "có phiên bản"),
        (r"\baudit\b", "lưu vết"),
        (r"\baudit trail\b", "nhật ký lưu vết"),
        (r"\bhold\b", "giữ chỗ"),
        (r"\bHold\b", "Giữ chỗ"),
        (r"\bavailable\b", "số còn dùng được"),
        (r"\bdeducted\b", "đã trừ"),
        (r"\bworking-day\b", "ngày làm việc"),
        (r"\bcalendar day\b", "ngày lịch"),
        (r"\boverlapping\b", "chồng ngày"),
        (r"\brecalculate\b", "tính lại"),
        (r"\bAccrual\b", "Cấp dần"),
        (r"\baccrual\b", "cấp dần"),
        (r"\bcarry\b", "phép mang sang"),
        (r"\bpayout\b", "trả tiền phép"),
        (r"\bterminate\b", "nghỉ việc"),
        (r"\boffboard\b", "thủ tục nghỉ việc"),
        (r"\bsplit-month\b", "gộp lương giữa tháng"),
        (r"\bsplit\b", "gộp giữa tháng"),
        (r"\bpro-rate\b", "tính tỷ lệ theo ngày"),
        (r"\btoggle\b", "bật/tắt"),
        (r"\bmembership\b", "quyền thành viên"),
        (r"\bscope\b", "phạm vi công ty"),
        (r"\bdraft\b", "bản nháp"),
        (r"\bActive\b", "Hoạt động"),
        (r"\bactive\b", "đang áp dụng"),
        (r"\bre-open\b", "mở lại"),
        (r"\bspawn\b", "tạo"),
        (r"\bbypass\b", "bỏ qua"),
        (r"\bfunnel\b", "phễu theo dõi"),
        (r"\bFunnel\b", "Phễu theo dõi"),
        (r"\bdashboard\b", "bảng theo dõi"),
        (r"\bDashboard\b", "Bảng theo dõi"),
        (r"\bAS-IS\b", "hiện trạng"),
        (r"\bTO-BE\b", "mục tiêu"),
        (r"\bFR\b", "mô tả chức năng"),
        (r"\bAC\b", "tiêu chí chấp nhận"),
        (r"\bBR\b", "quy tắc"),
        (r"\bPPT\b", "trang trình bày"),
        (r"\bLE\b", "pháp nhân"),
        (r"\bPB\b", "phòng ban"),
        (r"\bTP\b", "trưởng phòng"),
        (r"\bTT\b", "thanh toán"),
        (r"\bNS\b", "ngân sách"),
        (r"\bKH vs TT\b", "kế hoạch so với thực tế"),
        (r"\bKH vs thực tế\b", "kế hoạch so với thực tế"),
        (r"\bgeo\b", "vị trí"),
        (r"\bGEO\b", "quy tắc vị trí"),
        (r"\bCC\b", "gửi kèm"),
        (r"\bno-show\b", "không đến phỏng vấn"),
        (r"\bonboard\b", "nhận việc"),
        (r"\bOnboard\b", "Nhận việc"),
        (r"\bserial\b", "số máy"),
        (r"\bstub\b", "bản ghi tạm giai dõi"),
        (r"\bfull Asset\b", "module tài sản đầy đủ"),
        (r"\bphase\b", "giai đoạn"),
        (r"\bpolicy\b", "chính sách"),
        (r"\bPolicy\b", "Chính sách"),
        (r"\bgrace\b", "thời gian ân hạn"),
        (r"\btimestamp\b", "mốc giờ ghi nhận"),
        (r"\bpipeline\b", "chuỗi xử lý"),
        (r"\bunit=hour\b", "đơn vị tính theo giờ"),
        (r"\bunit\b", "đơn vị tính"),
        (r"\bcode path\b", "lối xử lý phần mềm"),
        (r"\bFork code\b", "Tách bản phần mềm"),
        (r"\bdependency\b", "phụ thuộc"),
        (r"\bmaster\b", "bản gốc chuẩn"),
        (r"\boverride\b", "ghi đè ưu tiên"),
        (r"\bglobal\b", "chung toàn hệ"),
        (r"\brelease\b", "bản phát hành"),
        (r"\bXbot\b", "lớp cấu hình tập đoàn"),
        (r"\bXBOS\b", "hệ điều hành doanh nghiệp"),
        (r"\bcatalog\b", "danh mục"),
        (r"\bDecision\b", "Quyết định cần chốt"),
        (r"\bMISSING\b", "còn thiếu so với mục tiêu"),
        (r"\bPARTIAL\b", "đã có một phần"),
        (r"\bCONFLICT\b", "mâu thuẫn cần chốt"),
        (r"\bDECISION\b", "đang chờ quyết định"),
        (r"  +", " "),
        (r"\s{2,}", " "),
    ]
    for a, b in reps:
        t = re.sub(a, b, t)
    return t.strip(" ;·-|")


def parse_matrix_rows(md_path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for line in md_path.read_text(encoding="utf-8").splitlines():
        if not line.startswith("| UC-BP-"):
            continue
        parts = [p.strip() for p in line.strip().strip("|").split("|")]
        if len(parts) < 12:
            continue
        rows.append({
            "uc": parts[0],
            "mod": parts[1],
            "task_wbs_label": parts[2],
            "req": parts[3],
            "ppt": parts[4],
            "dien_bien": soft_vn(parts[5]),
            "br": parts[6],
            "pass": soft_vn(parts[7]),
            "fail": soft_vn(parts[8]),
            "edge": soft_vn(parts[9]),
            "rui_ro": soft_vn(parts[10]),
            "gap": soft_vn(parts[11]),
        })
    return rows


def quy_tac_tu_br(br: str, pass_txt: str) -> str:
    """Gộp quy tắc + kiểm tra thành câu khách đọc."""
    br_map = {
        "BR-BP-HC-01": "Mỗi ô tháng trên lưới định biên chỉ mang một trạng thái (Hiện tại / Dự kiến / Cần tuyển). «Cần tuyển» chỉ gắn đúng tháng kích hoạt. Sau duyệt không chỉnh tay tùy tiện.",
        "BR-BP-HC-04": "Mỗi ô «Cần tuyển» đã duyệt sinh đúng một yêu cầu tuyển gắn tháng, vị trí và số lượng. Mở lại cùng phiên bản định biên không sinh trùng.",
        "BR-BP-HC-05": "Trong định biên đã duyệt: luồng duyệt ngắn hơn — không bắt buộc Ban giám đốc nếu chính sách pháp nhân cho phép; vẫn qua trưởng phòng và nhân sự tối thiểu.",
        "BR-BP-HC-06": "Ngoài định biên / vượt kế hoạch: nhánh duyệt dài hơn. Thiếu cấp Ban giám đốc (khi cấu hình yêu cầu) thì chưa được mở tin tuyển.",
        "BR-BP-HC-03": "Giai đoạn 2: chiến dịch liệt kê đủ yêu cầu nguồn; đóng không xóa lịch sử hồ sơ; không trộn pháp nhân. Giai đoạn hiện tại: trạng thái đăng tin / hồ sơ gắn trên yêu cầu tuyển.",
        "BR-BP-JD-01": "Mô tả công việc chuẩn là đầu vào tái sử dụng cho yêu cầu tuyển; không bắt nhập lại toàn bộ mô tả mỗi lần tạo yêu cầu khi đã chọn bản gốc còn hiệu lực.",
        "BR-BP-CV-01": "Trước khi đăng kênh ngoài phải quét kho nội bộ (hoặc bỏ qua có lý do và đủ quyền).",
        "BR-BP-CV-02": "Giữ lịch sử nguồn, từ chối đề nghị, mức lương mong muốn — không ghi đè mất quá khứ.",
        "BR-BP-MAIL-01": "Thư lịch phỏng vấn bắt buộc gửi kèm toàn bộ người phỏng vấn; mọi lần gửi lưu vết; đánh giá đạt/không đạt lưu được.",
        "BR-BP-LC-01": "Nhận đề nghị tạo hồ sơ chờ hoàn thiện từ dữ liệu ứng viên — không nhập lại các trường đã có. Tuyển dụng không trao đổi trực tiếp với tiền lương.",
        "BR-BP-SEC-01": "Hồ sơ chung không hiển thị lương cơ bản, phụ cấp tiền, mã số thuế, tài khoản ngân hàng, số sổ bảo hiểm.",
        "BR-BP-SEC-02": "Chỉ chuyên viên lương thưởng và phúc lợi mới đọc/ghi phần mật; mọi truy cập và sửa được lưu vết; hiệu lực theo ngày.",
        "BR-BP-SEC-03": "Nhóm trường hồ sơ cấu hình theo pháp nhân; vẫn tôn trọng tách công khai / mật.",
        "BR-BP-DOC-01": "Thiếu giấy bắt buộc thì chưa hoàn thiện / chưa kích hoạt hồ sơ; nêu rõ từng loại còn thiếu.",
        "BR-BP-OCR-01": "Trường đã đọc từ bản quét không bắt nhập lại; được sửa từng trường lệch.",
        "BR-BP-AST-01": "Tài sản «Đang sử dụng» gắn nhân viên; biên bản có xác nhận hai bên; lưu mã và số máy.",
        "BR-BP-AST-02": "Lệnh nghỉ việc đưa 100% tài sản đang dùng vào danh sách thu hồi; còn món bắt buộc chưa thu thì chưa tất toán cuối.",
        "BR-BP-RD-01": "Thưởng/phạt đang hoặc đã thi hành xuất hiện đúng kỳ lương đích; hủy thi hành thì không vào kỳ chưa chốt; sau chốt kỳ không sửa phiếu đã khóa.",
        "BR-BP-CTR-01": "Hợp đồng in từ mẫu khớp bản gốc (mã ký, hiệu lực, vị trí, lương); phụ lục đổi lương tạo phiên bản và nuôi luồng gộp giữa tháng.",
        "BR-BP-SI-01": "Trạng thái bảo hiểm (Hoạt động / Ngừng / Tạm hoãn) phản ánh đúng kỳ; đổi hàng loạt có xem trước và lưu vết.",
        "BR-BP-LC-02": "Chờ hoàn thiện → đủ checklist → Hoạt động → mở phép và ca → sẵn sàng lương.",
        "BR-BP-SHF-01": "Công tính theo ca đang gán thực tế của bộ phận/nhóm — không một bộ quy tắc chung ghi đè mọi đơn vị.",
        "BR-BP-SHF-02": "Chỉ một chế độ phạt (phút hoặc khối thời gian hoặc bậc). Chỉ áp khi nguồn điểm danh hợp lệ theo cấu hình đơn vị.",
        "BR-BP-ATT-01": "Bản ghi điểm danh có nguồn và vị trí/địa chỉ mạng khi bắt buộc; sau duyệt giải trình thì công cập nhật kèm lưu vết.",
        "BR-BP-HOL-01": "Phép và bảng công dùng chung bộ lịch năm (dương + ngày âm đã cấu hình theo pháp nhân).",
        "BR-BP-LV-01": "Số dư phép = tổng thành phần cấu hình (cấp dần theo tháng, thâm niên, chức vụ); nửa tháng vào/nghỉ tính nửa ngày gốc.",
        "BR-BP-LV-07": "Tắt ứng trước thì chặn đơn vượt số dư; bật thì cho đến trần; nghỉ không lương có loại riêng rồi bù trừ khi có quỹ.",
        "BR-BP-LV-02": "Phép mang sang thường cắt hết sau quý 1; nghỉ việc trả tiền phép còn theo đơn giá lương cơ bản đóng bảo hiểm theo chính sách.",
        "BR-BP-LV-03": "Chỉ cộng phép nghỉ bù khi tăng ca đã duyệt và công ty bật chế độ; tỷ lệ giờ sang ngày cấu hình được.",
        "BR-BP-LV-04": "Nghỉ ốm đi đúng một nhánh: bảo hiểm xã hội hoặc công ty hỗ trợ đủ — không áp hai nhánh cùng lúc không quy tắc.",
        "BR-BP-LV-05": "Đơn từ Thứ sáu đến Thứ ba chỉ trừ ngày làm việc (ví dụ chuẩn: trừ 2 ngày); Thứ bảy, Chủ nhật, lễ = 0. Đơn vị tối thiểu nửa ngày hoặc một giờ theo cấu hình.",
        "BR-BP-LV-06": "Gửi đơn = giữ chỗ quỹ theo số ngày trừ dự kiến; từ chối hoàn đủ giữ chỗ; duyệt chuyển sang đã trừ; chặn hai đơn chồng ngày.",
        "BR-BP-TS-01": "Một kỳ một bảng; giờ tăng ca vào bảng đã nhân hệ số — tiền lương không nhân lại lần nữa.",
        "BR-BP-TS-02": "Đủ chữ ký/xác nhận bắt buộc mới mở lệnh tính lương; hủy chốt có lý do và quyền.",
        "BR-BP-TS-03": "Tiền lương chỉ đọc bảng công đã chốt — không lấy giờ thẳng từ tăng ca hay phép để tính lương.",
        "BR-BP-PAY-01": "Công thức có phiên bản; soạn và phát hành tách quyền; không cố định cứng công thức trong mỗi lần phát hành phần mềm.",
        "BR-BP-PAY-02": "Giảm trừ gia cảnh lấy từ hồ sơ đủ quyền; không nhập trùng trên bảng lương.",
        "BR-BP-SPL-01": "Hai đoạn lương → một phiếu thực nhận; phần cộng dồn cộng hai đoạn; giảm trừ gia cảnh, thuế, trần bảo hiểm tính một lần trên tổng hợp.",
        "BR-BP-SPL-02": "Trần bảo hiểm không áp hai lần từng đoạn khi gộp giữa tháng.",
        "BR-BP-PAY-03": "Nhân viên chỉ mở phiếu của mình; có trạng thái thanh toán / công nợ; xem trước trước khi gửi chính thức.",
        "BR-BP-PAY-04": "Mỗi nhân viên thuộc một nhóm bảng lương đang áp dụng (hoặc quy tắc ưu tiên rõ).",
        "BR-BP-LC-04": "Chạy lương khi bảng công đã chốt và nhân sự đang Hoạt động.",
        "BR-BP-LC-05": "Nghỉ việc: cắt bảo hiểm, trả phép còn, thu hồi tài sản, thưởng/phạt đang thi hành vào kỳ cuối.",
    }
    base = br_map.get(br, soft_vn(pass_txt))
    return base


def validate_tu_pass_fail(pass_txt: str, fail_txt: str, edge: str) -> str:
    parts = []
    if pass_txt:
        parts.append("Khi hợp lệ: " + pass_txt)
    if fail_txt:
        parts.append("Khi không hợp lệ: " + fail_txt)
    if edge:
        parts.append("Tình huống đặc biệt cần nhớ: " + edge)
    return " | ".join(parts)


def gap_vi(g: str) -> str:
    g = g.upper()
    if "DECISION" in g:
        return "Cần chốt quyết định với khách trước khi viết sâu"
    if "CONFLICT" in g:
        return "Hiện trạng có thể mâu thuẫn — cần chốt hướng xử lý"
    if "MISSING" in g:
        return "Còn thiếu so với mục tiêu — ưu tiên mô tả đủ trước khi làm"
    if "PARTIAL" in g:
        return "Đã có một phần — cần bổ sung nhánh đặc biệt"
    return soft_vn(g) or "Đang theo dõi"


# Extra UC rows not always in matrix table (REC-07, CORE-07, ATT-12, PAY-06/07)
EXTRA_UC: list[dict[str, str]] = [
    {
        "uc": "UC-BP-REC-00", "mod": "REC", "task_wbs_label": "Thư viện mô tả công việc",
        "req": "REQ_REC_003", "ppt": "4",
        "dien_bien": "Nhân sự tuyển dụng / trưởng bộ phận tạo và duy trì mô tả công việc chuẩn theo vị trí; yêu cầu tuyển tham chiếu bản gốc còn hiệu lực thay vì gõ lại mô tả mỗi lần.",
        "br": "BR-BP-JD-01",
        "pass": "Tạo yêu cầu tuyển chọn được mô tả công việc còn hiệu lực; không bắt nhập lại toàn bộ mô tả khi đã gắn bản gốc.",
        "fail": "Mỗi lần tạo yêu cầu phải gõ lại toàn bộ mô tả dù đã có bản gốc chuẩn.",
        "edge": "Vị trí chưa có mô tả chuẩn — cho phép soạn kèm hoặc bổ sung thư viện trước khi duyệt yêu cầu (theo chính sách đơn vị).",
        "rui_ro": "Mô tả lệch giữa các yêu cầu cùng vị trí → tuyển sai hồ sơ / sai kỳ vọng.",
        "gap": "PARTIAL",
    },
    {
        "uc": "UC-BP-REC-08", "mod": "REC", "task_wbs_label": "Báo cáo tuyển — kế hoạch so với thực tế",
        "req": "REQ_REC_005", "ppt": "4",
        "dien_bien": "Theo dõi phễu hồ sơ → phỏng vấn → chốt người gắn yêu cầu tuyển; đối chiếu kế hoạch với thực tế theo thời gian × phòng ban × cấp; trả lời «bao giờ đủ người».",
        "br": "BR-BP-HC-01",
        "pass": "Bảng theo dõi hiển thị tiến độ theo tháng × phòng ban; chỉ số «khi nào đủ người» đọc được từ định biên và yêu cầu đang mở.",
        "fail": "Chỉ có tổng số hồ sơ, không gắn tháng kế hoạch / phòng ban; hoặc số liệu trộn nhiều pháp nhân.",
        "edge": "Cùng vị trí cần tuyển tháng 3 và tháng 8 — hai nhu cầu độc lập trên báo cáo.",
        "rui_ro": "Báo cáo sai tháng → quyết định tuyển muộn hoặc thừa người.",
        "gap": "PARTIAL",
    },
    {
        "uc": "UC-BP-REC-07", "mod": "REC", "task_wbs_label": "Offer → hồ sơ nhân sự",
        "req": "REQ_REC_004", "ppt": "13",
        "dien_bien": "Ứng viên chấp nhận đề nghị nhận việc → hệ thống tạo hồ sơ nhân sự ở trạng thái chờ hoàn thiện, mang theo thông tin đã có từ tuyển dụng.",
        "br": "BR-BP-LC-01",
        "pass": "Hồ sơ mới không bắt nhập lại họ tên, liên hệ, vị trí đã có trên hồ sơ ứng viên; tuyển dụng không đẩy số thẳng sang tiền lương.",
        "fail": "Nhân sự phải gõ lại toàn bộ; hoặc tuyển dụng ghi thẳng vào bảng lương.",
        "edge": "Một phần thông tin mật (lương đề nghị) chỉ mở cho chuyên viên lương thưởng sau khi nhận việc.",
        "rui_ro": "Nhập lại tay → lệch dữ liệu; nối thẳng tuyển–lương → phá ranh giới kiểm soát.",
        "gap": "PARTIAL",
    },
    {
        "uc": "UC-BP-CORE-07", "mod": "CORE", "task_wbs_label": "Kích hoạt hồ sơ Hoạt động",
        "req": "HR-003", "ppt": "13",
        "dien_bien": "Khi đủ giấy tờ bắt buộc, chuyển hồ sơ sang Hoạt động → mở quỹ phép năm và gắn ca mặc định → sẵn sàng vào kỳ lương.",
        "br": "BR-BP-LC-02",
        "pass": "Chưa đủ giấy bắt buộc thì không Hoạt động; sau Hoạt động mới mở phép/ca.",
        "fail": "Kích hoạt sớm khi còn thiếu giấy; hoặc mở chấm công khi hồ sơ còn chờ hoàn thiện.",
        "edge": "Nghỉ việc sau đó kích hoạt chuỗi cắt bảo hiểm, tất toán phép, thu hồi tài sản.",
        "rui_ro": "Hoạt động sớm → chấm công/lương sai căn cứ pháp lý nội bộ.",
        "gap": "PARTIAL",
    },
    {
        "uc": "UC-BP-ATT-12", "mod": "ATT", "task_wbs_label": "Mở quỹ phép & ca khi Hoạt động",
        "req": "REQ_NP_001; TIME-001", "ppt": "13",
        "dien_bien": "Sự kiện hồ sơ chuyển Hoạt động kéo theo mở quỹ phép theo cấu hình pháp nhân và gắn ca mặc định bộ phận.",
        "br": "BR-BP-LC-02",
        "pass": "Nhân viên mới Hoạt động thấy số phép ban đầu đúng thành phần cấu hình và ca đang gán.",
        "fail": "Có chấm công nhưng chưa có quỹ phép; hoặc ca trống sau kích hoạt.",
        "edge": "Ngày vào giữa tháng → cấp nửa ngày gốc theo quy tắc cấp dần.",
        "rui_ro": "Quên mở phép/ca → nhân viên không nộp đơn / công sai.",
        "gap": "PARTIAL",
    },
    {
        "uc": "UC-BP-PAY-06", "mod": "PAY", "task_wbs_label": "Tính lương kỳ",
        "req": "REQ_L_001", "ppt": "11",
        "dien_bien": "Chạy tính lương kỳ khi nhân sự đang Hoạt động và bảng công kỳ đã ký chốt; lấy biến từ bảng công chốt và hồ sơ lương thưởng.",
        "br": "BR-BP-LC-04",
        "pass": "Kỳ chạy thành công chỉ khi bảng công chốt và hồ sơ đủ trạng thái; kết quả khớp công thức đã phát hành.",
        "fail": "Chạy lương trên bảng công còn nháp; hoặc nhân sự chưa Hoạt động.",
        "edge": "Kỳ có gộp giữa tháng vẫn một phiếu thực nhận.",
        "rui_ro": "Chạy sớm → tranh chấp sau phát lương.",
        "gap": "PARTIAL",
    },
    {
        "uc": "UC-BP-PAY-07", "mod": "PAY", "task_wbs_label": "Tất toán nghỉ việc",
        "req": "REQ_NP_005; HR-006; HR-004; HR-005", "ppt": "13",
        "dien_bien": "Lệnh nghỉ việc kích hoạt chuỗi: cắt/đóng bảo hiểm theo ngày, trả tiền phép còn, thu hồi tài sản, đưa thưởng/phạt đang thi hành vào kỳ cuối.",
        "br": "BR-BP-LC-05",
        "pass": "Kỳ cuối phản ánh đủ các khoản liên quan; còn tài sản bắt buộc chưa thu thì chưa tất toán xong.",
        "fail": "Nghỉ xong vẫn còn tài sản «Đang sử dụng» không việc; hoặc quên thưởng/phạt kỳ cuối.",
        "edge": "Phép mang sang và phép năm mới — thứ tự trừ và đơn giá trả tiền theo chính sách đã chốt.",
        "rui_ro": "Thiếu bước → mất tài sản hoặc tranh chấp quyền lợi.",
        "gap": "PARTIAL",
    },
]


DECISIONS = [
    ("Q-REC-HEADCOUNT", "Tuyển dụng", "Trong định biên đã duyệt so với ngoài định biên: ai được bỏ qua Ban giám đốc? Vượt định biên thì cảnh báo hay chặn mở tin?", "Đề xuất: ngoài định biên bắt buộc Ban giám đốc; vượt định biên thì chặn mở tin đến khi duyệt.", "WBS-REC-02; WBS-REC-02b", "UC-BP-REC-02; UC-BP-REC-02b", "Mở — cần chốt"),
    ("Q-PAY-FORMULA", "Tiền lương", "Công thức lương: chuyên viên lương thưởng cấu hình trên màn hình, hay bộ phận kỹ thuật thiết lập trên hệ thống?", "Đề xuất: công thức có phiên bản + hai bước (chuyên viên lương thưởng soạn, kiểm soát kỹ thuật phát hành) — không cố định cứng mỗi kỳ.", "WBS-PAY-01", "UC-BP-PAY-02", "Mở — cần chốt"),
    ("Q-XBOT-PROFILE", "Nhân sự", "Nhóm thông tin hồ sơ cấu hình qua lớp danh mục tập đoàn hay hệ riêng?", "Đề xuất: dùng lớp cấu hình/danh mục tập đoàn trong phạm vi giai án.", "WBS-CORE-01", "UC-BP-CORE-02b", "Mở — cần chốt"),
    ("Q-ASSET-MODULE", "Nhân sự", "Giai đoạn này chỉ theo dõi mã/số máy + biên bản, hay mở module tài sản đầy đủ?", "Đề xuất giai đoạn 1: theo dõi mã/số máy + biên bản; giai đoạn sau mở đầy đủ nếu cần.", "WBS-CORE-04", "UC-BP-CORE-05; UC-BP-CORE-06", "Mở — cần chốt"),
    ("Q-LEAVE-ACCRUAL", "Chấm công & Nghỉ phép", "Cách cộng phép: 1 ngày/tháng + thâm niên + chức vụ — thông số mặc định theo từng pháp nhân?", "Đề xuất: từng thành phần cấu hình được, ghi tách dòng để đối chiếu.", "WBS-ATT-04", "UC-BP-ATT-04", "Mở — cần chốt"),
    ("Q-LEAVE-UNIT", "Chấm công & Nghỉ phép", "Đơn vị trừ phép tối thiểu: nửa ngày hay một giờ (theo loại phép / ca)?", "Đề xuất: cấu hình theo loại phép; một loại chỉ một đơn vị chuẩn.", "WBS-ATT-05", "UC-BP-ATT-08", "Mở — cần chốt"),
    ("Q-SI-SUSPEND", "Nhân sự", "Khi nghỉ ốm dài, bảo hiểm chuyển tạm hoãn theo ngày hiệu lực như thế nào?", "Đề xuất: gắn với đơn nghỉ dài bên chấm công / phép theo chính sách pháp nhân.", "WBS-CORE-07", "UC-BP-CORE-10", "Mở — cần chốt"),
]


def style_header(ws, row: int, ncols: int, fill: str = "1E40AF") -> None:
    f = PatternFill("solid", fgColor=fill)
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = f
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin


def style_body(ws, r1: int, r2: int, ncols: int, highlight_cols: set[int] | None = None) -> None:
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    font = Font(name="Calibri", size=10, color="0F172A")
    alt = PatternFill("solid", fgColor="F8FAFC")
    hi = PatternFill("solid", fgColor="FEF3C7")
    for r in range(r1, r2 + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin
            if highlight_cols and c in highlight_cols:
                cell.fill = hi
            elif r % 2 == 0:
                cell.fill = alt


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build() -> Path:
    matrix = parse_matrix_rows(ROOT / "UC_BR_MATRIX_DEPTH.md")
    by_uc = {r["uc"]: r for r in matrix}
    for ex in EXTRA_UC:
        by_uc.setdefault(ex["uc"], ex)

    # Ordered UC list from inventory order
    order = list(UC_TEN.keys())
    for u in by_uc:
        if u not in order:
            order.append(u)

    step_rows = build_all_steps(order, UC_TEN, WBS_FOR_UC, by_uc, ACTORS)

    wb = Workbook()

    # --- Sheet 0: Hướng dẫn ---
    ws = wb.active
    ws.title = "00_Huong_dan"
    ws["A1"] = "WBS & Nghiệp vụ HRM doanh nghiệp — bản gửi khách"
    ws["A1"].font = Font(bold=True, size=16, color="1E40AF", name="Calibri")
    lines = [
        "",
        "Mục đích: Khách đọc và chốt logic nghiệp vụ trên giấy trước khi đi vào đặc tả kỹ thuật và phát triển.",
        "Ngôn ngữ: 100% tiếng Việt — ưu tiên diễn giải đủ nghĩa, không dùng thuật ngữ lập trình.",
        "",
        "Cách đọc file:",
        "1) 01_Bon_khoi — bốn khối nghiệp vụ và câu hỏi quản trị cốt lõi.",
        "2) 02_Hang_muc_WBS — danh sách hạng mục công việc (module → hạng mục → mục đích).",
        "3) 03_Tinh_huong_nghiep_vu — từng tình huống: quy tắc, kiểm tra hợp lệ, tình huống đặc biệt (tóm tắt).",
        "4) 03b_Dien_bien_chi_tiet — BẮT BUỘC đọc khi làm việc: mỗi dòng = một bước (ai làm · người dùng làm gì · hệ thống xử lý · thành công/thất bại · bước kế). Lọc theo Module hoặc mã tình huống. Dùng cho khách chốt + SA / FE / BE / Mobile / QA / QC cùng hiểu để triển khai và kiểm thử.",
        "5) 04_Lien_ket_khoi — dữ liệu mang theo giữa các khối và điều cấm.",
        "6) 05_Can_chot_voi_khach — các quyết định còn mở (đánh dấu «Đồng ý / Cần trao đổi»).",
        "7) 99_Tham_chieu_noi_bo — mã nội bộ để đội ngũ đối chiếu khi chỉnh đặc tả (khách có thể bỏ qua).",
        "",
        "Gợi ý khi họp chốt: duyệt theo thứ tự Tuyển dụng → Nhân sự → Chấm công & phép → Tiền lương; ưu tiên các dòng «Ưu tiên gửi chốt».",
        "Khi giao Dev/QA: mở sheet 03b, lọc đúng mã tình huống, đi từng bước — cột «Gợi ý đội làm» hướng lane kỹ thuật.",
        "Phiên bản nguồn: WBS 0.4 · Inventory UC 0.3.3 · Ma trận độ sâu 1.1 · SRS 0.7 · 45 tình huống · 28 hạng mục.",
        "Phạm vi giai đoạn hiện tại: bốn khối Tuyển · Nhân sự · Chấm công/phép · Tiền lương (đã họp xong bốn trụ). Chiến dịch tuyển / hub tin đa kênh = giai đoạn 2. Loại phép: năm · thâm niên · bù tăng ca · chuyển kỳ · ứng; nghỉ ốm xét BH / hỗ trợ công ty. Bảng công chốt = đầu vào lương. Cờ cách lắp công thức lương không mang nghĩa «chưa họp lương».",
        "Quản lý công việc / dự án = module riêng, không gộp vào nhân sự. Tạm dừng viết phần mềm / demo đến khi xác nhận tài liệu trên giấy.",
        "Tài liệu này không khẳng định phần mềm đã nghiệm thu, đã ký nghiệm thu, hay đã triển khai xong.",
        "Bản chốt theo UC + màn chấm công (#1–46) + sheet chức năng sâu (18): WBS_HRM_ENTERPRISE_UC_CHOT.xlsx v1.1 · hướng dẫn WBS_UC_CHOT_README.md.",
    ]
    for i, line in enumerate(lines, 2):
        ws.cell(i, 1, line)
        ws.cell(i, 1).font = Font(name="Calibri", size=11)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[1].height = 28

    # --- Sheet 1: Bốn khối ---
    ws = wb.create_sheet("01_Bon_khoi")
    headers = ["STT", "Khối", "Tên khối", "Câu hỏi quản trị cốt lõi", "Vai trò chính tham gia", "Ghi chú liên kết"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(headers))
    blocks = [
        (1, "REC", "Tuyển dụng", "Bao giờ có người làm? Kế hoạch so với thực tế?", ACTORS["REC"], "MVP: mô tả công việc + yêu cầu tuyển (trong/ngoài định biên, tuyển mới/thay thế) + ứng viên gắn yêu cầu + báo cáo. Chiến dịch đa kênh = giai đoạn 2. Sau nhận việc → hồ sơ; không nối thẳng sang tiền lương."),
        (2, "CORE", "Nhân sự", "Hồ sơ đã đủ giấy tờ chưa? Ai được xem dữ liệu mật?", ACTORS["CORE"], "Hồ sơ công khai tách vòng hợp đồng & bảo hiểm mật. Quản lý công việc / dự án = module riêng (không thuộc khối này). Hồ sơ Hoạt động mới mở phép/ca."),
        (3, "ATT", "Chấm công & Nghỉ phép", "Giờ công đúng ca? Các loại phép đủ chưa? Phép xuyên cuối tuần/lễ trừ đúng ngày làm?", ACTORS["ATT"], "Loại phép: năm · thâm niên · bù tăng ca · chuyển kỳ · ứng. Bảng công tổng hợp đã chốt = đầu vào duy nhất để tính lương."),
        (4, "PAY", "Tiền lương & Phúc lợi", "Công thức ai cấu hình? Đổi lương giữa kỳ có một phiếu thực nhận?", ACTORS["PAY"], "Đã thống nhất trong họp: chỉ đọc bảng công đã chốt; căn cứ lương/bảo hiểm từ vòng hợp đồng–bảo hiểm mật — không lấy giờ thẳng từ tăng ca/phép."),
    ]
    for i, row in enumerate(blocks, 2):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
    style_body(ws, 2, 5, len(headers))
    set_widths(ws, [6, 10, 22, 48, 50, 48])
    ws.row_dimensions[1].height = 32
    for r in range(2, 6):
        ws.row_dimensions[r].height = 48

    # --- Sheet 2: WBS ---
    ws = wb.create_sheet("02_Hang_muc_WBS")
    h2 = [
        "STT", "Module", "Mã hạng mục", "Tên hạng mục", "Mục đích nghiệp vụ (để khách hiểu)",
        "Các tình huống gắn kèm", "Ai liên quan chính", "Độ ưu tiên gửi chốt",
        "Cần chốt với khách?", "Nội dung cần chốt (nếu có)",
        # internal last
        "Tham chiếu — mã yêu cầu đối tác", "Tham chiếu — trang trình bày", "Tham chiếu — mã tình huống",
    ]
    for c, h in enumerate(h2, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(h2))
    # tint internal headers
    for c in range(11, 14):
        ws.cell(1, c).fill = PatternFill("solid", fgColor="64748B")

    for i, t in enumerate(WBS_TASKS, 2):
        need_q = "Có" if t["q"] else "Không"
        vals = [
            i - 1,
            MODULE_VI[t["mod"]],
            t["wbs"],
            t["ten"],
            t["muc_dich"],
            t["uc"].replace("UC-BP-", ""),
            ACTORS[t["mod"]],
            t["uu_tien"],
            need_q,
            t["q"] if t["q"] else "—",
            t["req"],
            t["ppt"],
            t["uc"],
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)
    style_body(ws, 2, 1 + len(WBS_TASKS), len(h2), highlight_cols={8, 9})
    set_widths(ws, [5, 18, 14, 42, 55, 28, 40, 22, 14, 36, 28, 12, 36])
    ws.auto_filter.ref = f"A1:M{1 + len(WBS_TASKS)}"
    ws.freeze_panes = "A2"
    for r in range(2, 2 + len(WBS_TASKS)):
        ws.row_dimensions[r].height = 55

    # --- Sheet 3: UC business ---
    ws = wb.create_sheet("03_Tinh_huong_nghiep_vu")
    h3 = [
        "STT", "Module", "Mã hạng mục", "Tên tình huống nghiệp vụ",
        "Mục đích / câu hỏi giải quyết", "Ai thực hiện",
        "Diễn biến — tóm tắt + chỉ sang sheet chi tiết", "Quy tắc nghiệp vụ",
        "Kiểm tra hợp lệ & điều kiện lưu",
        "Được chấp nhận khi", "Không chấp nhận khi",
        "Tình huống đặc biệt (edge)", "Rủi ro nếu hiểu sai",
        "Độ ưu tiên gửi chốt", "Trạng thái so với mục tiêu", "Ghi chú khách (điền khi họp)",
        # internal
        "Tham chiếu — mã tình huống", "Tham chiếu — mã quy tắc",
        "Tham chiếu — mã yêu cầu đối tác", "Tham chiếu — trang trình bày",
        "Tham chiếu — mục SRS cần cập nhật",
    ]
    for c, h in enumerate(h3, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, 16, fill="1E40AF")
    for c in range(17, 22):
        ws.cell(1, c).fill = PatternFill("solid", fgColor="64748B")
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        ws.cell(1, c).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    stt = 0
    for uc in order:
        r = by_uc.get(uc)
        if not r:
            continue
        stt += 1
        mod = r["mod"]
        wbs = WBS_FOR_UC.get(uc, "")
        ten = UC_TEN.get(uc, r.get("task_wbs_label", uc))
        # Mục đích: first sentence of dien_bien or from WBS
        muc = soft_vn(r["dien_bien"].split("→")[0] if "→" in r["dien_bien"] else r["dien_bien"][:160])
        if wbs:
            for t in WBS_TASKS:
                if t["wbs"] == wbs and uc in t["uc"]:
                    muc = t["muc_dich"]
                    break
        quy = quy_tac_tu_br(r["br"], r["pass"])
        kiem = validate_tu_pass_fail(r["pass"], r["fail"], "")
        prio = "Ưu tiên gửi chốt" if uc in FR_PRIORITY else "Bổ sung sau khi chốt khung"
        srs_hint = f"SRS — FR gắn {uc}" if uc in FR_PRIORITY else f"SRS — bổ sung FR {uc} sau chốt khung"
        dien_tom_tat = (
            f"→ Mở sheet «03b_Dien_bien_chi_tiet», lọc cột «Tham chiếu — mã tình huống» = {uc}. "
            + summary_for_uc(uc, step_rows)
            + f" | Tóm tắt nghiệp vụ: {r['dien_bien']}"
        )
        row_vals = [
            stt,
            MODULE_VI.get(mod, mod),
            wbs,
            ten,
            muc,
            ACTORS.get(mod, ""),
            dien_tom_tat,
            quy,
            kiem,
            r["pass"],
            r["fail"],
            r["edge"],
            r["rui_ro"],
            prio,
            gap_vi(r.get("gap", "")),
            "",
            uc,
            r["br"],
            r["req"],
            r["ppt"],
            srs_hint,
        ]
        for c, v in enumerate(row_vals, 1):
            ws.cell(1 + stt, c, v)

    last = 1 + stt
    style_body(ws, 2, last, 16, highlight_cols={14})
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    for r in range(2, last + 1):
        for c in range(17, 22):
            cell = ws.cell(r, c)
            cell.border = thin
            cell.font = Font(name="Calibri", size=9, color="475569")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F1F5F9")

    set_widths(ws, [5, 18, 14, 40, 36, 36, 48, 48, 48, 40, 40, 36, 32, 20, 28, 24, 16, 14, 22, 10, 28])
    ws.auto_filter.ref = f"A1:U{last}"
    ws.freeze_panes = "D2"
    for r in range(2, last + 1):
        ws.row_dimensions[r].height = 88
    ws.row_dimensions[1].height = 40

    # --- Sheet 3b: Diễn biến từng bước (chi tiết làm việc) ---
    ws = wb.create_sheet("03b_Dien_bien_chi_tiet")
    h3b = [
        "STT", "Module", "Mã hạng mục", "Tên tình huống",
        "Bước số", "Loại bước", "Ai thực hiện",
        "Người dùng / người nghiệp vụ làm gì (chi tiết)",
        "Hệ thống xử lý thế nào (chi tiết)",
        "Điều kiện / quy tắc tại bước này",
        "Thành công — thấy gì (màn hình / dữ liệu)",
        "Thất bại — chặn / thông báo gì",
        "Bước hoặc tình huống tiếp theo",
        "Gợi ý đội làm (SA · FE · BE · Mobile · QA · QC)",
        "Tham chiếu — mã tình huống", "Tham chiếu — mã quy tắc", "Nguồn diễn giải",
    ]
    for c, h in enumerate(h3b, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, 14, fill="0F766E")
    for c in range(15, 18):
        ws.cell(1, c).fill = PatternFill("solid", fgColor="64748B")
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        ws.cell(1, c).alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for i, sr in enumerate(step_rows, 2):
        vals = [
            sr["stt"], sr["module"], sr["wbs"], sr["ten"],
            sr["buoc"], sr["loai"], sr["ai"],
            sr["user_do"], sr["sys_do"], sr["dieu_kien"],
            sr["thanh_cong"], sr["that_bai"], sr["buoc_ke"],
            sr["lane"], sr["uc"], sr["br"], sr["nguon"],
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)
    last_b = 1 + len(step_rows)
    style_body(ws, 2, last_b, 14)
    thin_b = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    for r in range(2, last_b + 1):
        for c in range(15, 18):
            cell = ws.cell(r, c)
            cell.border = thin_b
            cell.font = Font(name="Calibri", size=9, color="475569")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F1F5F9")
        ws.row_dimensions[r].height = 78
    set_widths(ws, [6, 16, 12, 36, 8, 18, 28, 48, 48, 36, 36, 32, 28, 36, 16, 12, 28])
    ws.auto_filter.ref = f"A1:Q{last_b}"
    ws.freeze_panes = "E2"
    ws.row_dimensions[1].height = 44

    # --- Sheet 4: Liên kết ---
    ws = wb.create_sheet("04_Lien_ket_khoi")
    h4 = ["STT", "Từ khối", "Sang khối", "Dữ liệu mang theo", "Điều cấm / không làm", "Vì sao quan trọng"]
    for c, h in enumerate(h4, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(h4))
    links = [
        (1, "Tuyển dụng", "Nhân sự", "Ứng viên chấp nhận đề nghị → hồ sơ nhân sự mới với thông tin đã có", "Không bắt nhập lại các trường đã có trên hồ sơ ứng viên", "Tránh lệch dữ liệu và chậm nhận việc"),
        (2, "Nhân sự", "Chấm công & Nghỉ phép", "Hồ sơ ở trạng thái Hoạt động → mở phép năm và gắn ca mặc định", "Không mở chấm công khi hồ sơ còn chờ hoàn thiện; không gộp quản lý công việc vào nhân sự", "Chỉ người đủ giấy tờ mới vào vận hành giờ công"),
        (3, "Chấm công & Nghỉ phép", "Tiền lương", "Bảng công tổng hợp đã ký chốt (chấm + phép + tăng ca) — đầu vào tính lương", "Tiền lương không lấy giờ thẳng từ tăng ca hay đơn phép", "Một nguồn số duy nhất — tránh lệch phiếu lương"),
        (4, "Nhân sự", "Tiền lương", "Lương cơ bản, phụ cấp, bảo hiểm, giảm trừ gia cảnh từ vòng hợp đồng–bảo hiểm mật; nghỉ việc; thưởng/phạt đang thi hành", "Không lấy lương/mã số thuế/ngân hàng từ hồ sơ công khai", "Căn cứ tính lương và tất toán"),
        (5, "Tuyển dụng", "Tiền lương", "— (không trao đổi trực tiếp)", "Cấm kết nối trực tiếp tuyển dụng với tiền lương", "Kiểm soát tách bạch nhu cầu tuyển và chính sách lương"),
    ]
    for i, row in enumerate(links, 2):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)
    style_body(ws, 2, 6, len(h4))
    set_widths(ws, [5, 22, 22, 55, 48, 40])
    for r in range(2, 7):
        ws.row_dimensions[r].height = 48

    # --- Sheet 5: Decisions ---
    ws = wb.create_sheet("05_Can_chot_voi_khach")
    h5 = [
        "STT", "Mã quyết định", "Module", "Câu hỏi cần chốt", "Phương án đề xuất",
        "Hạng mục liên quan", "Tình huống liên quan", "Trạng thái",
        "Ý kiến khách (Đồng ý / Điều chỉnh / Từ chối)", "Ghi chú họp",
    ]
    for c, h in enumerate(h5, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(h5), fill="B45309")
    for i, d in enumerate(DECISIONS, 2):
        vals = [i - 1, d[0], d[1], d[2], d[3], d[4], d[5], d[6], "", ""]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)
    style_body(ws, 2, 1 + len(DECISIONS), len(h5), highlight_cols={9})
    set_widths(ws, [5, 18, 20, 55, 55, 28, 28, 16, 28, 30])
    ws.freeze_panes = "A2"
    for r in range(2, 2 + len(DECISIONS)):
        ws.row_dimensions[r].height = 60
    dv = DataValidation(type="list", formula1='"Đồng ý,Điều chỉnh,Từ chối,Cần trao đổi thêm"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"I2:I{1 + len(DECISIONS)}")

    # --- Sheet 6: internal index ---
    ws = wb.create_sheet("99_Tham_chieu_noi_bo")
    ws["A1"] = "Bảng tham chiếu nội bộ — dùng khi chỉnh SRS / ma trận; khách không bắt buộc đọc"
    ws["A1"].font = Font(bold=True, color="64748B", size=12)
    h6 = ["Mã tình huống", "Mã hạng mục", "Mã quy tắc", "Mã yêu cầu đối tác", "Trang trình bày", "File nguồn", "Ghi chú cập nhật SRS"]
    for c, h in enumerate(h6, 1):
        ws.cell(2, c, h)
    style_header(ws, 2, len(h6), fill="64748B")
    rr = 3
    for uc in order:
        r = by_uc.get(uc)
        if not r:
            continue
        ws.cell(rr, 1, uc)
        ws.cell(rr, 2, WBS_FOR_UC.get(uc, ""))
        ws.cell(rr, 3, r["br"])
        ws.cell(rr, 4, r["req"])
        ws.cell(rr, 5, r["ppt"])
        ws.cell(rr, 6, "UC_BR_MATRIX_DEPTH.md · WBS_HRM_ENTERPRISE.md · UC_INVENTORY.md")
        ws.cell(rr, 7, f"Khi sửa nghiệp vụ dòng này → cập nhật FR trong SRS_HRM_ENTERPRISE.md gắn {uc}")
        rr += 1
    style_body(ws, 3, rr - 1, len(h6))
    set_widths(ws, [18, 14, 14, 28, 12, 55, 55])
    ws.freeze_panes = "A3"

    # Cover note row count check
    assert stt == 45, f"Expected 45 UC, got {stt}"

    try:
        wb.save(OUT)
        return OUT
    except PermissionError:
        wb.save(OUT_FALLBACK)
        return OUT_FALLBACK


if __name__ == "__main__":
    path = build()
    print(f"OK: {path}")
    print(f"size_kb: {path.stat().st_size // 1024}")
    if path.name == "WBS_HRM_ENTERPRISE_KHACH.xlsx":
        print("NOTE: file MOI dang mo — da ghi ban KHACH.xlsx; dong Excel MOI roi chay lai de ghi de MOI.")
