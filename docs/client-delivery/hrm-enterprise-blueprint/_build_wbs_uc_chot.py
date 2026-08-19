# -*- coding: utf-8 -*-
"""WBS chốt họp theo UC — khách v1.1 (gap matrix v1.1 + ATT-DEEP 18 MISSING + browser QA)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "WBS_HRM_ENTERPRISE_UC_CHOT.xlsx"
OUT_FALLBACK = ROOT / "WBS_HRM_ENTERPRISE_UC_CHOT_ALT.xlsx"
VERSION = "1.1"

MODULE_VI = {
    "REC": "Tuyển dụng",
    "CORE": "Nhân sự",
    "ATT": "Chấm công & Nghỉ phép",
    "PAY": "Tiền lương & Phúc lợi",
}

GAP_VI = {
    "COVERED": "Đã phủ đủ để chốt giấy",
    "SRS_THIN": "Mô tả chức năng còn khung — cần bổ sung chi tiết sau chốt",
    "PRODUCT_STUB": "Màn hình đang phát triển",
    "PRODUCT_MISSING": "Chưa có màn hình tương ứng / chưa liệt kê trên bản đồ màn",
    "MEETING_ONLY_GĐ2": "Ngoài phạm vi hiện tại (giai đoạn 2)",
    "UNMAPPED_PRODUCT": "Có trên phần mềm — cần chốt có/không thuộc phạm vi",
    "SPEC_GAP": "Còn thiếu mô tả chấp nhận đo được",
    "ALIAS_SAME_UI": "Nhãn menu khác nhưng cùng một màn nghỉ phép — cần chốt trung thực",
}

# Browser runtime overlay — PO-HRM-BP-ATT-DEEP-QA-01 (fid #1–46)
ATT_BROWSER: dict[int, str] = {
    1: "LIVE", 2: "LIVE", 3: "LIVE", 4: "LIVE", 5: "LIVE",
    6: "LIVE", 7: "LIVE", 8: "PARTIAL", 9: "GĐ2-HOLD", 10: "LIVE",
    11: "LIVE", 12: "LIVE", 13: "LIVE", 14: "LIVE", 15: "LIVE",
    16: "LIVE", 17: "STUB_UI", 18: "STUB_UI",
    19: "LIVE", 20: "LIVE", 21: "LIVE", 22: "LIVE", 23: "LIVE", 24: "LIVE",
    25: "LIVE", 26: "LIVE", 27: "LIVE",
    28: "LIVE", 29: "LIVE", 30: "PARTIAL",
    31: "LIVE", 32: "LIVE", 33: "PARTIAL", 34: "LIVE", 35: "LIVE", 36: "LIVE",
    37: "STUB_UI", 38: "STUB_UI", 39: "STUB_UI",
    40: "STUB_UI", 41: "STUB_UI", 42: "STUB_UI", 43: "STUB_UI",
    44: "STUB_UI", 45: "STUB_UI", 46: "STUB_UI",
}

ATT_INV_PARENT: dict[int, str] = {
    1: "S01", 2: "S05", 3: "S06", 4: "S08", 5: "S09", 6: "S10", 7: "S11",
    8: "S13", 9: "S17", 10: "S20", 11: "S23", 12: "S24", 13: "S26", 14: "S31",
    15: "S34", 16: "S35", 17: "S40", 18: "S41", 19: "S42", 20: "S48", 21: "S50",
    22: "S52", 23: "S54", 24: "S56", 25: "S58", 26: "S59", 27: "S60", 28: "S61",
    29: "S62", 30: "S63", 31: "S64", 32: "S67", 33: "S68", 34: "S69", 35: "S72",
    36: "S73", 37: "S76", 38: "S77", 39: "S78", 40: "S79", 41: "S80", 42: "S81",
    43: "S82", 44: "S83", 45: "S84", 46: "S85",
}

# ATT-DEEP §6.1 — 18 MISSING (gap_class per matrix §6.0) + alias #25–27 honesty
# inv, menu, meaning, parent_fid, browser_note, gap, mvp, decision, proposed_uc, wbs
ATT_DEEP_ROWS: list[tuple] = [
    ("S03", "Tổng quan → Chấm công ngay", "Nút mở chấm từ tổng quan", "ATT-FID#1/#6",
     "Overview LIVE (CTA không probe riêng)", "UNMAPPED_PRODUCT", "MVP",
     "Giữ nút hay gộp vào vào/ra?", "Note #1/#6", "WBS-ATT-02"),
    ("S04", "Tổng quan → Tùy chỉnh layout", "Nút tùy chỉnh bố cục (đang khóa)", "ATT-FID#1",
     "Overview LIVE shell · CTA khóa (STUB)", "PRODUCT_STUB", "GĐ2",
     "Ngoài phạm vi hiện tại?", "—", "—"),
    ("S07", "Tổng quan → Phân tích loại nghỉ (biểu đồ tròn)", "Biểu đồ loại nghỉ", "—",
     "Overview LIVE (panel không probe riêng)", "SPEC_GAP", "MVP",
     "Thuộc phạm vi hiện tại / giai đoạn 2?", "ACCEPTED hoặc FR GĐ2", "WBS-ATT-04"),
    ("S15", "Clock-In → QR → Thẻ QR nhân viên", "Chọn NV · xem/tải/in mã QR", "ATT-FID#8",
     "QA chưa mở dialog (QR shell PARTIAL) — residual", "UNMAPPED_PRODUCT", "MVP",
     "Thẻ QR nhân viên bắt buộc giai đoạn hiện tại?", "UC-BP-ATT-03e", "WBS-ATT-02"),
    ("S16", "Clock-In → QR → Hộp thoại phóng to QR", "Phóng to mã QR", "ATT-FID#8",
     "QA chưa mở — residual", "UNMAPPED_PRODUCT", "MVP",
     "—", "nested S15", "WBS-ATT-02"),
    ("S25", "Bảng chấm công → Xóa bảng", "Hộp thoại xác nhận xóa kỳ", "ATT-FID#11",
     "QA RO chưa mở Alert xóa — residual", "UNMAPPED_PRODUCT", "MVP",
     "Ai được xóa bảng đã tạo?", "UC-BP-ATT-11b", "WBS-ATT-06"),
    ("S28", "Bản ghi → Xóa bản ghi", "Hộp thoại xác nhận xóa dòng", "ATT-FID#13",
     "QA RO chưa mở Alert xóa — residual", "UNMAPPED_PRODUCT", "MVP",
     "—", "Expand #13", "WBS-ATT-02"),
    ("S29", "Bản ghi → Xuất tệp", "Xuất từ sổ bản ghi", "ATT-FID#13/#30",
     "Xuất báo cáo #30 PARTIAL đã mở; path sổ ≠ báo cáo", "SPEC_GAP", "MVP",
     "Xuất từ sổ hay chỉ từ báo cáo?", "Gộp #30", "—"),
    ("S32", "Chấm công tuần → Chi tiết ô", "Hộp thoại xem/sửa ô tuần", "ATT-FID#14",
     "Tuần LIVE; ô chi tiết chưa mở", "UNMAPPED_PRODUCT", "MVP",
     "Thuộc phạm vi lưới tuần?", "UC-BP-ATT-18", "WBS-ATT-01"),
    ("S33", "Chấm công tuần → Biểu tượng bút/cài/tải", "Nút chưa gắn hành động", "ATT-FID#14",
     "Tuần LIVE · icon no-op (STUB)", "PRODUCT_STUB", "MVP",
     "Ẩn nút trống hay hoàn thiện?", "—", "—"),
    ("S39", "Ca → Sao chép", "Nút sao chép ca (chưa gắn hành động)", "ATT-FID#16",
     "Danh sách ca LIVE · copy chưa bấm", "PRODUCT_STUB", "MVP",
     "—", "—", "WBS-ATT-01"),
    ("S43", "Nghỉ phép → Quỹ phép (số dư)", "Panel xem số dư quỹ phép", "ATT-FID#19/#28",
     "Tab nghỉ LIVE; panel quỹ chưa stamp riêng", "SRS_THIN", "MVP",
     "Bắt buộc hiển thị quỹ khi nộp đơn?", "UC-BP-ATT-05b", "WBS-ATT-04"),
    ("S65", "Cài đặt → NV → Nhập khẩu", "Hộp thoại nhập danh sách NV chấm", "ATT-FID#31",
     "Settings NV LIVE; Import chưa mở lại", "UNMAPPED_PRODUCT", "MVP",
     "—", "Expand #31", "—"),
    ("S66", "Cài đặt → NV → Lọc / Tải xuống (icon)", "Nút chưa gắn hành động", "ATT-FID#31",
     "STUB candidate (no-op)", "PRODUCT_STUB", "MVP",
     "Ẩn hay hoàn thiện?", "—", "—"),
    ("S70", "Quy tắc → Tùy chỉnh → Reset/Xem trước/Thêm…", "Nút chưa gắn hành động", "ATT-FID#34",
     "Rules tùy chỉnh LIVE tĩnh · CTA no-op", "PRODUCT_STUB", "GĐ2",
     "—", "—", "—"),
    ("S71", "Quy tắc → Gợi ý phương thức chấm", "Nút chưa gắn hành động", "ATT-FID#34/#36",
     "STUB candidate (no-op)", "PRODUCT_STUB", "MVP",
     "—", "—", "—"),
    ("S74", "Ứng dụng → Danh sách địa điểm GPS", "Danh sách vùng chấm theo vị trí", "— (không thuộc Thiết bị)",
     "App rules #36 LIVE; dialog thêm chưa mở (CFG)", "UNMAPPED_PRODUCT", "MVP",
     "Cấu hình vùng GPS bắt buộc giai đoạn hiện tại?", "UC-BP-ATT-03d", "WBS-ATT-02"),
    ("S75", "Ứng dụng → Thêm địa điểm GPS", "Thêm tên / tọa độ / bán kính", "—",
     "QA chưa mở dialog thêm — residual", "UNMAPPED_PRODUCT", "MVP",
     "—", "nested 03d", "WBS-ATT-02"),
    # Alias honesty #25–27
    ("S58-ALIAS", "Đơn từ → Tổng hợp nghỉ", "Nhãn menu — cùng màn nghỉ phép", "ATT-FID#25",
     "LIVE (ALIAS LeaveTab)", "ALIAS_SAME_UI", "MVP",
     "Giữ nhãn hay gộp một mục «Nghỉ phép»?", "ACCEPTED", "WBS-ATT-04"),
    ("S59-ALIAS", "Đơn từ → Tổng hợp nghỉ bù", "Nhãn menu — cùng màn nghỉ phép", "ATT-FID#26",
     "LIVE (ALIAS LeaveTab)", "ALIAS_SAME_UI", "MVP",
     "Giữ nhãn hay gộp?", "ATT-06", "WBS-ATT-04"),
    ("S60-ALIAS", "Đơn từ → Kế hoạch nghỉ", "Nhãn menu — cùng màn nghỉ phép · họp GĐ2", "ATT-FID#27",
     "LIVE wire · họp = GĐ2", "ALIAS_SAME_UI", "GĐ2",
     "Ẩn mục kế hoạch nghỉ ở giai đoạn hiện tại?", "—", "—"),
]

# uc, ten, pillar, gd, meeting, srs_status, wbs, product_rt, gap, decision, prio, note
UC_ROWS: list[tuple[str, ...]] = [
    ("UC-BP-REC-00", "Thư viện mô tả công việc (bản gốc chuẩn)", "REC", "MVP", "D2 · R6", "Khung", "WBS-REC-00", "Đang vận hành", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Đầu vào yêu cầu tuyển"),
    ("UC-BP-REC-01", "Quản trị định biên 12 tháng (chỉ số cần tuyển)", "REC", "MVP", "D4 · R5", "Đủ 7 mục", "WBS-REC-01", "Đang vận hành", "COVERED", "Chốt quyền duyệt trong/ngoài định biên", "Ưu tiên gửi chốt", "Lưới tháng chỉ «Cần tuyển»"),
    ("UC-BP-REC-01b", "Tự tạo yêu cầu tuyển theo tháng «Cần tuyển»", "REC", "MVP", "D2 · R5", "Đủ 7 mục", "WBS-REC-01b", "Đang phát triển một phần", "PRODUCT_STUB", "Chốt lịch tự tạo yêu cầu", "Ưu tiên gửi chốt", "Mỗi ô Cần tuyển đã duyệt → một yêu cầu"),
    ("UC-BP-REC-02", "Yêu cầu tuyển trong định biên (tuyển mới / thay thế)", "REC", "MVP", "D3 · R2", "Đủ 7 mục", "WBS-REC-02", "Đang vận hành", "COVERED", "Chốt quyền duyệt trong định biên", "Ưu tiên gửi chốt", "Luồng duyệt rút gọn"),
    ("UC-BP-REC-02b", "Yêu cầu tuyển ngoài định biên (tuyển mới / thay thế)", "REC", "MVP", "D3 · R2", "Đủ 7 mục", "WBS-REC-02b", "Đang vận hành", "COVERED", "Chốt bắt buộc Ban giám đốc ngoài định biên", "Ưu tiên gửi chốt", "Nhánh duyệt chặt hơn"),
    ("UC-BP-REC-03", "Chiến dịch tuyển / hub tin đa kênh", "REC", "GĐ2", "D1 · R1", "GĐ2", "WBS-REC-02c", "Menu còn (không dùng MVP)", "MEETING_ONLY_GĐ2", "Ẩn hay giữ menu chiến dịch ở giai đoạn hiện tại", "Giai đoạn 2", "Chỉ khi có đối tác đăng tin"),
    ("UC-BP-REC-04", "Quét kho ứng viên nội bộ trước kênh ngoài", "REC", "MVP", "R3 · R6", "Khung", "WBS-REC-03", "Đang vận hành", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Ưu tiên kho nội bộ"),
    ("UC-BP-REC-05", "Lịch sử trạng thái ứng viên gắn yêu cầu tuyển", "REC", "MVP", "D2 · R3", "Khung", "WBS-REC-03", "Đang vận hành một phần", "SRS_THIN", "Ứng viên bắt buộc gắn yêu cầu tuyển", "Bổ sung sau chốt khung", "Trạng thái theo từng yêu cầu"),
    ("UC-BP-REC-06", "Gửi thư tuyển và đánh giá phỏng vấn", "REC", "MVP", "R6", "Khung", "WBS-REC-04", "Đang vận hành một phần", "SRS_THIN", "Bắt buộc gửi kèm người phỏng vấn?", "Bổ sung sau chốt khung", "Trong pipeline gắn yêu cầu"),
    ("UC-BP-REC-07", "Chấp nhận đề nghị → tạo hồ sơ nhân sự", "REC", "MVP", "R3", "Khung", "WBS-REC-05", "Đang phát triển một phần", "PRODUCT_STUB", "—", "Bổ sung sau chốt khung", "Không nhập lại; không nối thẳng lương"),
    ("UC-BP-REC-08", "Báo cáo tuyển — kế hoạch so với thực tế", "REC", "MVP", "R4", "Đủ 7 mục", "WBS-REC-06", "Đang vận hành", "COVERED", "Đủ chiều thời gian × phòng × cấp?", "Ưu tiên gửi chốt", "«Bao giờ đủ người»"),
    ("UC-BP-CORE-01", "Hồ sơ vòng công khai (hành chính / phúc lợi)", "CORE", "MVP", "C1", "Đủ 7 mục", "WBS-CORE-01", "Đang vận hành", "COVERED", "—", "Ưu tiên gửi chốt", "Không hiện lương / MST / ngân hàng / BH mật"),
    ("UC-BP-CORE-02", "Hồ sơ vòng hợp đồng & bảo hiểm mật", "CORE", "MVP", "C2 · D5 · P2", "Đủ 7 mục", "WBS-CORE-01", "Đang vận hành", "COVERED", "—", "Ưu tiên gửi chốt", "Chỉ đúng quyền xem"),
    ("UC-BP-CORE-02b", "Cấu hình nhóm thông tin trên hồ sơ", "CORE", "MVP", "C1", "Khung", "WBS-CORE-01", "Đang vận hành một phần", "SRS_THIN", "Cấu hình danh mục tập đoàn hay riêng?", "Bổ sung sau chốt khung", "Vẫn tách công khai / mật"),
    ("UC-BP-CORE-03", "Danh mục giấy tờ động", "CORE", "MVP", "C9", "Khung", "WBS-CORE-03", "Đang phát triển một phần", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Bắt buộc / tùy chọn theo vị trí"),
    ("UC-BP-CORE-04", "Đọc giấy tờ tự động — điền sẵn", "CORE", "MVP", "C9", "Khung", "WBS-CORE-03", "Chưa có màn rõ", "PRODUCT_MISSING", "Giai đoạn hiện tại có đọc giấy tờ tự động không?", "Bổ sung sau chốt khung", "Không nhập lại trường đã có"),
    ("UC-BP-CORE-05", "Cấp phát tài sản và biên bản bàn giao", "CORE", "MVP", "C6", "Khung", "WBS-CORE-04", "Đang vận hành", "SRS_THIN", "Chỉ theo dõi mã máy hay module tài sản đầy đủ?", "Bổ sung sau chốt khung", "Xác nhận hai bên"),
    ("UC-BP-CORE-06", "Thu hồi tài sản khi nghỉ việc", "CORE", "MVP", "C6 · C8", "Khung", "WBS-CORE-04", "Đang phát triển một phần", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Đủ thu hồi trước tất toán"),
    ("UC-BP-CORE-07", "Chuyển hồ sơ sang «Hoạt động» khi đủ giấy tờ", "CORE", "MVP", "C9", "Khung", "WBS-CORE-05", "Đang phát triển một phần", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Mở phép + ca"),
    ("UC-BP-CORE-08", "Khen thưởng & kỷ luật — thi hành vào bảng lương", "CORE", "MVP", "C5 · P3", "Đủ 7 mục", "WBS-CORE-06", "Đang vận hành", "COVERED", "Ánh xạ đúng kỳ lương đích", "Ưu tiên gửi chốt", "Đã thi hành → kỳ đích"),
    ("UC-BP-CORE-09", "Hợp đồng lao động — mẫu in điền sẵn", "CORE", "MVP", "C2", "Khung", "WBS-CORE-02", "Đang vận hành", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Nuôi dữ liệu lương"),
    ("UC-BP-CORE-10", "Bảo hiểm xã hội theo vòng đời", "CORE", "MVP", "C4 · A4", "Khung", "WBS-CORE-07", "Đang vận hành", "SRS_THIN", "Tạm hoãn BH khi nghỉ ốm dài?", "Bổ sung sau chốt khung", "Đóng / ngừng / tạm hoãn"),
    ("UC-BP-ATT-01", "Thiết lập quy tắc ca theo bộ phận / nhóm", "ATT", "MVP", "A1 · A5", "Khung", "WBS-ATT-01", "Danh sách ca OK · phân ca đang phát triển", "SRS_THIN", "Rule theo bộ phận đủ chưa?", "Bổ sung sau chốt khung", "Không một quy tắc chung ghi đè"),
    ("UC-BP-ATT-02", "Phạt muộn / về sớm (phút / khối / bậc)", "ATT", "MVP", "A1 · A5", "Đủ 7 mục", "WBS-ATT-01", "Quy tắc chung OK · cấu hình muộn chuyên biệt đang phát triển", "PRODUCT_STUB", "Chế độ phạt: phút / khối / bậc", "Ưu tiên gửi chốt", "Chỉ khi điểm danh hợp lệ"),
    ("UC-BP-ATT-03", "Thu nhận điểm danh nhiều nguồn → giờ công thô", "ATT", "MVP", "A6", "Khung", "WBS-ATT-02", "Thủ công/GPS OK · QR một phần · khuôn mặt GĐ2", "SRS_THIN", "Kênh bắt buộc giai đoạn hiện tại?", "Bổ sung sau chốt khung", "Ứng dụng · mạng · vị trí · máy"),
    ("UC-BP-ATT-03b", "Lịch lễ / Tết (dương và âm theo năm)", "ATT", "MVP", "A5", "Khung", "WBS-ATT-03", "Chưa có màn riêng", "PRODUCT_MISSING", "Cần màn cấu hình lễ âm/dương?", "Bổ sung sau chốt khung", "Dùng chung phép và bảng công"),
    ("UC-BP-ATT-04", "Cấp phát phép năm + danh mục loại phép", "ATT", "MVP", "A3", "Khung", "WBS-ATT-04", "Cấu hình quy tắc nghỉ đang phát triển", "PRODUCT_STUB", "Đủ 5 loại: năm · thâm niên · bù tăng ca · chuyển kỳ · ứng?", "Ưu tiên gửi chốt", "Khóa danh mục loại phép đã họp"),
    ("UC-BP-ATT-04b", "Ứng phép và thời điểm cấp / nghỉ không lương", "ATT", "MVP", "A3", "Khung", "WBS-ATT-04", "Cấu hình đang phát triển · đơn nghỉ vận hành", "SRS_THIN", "Bật/tắt ứng phép", "Bổ sung sau chốt khung", "Trần ứng; bù trừ khi có quỹ"),
    ("UC-BP-ATT-05", "Bảo lưu phép năm cũ đến hết quý 1", "ATT", "MVP", "A3", "Khung", "WBS-ATT-04", "Đang phát triển một phần", "SRS_THIN", "Cắt hết sau quý 1?", "Bổ sung sau chốt khung", "Chính sách mang sang"),
    ("UC-BP-ATT-06", "Phép nghỉ bù từ tăng ca (khi công ty bật)", "ATT", "MVP", "A3", "Khung", "WBS-ATT-04", "Tổng hợp nghỉ bù (dây chuyền đơn nghỉ)", "SRS_THIN", "Bật phép bù tăng ca?", "Bổ sung sau chốt khung", "Chỉ khi tăng ca đã duyệt"),
    ("UC-BP-ATT-07", "Nghỉ ốm — bảo hiểm hoặc hỗ trợ công ty", "ATT", "MVP", "A4", "Khung", "WBS-ATT-04", "Đơn nghỉ vận hành", "SRS_THIN", "Nhánh BHXH vs hỗ trợ công ty", "Bổ sung sau chốt khung", "Không chồng hai chế độ không quy tắc"),
    ("UC-BP-ATT-08", "Tính ngày trừ phép xuyên cuối tuần và lễ", "ATT", "MVP", "A3 · A5", "Đủ 7 mục", "WBS-ATT-05", "Đang vận hành", "COVERED", "Đơn vị trừ: nửa ngày hay một giờ?", "Ưu tiên gửi chốt", "Chỉ trừ ngày làm"),
    ("UC-BP-ATT-09", "Nộp & duyệt phép — giữ chỗ quỹ khi gửi đơn", "ATT", "MVP", "A5", "Đủ 7 mục", "WBS-ATT-05", "Đang vận hành", "COVERED", "Giữ chỗ khi gửi đơn", "Ưu tiên gửi chốt", "Từ chối hoàn đủ; duyệt = trừ thật"),
    ("UC-BP-ATT-10", "Tổng hợp bảng công (đầu vào tính lương)", "ATT", "MVP", "A2 · D8 · P1", "Đủ 7 mục", "WBS-ATT-06", "Đang vận hành (tổng hợp ≈ sổ bản ghi)", "SPEC_GAP", "Tổng hợp riêng hay chấp nhận gộp từ sổ bản ghi?", "Ưu tiên gửi chốt", "Gộp chấm + phép + tăng ca"),
    ("UC-BP-ATT-11", "Ký chốt bảng công trước khi tính lương", "ATT", "MVP", "A2 · D8 · P1", "Đủ 7 mục", "WBS-ATT-06", "Đang vận hành", "COVERED", "Ai ký chốt?", "Ưu tiên gửi chốt", "Đủ chữ ký mới mở tính lương"),
    ("UC-BP-ATT-12", "Mở quỹ phép và ca mặc định khi hồ sơ Hoạt động", "ATT", "MVP", "A5", "Khung", "WBS-CORE-05", "Đang phát triển một phần", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Kéo theo kích hoạt hồ sơ"),
    ("UC-BP-PAY-01", "Lương chỉ lấy số từ bảng công đã chốt", "PAY", "MVP", "P1 · D8", "Đủ 7 mục", "WBS-ATT-06", "Màn dữ liệu lương–chấm đang phát triển", "PRODUCT_STUB", "Chỉ đọc bảng công đã chốt", "Ưu tiên gửi chốt", "Không lấy giờ thẳng từ tăng ca/phép"),
    ("UC-BP-PAY-02", "Lắp ráp và chạy công thức lương cấu hình được", "PAY", "MVP", "P4 · P5", "Đủ 7 mục", "WBS-PAY-01", "Màn tính lương / chính sách đang phát triển", "PRODUCT_STUB", "Chốt cách lắp công thức trên màn hình (Q-PAY-FORMULA)", "Ưu tiên gửi chốt", "Họp lương đã xong — còn chốt cách lắp"),
    ("UC-BP-PAY-03", "Giảm trừ gia cảnh từ hồ sơ (đủ quyền)", "PAY", "MVP", "P2", "Khung", "WBS-PAY-02", "Đang phát triển một phần", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Không nhập trùng trên bảng lương"),
    ("UC-BP-PAY-04", "Gộp lương khi đổi điều kiện giữa tháng", "PAY", "MVP", "P6", "Đủ 7 mục", "WBS-PAY-03", "Chưa thấy màn riêng", "PRODUCT_MISSING", "—", "Ưu tiên gửi chốt", "Một phiếu thực nhận"),
    ("UC-BP-PAY-05", "Trần bảo hiểm trên tổng hợp kỳ", "PAY", "MVP", "P2 · P6", "Khung", "WBS-PAY-02", "Chưa rõ màn", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Kể cả gộp giữa tháng"),
    ("UC-BP-PAY-06", "Tính lương kỳ khi đã Hoạt động + bảng công chốt", "PAY", "MVP", "P1 · P5", "Khung", "WBS-PAY-04", "Tạo bảng lương đang phát triển", "PRODUCT_STUB", "—", "Bổ sung sau chốt khung", "Điều kiện mở kỳ rõ ràng"),
    ("UC-BP-PAY-07", "Tất toán nghỉ việc", "PAY", "MVP", "P3 · P6 · C8", "Khung", "WBS-PAY-04", "Đang phát triển một phần", "SRS_THIN", "Tách nghỉ tự nguyện vs buộc thôi việc?", "Bổ sung sau chốt khung", "BH · phép · tài sản · thưởng/phạt kỳ cuối"),
    ("UC-BP-PAY-08", "Phiếu lương — xem trước, bảo mật, thanh toán", "PAY", "MVP", "P5", "Khung", "WBS-PAY-05", "Đang phát triển một phần", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Nhân viên chỉ xem phiếu của mình"),
    ("UC-BP-PAY-09", "Phân nhóm bảng lương (VP / KD / tài xế / vận hành)", "PAY", "MVP", "P5", "Khung", "WBS-PAY-05", "Chưa rõ màn", "SRS_THIN", "—", "Bổ sung sau chốt khung", "Mỗi người một nhóm đang áp dụng"),
]

# #, menu, meaning, status_vi, gd, gap, decision, note
ATT_SURFACES: list[tuple] = [
    (1, "Tổng quan", "Bảng điều khiển đi làm / muộn / nghỉ", "Đang vận hành", "MVP", "UNMAPPED_PRODUCT", "Dashboard bắt buộc giai đoạn hiện tại?", "Theo dõi điều hành"),
    (2, "Tổng quan → Biểu đồ nghỉ tháng", "Xu hướng nghỉ theo tháng", "Đang vận hành", "MVP", "SPEC_GAP", "Giữ như hiện tại hay mô tả chức năng riêng?", "Hỗ trợ cân nhân sự"),
    (3, "Tổng quan → Nghỉ theo phòng ban", "So sánh nghỉ giữa phòng", "Đang vận hành", "MVP", "SPEC_GAP", "Giữ như hiện tại hay mô tả chức năng riêng?", "Phân bổ nguồn lực"),
    (4, "Tổng quan → Danh sách muộn/sớm", "Cảnh báo muộn / về sớm", "Đang vận hành", "MVP", "UNMAPPED_PRODUCT", "—", "Liên quan kỷ luật"),
    (5, "Tổng quan → Đơn nghỉ gần đây", "Đơn nghỉ gần đây", "Đang vận hành", "MVP", "COVERED", "—", "Cầu nối module nghỉ"),
    (6, "Chấm công → Vào/ra", "Trung tâm chọn cách chấm", "Đang vận hành", "MVP", "COVERED", "—", "Nguồn giờ công thô"),
    (7, "Chấm công → Thủ công", "Chấm tay", "Đang vận hành", "MVP", "COVERED", "—", "Phương án dự phòng"),
    (8, "Chấm công → QR", "Chấm mã QR", "Đang phát triển một phần", "MVP", "SPEC_GAP", "QR giai đoạn hiện tại hay giai đoạn 2?", "Vỏ màn đã có"),
    (9, "Chấm công → Khuôn mặt", "Chấm / đăng ký khuôn mặt", "Giai đoạn 2 — đang phát triển", "GĐ2", "MEETING_ONLY_GĐ2", "—", "Ngoài phạm vi hiện tại"),
    (10, "Chấm công → Vị trí", "Chấm trong vùng cho phép", "Đang vận hành", "MVP", "COVERED", "—", "Cần cấu hình vùng (xem sheet 02b GPS)"),
    (11, "Bảng chấm công", "Kỳ / bảng công theo đơn vị", "Đang vận hành", "MVP", "COVERED", "—", "Chốt trước tính lương"),
    (12, "Bảng chấm công → Thêm bảng", "Tạo kỳ chấm mới", "Đang vận hành", "MVP", "COVERED", "—", "Gắn phòng / khoảng thời gian"),
    (13, "Bản ghi chấm công", "Sổ chi tiết ngày/giờ", "Đang vận hành", "MVP", "COVERED", "—", "Điều chỉnh trước tổng hợp"),
    (14, "Chấm công tuần", "Lưới coverage theo tuần", "Đang vận hành", "MVP", "UNMAPPED_PRODUCT", "Thuộc phạm vi hiện tại?", "Chưa thay phân ca đầy đủ"),
    (15, "Tổng hợp công", "Tổng hợp trước khi chốt", "Đang vận hành (gộp sổ bản ghi)", "MVP", "SPEC_GAP", "Cần tổng hợp riêng hay chấp nhận gộp sổ?", "Đầu vào ký chốt"),
    (16, "Ca → Danh sách ca", "Danh mục ca", "Đang vận hành", "MVP", "SRS_THIN", "—", "Nền hệ số tăng ca"),
    (17, "Ca → Phân ca (lịch)", "Xếp lịch ca theo nhân viên", "Đang phát triển", "GĐ2", "PRODUCT_STUB", "Xếp ca đầy đủ giai đoạn hiện tại hay giai đoạn 2?", "Menu có — chức năng chưa đủ"),
    (18, "Ca → Tăng ca (ca OT)", "Ca tăng ca riêng", "Đang phát triển", "GĐ2", "PRODUCT_STUB", "—", "Tách hệ số — chờ hoàn thiện"),
    (19, "Đơn từ → Nghỉ phép", "Tạo / duyệt đơn nghỉ", "Đang vận hành", "MVP", "COVERED", "—", "Giữ chỗ quỹ khi gửi"),
    (20, "Đơn từ → Đi muộn / Về sớm", "Giải trình muộn sớm", "Đang vận hành", "MVP", "UNMAPPED_PRODUCT", "Cần tình huống sử dụng riêng trên giấy?", "Ảnh hưởng công chuẩn"),
    (21, "Đơn từ → Tăng ca", "Đơn tăng ca", "Đang vận hành", "MVP", "UNMAPPED_PRODUCT", "Cần tình huống sử dụng riêng?", "Vào bảng công"),
    (22, "Đơn từ → Công tác", "Đơn công tác", "Đang vận hành", "MVP", "UNMAPPED_PRODUCT", "Thuộc phạm vi hiện tại?", "Tính công ngày đi"),
    (23, "Đơn từ → Cập nhật chấm công", "Đơn điều chỉnh bản ghi", "Đang vận hành", "MVP", "UNMAPPED_PRODUCT", "Gộp vào điểm danh hay tình huống riêng?", "Sau sự cố / quên chấm"),
    (24, "Đơn từ → Đổi ca", "Đơn đổi ca đã xếp", "Đang vận hành", "MVP", "UNMAPPED_PRODUCT", "Phụ thuộc xếp ca đầy đủ", "—"),
    (25, "Đơn từ → Tổng hợp nghỉ", "Nhãn menu — cùng màn nghỉ phép (không view riêng)", "Nhãn menu · cùng UI nghỉ phép", "MVP", "ALIAS_SAME_UI", "Giữ nhãn hay gộp một mục Nghỉ phép?", "Xem sheet 02b hàng ALIAS"),
    (26, "Đơn từ → Tổng hợp nghỉ bù", "Nhãn menu — cùng màn nghỉ phép", "Nhãn menu · cùng UI nghỉ phép", "MVP", "ALIAS_SAME_UI", "Giữ nhãn hay gộp?", "Xem sheet 02b hàng ALIAS"),
    (27, "Đơn từ → Kế hoạch nghỉ", "Nhãn menu — cùng màn nghỉ phép · họp GĐ2", "Nhãn menu · cùng UI · GĐ2", "GĐ2", "ALIAS_SAME_UI", "Ẩn kế hoạch nghỉ ở giai đoạn hiện tại?", "Xem sheet 02b hàng ALIAS"),
    (28, "Tab Nghỉ phép (cấp cao)", "Màn nghỉ cho HR / quản lý", "Đang vận hành", "MVP", "COVERED", "—", "Trùng nghiệp vụ đơn nghỉ"),
    (29, "Báo cáo chấm công", "Báo cáo tháng / đơn vị", "Đang vận hành", "MVP", "UNMAPPED_PRODUCT", "—", "Trước chốt lương"),
    (30, "Báo cáo → Xuất tệp", "Xuất Excel / PDF", "Đang phát triển một phần", "MVP", "SPEC_GAP", "Xuất máy chủ giai đoạn 2?", "Định dạng kiểm toán"),
    (31, "Cài đặt → Nhân viên chấm công", "Gán mã chấm / mapping", "Đang vận hành", "MVP", "SPEC_GAP", "Cần mã chấm công riêng?", "Đối chiếu máy"),
    (32, "Cài đặt → Quy tắc chung", "Ngày công chuẩn, chốt kỳ", "Đang vận hành", "MVP", "COVERED", "—", "Chính sách holding / công ty"),
    (33, "Cài đặt → Công chuẩn", "Định nghĩa công chuẩn tháng/giờ", "Đang phát triển một phần", "MVP", "SPEC_GAP", "Cột chi tiết giai đoạn 2?", "Liên quan bảng công"),
    (34, "Cài đặt → Tùy chỉnh", "Mở rộng công ty con", "Đang vận hành (tĩnh)", "MVP", "SPEC_GAP", "—", "Ghi đè cục bộ"),
    (35, "Cài đặt → Thiết bị", "Quy tắc máy chấm", "Đang vận hành", "MVP", "SRS_THIN", "—", "Cổng / kho"),
    (36, "Cài đặt → Ứng dụng", "Chính sách chấm điện thoại", "Đang vận hành", "MVP", "SRS_THIN", "—", "Vùng vị trí — GPS sites ở sheet 02b"),
    (37, "Cài đặt → Máy tính bảng", "Chấm kiosk / tablet", "Đang phát triển", "GĐ2", "PRODUCT_STUB", "Giai đoạn 2?", "Chưa mở đủ"),
    (38, "Cài đặt → Ủy quyền chấm", "Chấm hộ có kiểm soát", "Giai đoạn 2 — đang phát triển", "GĐ2", "MEETING_ONLY_GĐ2", "—", "Kiểm soát gian lận"),
    (39, "Cài đặt → Tự động", "Tự kết thúc ca theo giờ", "Đang phát triển", "MVP", "PRODUCT_STUB", "Tự checkout giai đoạn 2?", "Chờ chính sách"),
    (40, "Cài đặt → Quy tắc tăng ca", "Hệ số / ngưỡng tăng ca", "Đang phát triển (chuyển Cài đặt HRM)", "MVP", "PRODUCT_STUB", "—", "Ưu tiên trước lương thử"),
    (41, "Cài đặt → Quy tắc nghỉ", "Ánh xạ loại nghỉ ↔ cột công", "Đang phát triển (chuyển Cài đặt HRM)", "MVP", "PRODUCT_STUB", "Khớp 5 loại phép đã họp", "Ưu tiên cấu hình"),
    (42, "Cài đặt → Đi muộn / Về sớm", "Ngưỡng phút muộn sớm", "Đang phát triển (chuyển Cài đặt HRM)", "MVP", "PRODUCT_STUB", "—", "Bổ sung màn chuyên biệt"),
    (43, "Cài đặt → Quy tắc đơn từ", "Thời hạn / cấp duyệt đơn", "Đang phát triển (chuyển Cài đặt HRM)", "MVP", "PRODUCT_STUB", "—", "Quy trình duyệt"),
    (44, "Cài đặt → Người dùng", "Phân quyền module chấm công", "Đang phát triển", "MVP", "PRODUCT_STUB", "—", "Ai cấu hình / duyệt"),
    (45, "Cài đặt → Vai trò", "Vai trò quản trị chấm công", "Đang phát triển", "GĐ2", "PRODUCT_STUB", "—", "Có thể gộp phân quyền chung"),
    (46, "Cài đặt → Hệ thống", "Tham số hệ thống module", "Đang phát triển", "GĐ2", "PRODUCT_STUB", "—", "Thấp hơn nghiệp vụ lõi"),
]

GAP_ROWS: list[tuple[str, str, str, str, str]] = [
    ("G-01", "Tuyển dụng", "Chiến dịch / đăng tin đa kênh", "Giai đoạn 2 — khi có đối tác kết nối", "Xác nhận ẩn hoặc giữ menu ở giai đoạn hiện tại"),
    ("G-02", "Tuyển dụng", "Trong / ngoài định biên + tuyển mới / thay thế", "Đã thống nhất họp — xác nhận trên giấy", "Chốt quyền Ban giám đốc ngoài định biên"),
    ("G-03", "Tuyển dụng", "Lưới kế hoạch chỉ «Cần tuyển»", "Đã thống nhất — bỏ cột trùng", "Xác nhận lưới định biên"),
    ("G-04", "Nhân sự", "Lương / ngân hàng / MST / BH không trên hồ sơ mở", "Tách sang vòng hợp đồng–bảo hiểm mật", "Xác nhận quyền xem"),
    ("G-05", "Nhân sự", "Quản lý công việc / dự án", "Module riêng — ngoài khối nhân sự hành chính", "Xác nhận ngoài phạm vi"),
    ("G-06", "Chấm công & phép", "Năm loại phép + nghỉ ốm BH / hỗ trợ công ty", "Đã họp — cấu hình quy tắc nghỉ trên phần mềm đang phát triển", "Xác nhận đủ 5 loại: năm · thâm niên · bù tăng ca · chuyển kỳ · ứng"),
    ("G-07", "Chấm công & phép", "Bảng công đã chốt = đầu vào duy nhất tính lương", "Đã thống nhất bốn khối", "Xác nhận ranh giới lương"),
    ("G-08", "Chấm công (màn hình)", "Phân ca (lịch) và ca tăng ca", "Đang phát triển — chưa đủ xếp ca", "Chốt: giai đoạn hiện tại dùng danh sách ca + đơn đổi ca"),
    ("G-09", "Chấm công (màn hình)", "Cài đặt quy tắc tăng ca / nghỉ / muộn / đơn từ / phân quyền", "Đang phát triển (chuyển Cài đặt HRM)", "Chốt thứ tự ưu tiên cấu hình trước lương thử"),
    ("G-10", "Chấm công (màn hình)", "Khuôn mặt · ủy quyền chấm hộ · kế hoạch nghỉ tập thể", "Giai đoạn 2", "Xác nhận ngoài phạm vi hiện tại"),
    ("G-11", "Chấm công", "Lịch lễ / Tết (dương + âm)", "Chưa có màn cấu hình riêng", "Có cần màn riêng giai đoạn hiện tại?"),
    ("G-12", "Chấm công", "Tổng hợp công vs sổ bản ghi", "Phần mềm đang gộp sổ bản ghi — cần chốt chấp nhận hay tách", "Đồng ý gộp hoặc yêu cầu tổng hợp riêng"),
    ("G-13", "Tiền lương", "Cách lắp công thức lương trên màn hình", "Họp lương đã xong — còn chốt cách lắp (Q-PAY-FORMULA = authoring/engine, không phải họp lương chưa hoàn tất)", "Đồng ý / điều chỉnh phương án đề xuất"),
    ("G-14", "Tiền lương", "Màn dữ liệu chấm → lương và tính lương", "Đang phát triển — không nhận là đã nghiệm thu tính lương", "Xác nhận SoT bảng công chốt trước khi mở sâu"),
    ("G-15", "Tiền lương", "Gộp lương giữa tháng · tất toán · phiếu theo nhóm", "Khung giấy có — màn hình còn thiếu / đang phát triển", "Chốt tình huống ưu tiên trên giấy"),
    ("G-16", "Chung", "Tạm dừng viết phần mềm / demo đến khi xác nhận tài liệu", "Đã thống nhất trong họp", "Xác nhận / ký packet tài liệu"),
    ("G-17", "Chấm công", "Các đơn LIVE chưa có mã tình huống giấy (muộn, tăng ca, công tác, đổi ca, tuần…)", "Cần chốt thuộc phạm vi hoặc bổ sung tình huống", "Đánh dấu Có / Không / Giai đoạn 2 trên sheet màn hình"),
    ("G-18", "Chấm công (sâu)", "18 chức năng có trên phần mềm nhưng chưa trên bản đồ 46 màn (gồm quỹ phép, GPS vùng, xóa bảng, thẻ QR…)", "Đã liệt kê sheet «02b_Man_thieu_sau» — chốt Có/Không/GĐ2", "Duyệt sheet 02b khi họp"),
    ("G-19", "Chấm công (trung thực)", "Ba mục Tổng hợp nghỉ / nghỉ bù / kế hoạch nghỉ = cùng một màn nghỉ phép", "Ghi rõ trên giấy — không nhận là ba báo cáo riêng", "Chốt gộp nhãn hoặc giữ nhãn"),
]

ATT_MEETING = {
    1: "A2", 2: "—", 3: "—", 4: "A1", 5: "A3", 6: "A6", 7: "A6", 8: "A6", 9: "A6", 10: "A6",
    11: "A2 · D8", 12: "A2", 13: "A2 · A6", 14: "A1 · A5", 15: "A2 · D8", 16: "A1",
    17: "A1 · A5", 18: "A1", 19: "A3 · A5", 20: "A1", 21: "A2 · D8", 22: "—", 23: "A6", 24: "A1",
    25: "A3", 26: "A3", 27: "—", 28: "A3", 29: "A2 · D8", 30: "—", 31: "A6", 32: "A1 · A5",
    33: "A2 · D8", 34: "—", 35: "A6", 36: "A6", 37: "—", 38: "—", 39: "—", 40: "A2 · D8",
    41: "A3 · A5", 42: "A1", 43: "A5", 44: "—", 45: "—", 46: "—",
}


def style_header(ws, row: int, ncols: int, fill: str = "1E40AF") -> None:
    f = PatternFill("solid", fgColor=fill)
    font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = f
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin


def style_body(ws, r1: int, r2: int, ncols: int, highlight_cols: set[int] | None = None) -> None:
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    font = Font(name="Calibri", size=10, color="0F172A")
    alt = PatternFill("solid", fgColor="F8FAFC")
    hi = PatternFill("solid", fgColor="FEF3C7")
    for r in range(r1, r2 + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin
            if highlight_cols and c in highlight_cols:
                cell.fill = hi
            elif r % 2 == 0:
                cell.fill = alt


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def paint_status_rows(ws, r1: int, r2: int, status_col: int, ncols: int) -> None:
    for r in range(r1, r2 + 1):
        st = str(ws.cell(r, status_col).value or "").lower()
        gap = str(ws.cell(r, min(status_col + 2, ncols)).value or "").upper()
        fill = None
        if "đang phát triển" in st or "giai đoạn 2" in st or "stub" in st or gap in {
            "PRODUCT_STUB", "MEETING_ONLY_GĐ2", "GĐ2-HOLD",
        }:
            fill = PatternFill("solid", fgColor="FEE2E2")
        elif "một phần" in st or "chưa" in st or "partial" in st or gap in {
            "PRODUCT_MISSING", "SPEC_GAP", "UNMAPPED_PRODUCT", "SRS_THIN",
        }:
            fill = PatternFill("solid", fgColor="FFEDD5")
        if fill:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = fill


def build() -> Path:
    assert len(UC_ROWS) == 45, f"Expected 45 UC, got {len(UC_ROWS)}"
    assert len(ATT_SURFACES) == 46, f"Expected 46 ATT, got {len(ATT_SURFACES)}"
    missing_only = [r for r in ATT_DEEP_ROWS if str(r[0]).startswith("S") and "ALIAS" not in str(r[0])]
    assert len(missing_only) == 18, f"Expected 18 MISSING, got {len(missing_only)}"

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Huong_dan"
    ws["A1"] = "WBS chốt họp theo tình huống sử dụng (UC) — HRM doanh nghiệp"
    ws["A1"].font = Font(bold=True, size=15, color="1E40AF", name="Calibri")
    guide = [
        "",
        f"Phiên bản: {VERSION} — gộp ma trận khoảng trống họp × mô tả chức năng × màn hình (ATT-FID#1–46 + 18 chức năng sâu).",
        "Nguồn: tổng hợp quyết định họp · danh mục 45 tình huống · ma trận khoảng trống v1.1 · 46 màn chấm công · kiểm tra trình duyệt đọc-only.",
        "",
        "Cách dùng khi họp:",
        "1) «01_Danh_muc_UC» — duyệt MVP / GĐ2; cột khoảng trống; cột cần khách chốt; ký Đồng ý / Điều chỉnh / GĐ2.",
        "2) «02_Man_cham_cong» — ATT-FID#1–46 + tham chiếu họp + trạng thái trình duyệt (LIVE / PARTIAL / STUB_UI / GĐ2-HOLD).",
        "3) «02b_Man_thieu_sau» — 18 chức năng sâu (GPS vùng · thẻ QR · quỹ phép · xóa bảng…) + trung thực nhãn nghỉ.",
        "4) «03_Tom_tat_khoang_trong» — các điểm quyết định G-01…G-19.",
        "5) Diễn biến từng bước (sâu): WBS_HRM_ENTERPRISE_KHACH_MOI.xlsx.",
        "6) Mô tả chức năng: SRS_HRM_ENTERPRISE.md / PDF khách (v0.7).",
        "",
        "Khóa nghiệp vụ: chiến dịch tuyển = giai đoạn 2; khuôn mặt = giai đoạn 2;",
        "loại phép năm · thâm niên · bù tăng ca · chuyển kỳ · ứng; nghỉ ốm BH / hỗ trợ công ty;",
        "bảng công chốt = đầu vào lương; trụ tiền lương đã họp xong — còn chốt cách lắp công thức",
        "(Q-PAY-FORMULA = cách lắp / động cơ — không mang nghĩa họp lương chưa hoàn tất).",
        "Tạm dừng viết phần mềm / demo đến khi xác nhận tài liệu trên giấy.",
        "Tài liệu này không khẳng định khách đã ký nghiệm thu hay phần mềm đã hoàn tất.",
        "Không mở đặc tả kỹ thuật khi còn khoảng trống chưa chốt trên giấy.",
    ]
    for i, line in enumerate(guide, 2):
        ws.cell(i, 1, line)
        ws.cell(i, 1).font = Font(name="Calibri", size=11)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 118

    # 01 UC
    ws = wb.create_sheet("01_Danh_muc_UC")
    h1 = [
        "STT", "Mã UC", "Tên tình huống", "Khối", "Giai đoạn",
        "Tham chiếu họp", "Trạng thái mô tả chức năng", "Mã hạng mục WBS",
        "Trạng thái màn hình", "Khoảng trống (khách)", "Việc cần khách chốt",
        "Độ ưu tiên gửi chốt", "Ghi chú", "Cột ký họp",
    ]
    for c, h in enumerate(h1, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(h1))
    for i, row in enumerate(UC_ROWS, 2):
        (uc, ten, pillar, gd, mref, srs, wbs, prt, gap, dec, prio, note) = row
        vals = [
            i - 1, uc, ten, MODULE_VI[pillar], gd, mref, srs, wbs, prt,
            GAP_VI.get(gap, gap), dec, prio, note, "",
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)
    last = 1 + len(UC_ROWS)
    style_body(ws, 2, last, len(h1), highlight_cols={5, 12})
    paint_status_rows(ws, 2, last, 9, len(h1))
    set_widths(ws, [5, 16, 42, 18, 10, 14, 16, 14, 28, 28, 32, 20, 32, 22])
    ws.auto_filter.ref = f"A1:N{last}"
    ws.freeze_panes = "C2"
    for r in range(2, last + 1):
        ws.row_dimensions[r].height = 40

    # 02 ATT fidelity #1–46 + browser
    ws = wb.create_sheet("02_Man_cham_cong")
    h2 = [
        "STT", "ATT-FID", "Đường menu", "Ý nghĩa nghiệp vụ",
        "Tham chiếu họp", "Mã inventory sâu", "Trạng thái trình duyệt",
        "Trạng thái màn hình (khách)", "Giai đoạn", "Khoảng trống (khách)",
        "Việc cần khách chốt", "Ghi chú", "Cột ký họp",
    ]
    for c, h in enumerate(h2, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(h2), fill="0F766E")
    for i, row in enumerate(ATT_SURFACES, 2):
        num, menu, meaning, status, gd, gap, dec, note = row
        vals = [
            i - 1,
            f"ATT-FID#{num}",
            menu,
            meaning,
            ATT_MEETING.get(num, "—"),
            ATT_INV_PARENT.get(num, "—"),
            ATT_BROWSER.get(num, "UNKNOWN"),
            status,
            gd,
            GAP_VI.get(gap, gap),
            dec,
            note,
            "",
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)
    last2 = 1 + len(ATT_SURFACES)
    style_body(ws, 2, last2, len(h2), highlight_cols={7})
    paint_status_rows(ws, 2, last2, 7, len(h2))
    set_widths(ws, [5, 12, 36, 36, 12, 12, 16, 28, 10, 28, 32, 28, 18])
    ws.auto_filter.ref = f"A1:M{last2}"
    ws.freeze_panes = "C2"
    for r in range(2, last2 + 1):
        ws.row_dimensions[r].height = 40

    # 02b — 18 MISSING + aliases
    ws = wb.create_sheet("02b_Man_thieu_sau")
    h2b = [
        "STT", "Mã inventory sâu", "Đường menu", "Ý nghĩa",
        "ATT-FID cha", "Ghi chú trình duyệt", "Khoảng trống (khách)",
        "Giai đoạn", "Tình huống đề xuất (nội bộ)", "Mã WBS",
        "Việc cần khách chốt", "Cột ký họp",
    ]
    for c, h in enumerate(h2b, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(h2b), fill="7C2D12")
    for i, row in enumerate(ATT_DEEP_ROWS, 2):
        inv, menu, meaning, parent, browser, gap, mvp, dec, prop, wbs = row
        vals = [
            i - 1, inv, menu, meaning, parent, browser,
            GAP_VI.get(gap, gap), mvp, prop, wbs, dec, "",
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(i, c, v)
    last2b = 1 + len(ATT_DEEP_ROWS)
    style_body(ws, 2, last2b, len(h2b), highlight_cols={6, 11})
    paint_status_rows(ws, 2, last2b, 6, len(h2b))
    set_widths(ws, [5, 14, 40, 36, 16, 36, 28, 10, 18, 12, 36, 18])
    ws.auto_filter.ref = f"A1:L{last2b}"
    ws.freeze_panes = "C2"
    for r in range(2, last2b + 1):
        ws.row_dimensions[r].height = 44

    # 03 gaps
    ws = wb.create_sheet("03_Tom_tat_khoang_trong")
    h3 = [
        "STT", "Mã", "Khối", "Nội dung cần làm rõ",
        "Hướng xử lý đề xuất", "Việc khách cần làm", "Ý kiến khách",
    ]
    for c, h in enumerate(h3, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(h3), fill="B45309")
    for i, row in enumerate(GAP_ROWS, 2):
        code, khoi, nd, huong, viec = row
        for c, v in enumerate([i - 1, code, khoi, nd, huong, viec, ""], 1):
            ws.cell(i, c, v)
    last3 = 1 + len(GAP_ROWS)
    style_body(ws, 2, last3, len(h3), highlight_cols={7})
    set_widths(ws, [5, 8, 18, 48, 48, 40, 28])
    ws.freeze_panes = "D2"
    for r in range(2, last3 + 1):
        ws.row_dimensions[r].height = 48

    # 99
    ws = wb.create_sheet("99_Thong_ke")
    ws["A1"] = "Thống kê nội bộ (khách có thể bỏ qua)"
    ws["A1"].font = Font(bold=True, color="64748B")
    browser_counts: dict[str, int] = {}
    for v in ATT_BROWSER.values():
        browser_counts[v] = browser_counts.get(v, 0) + 1
    stats = [
        ("Tổng UC", 45),
        ("UC MVP", sum(1 for r in UC_ROWS if r[3] == "MVP")),
        ("UC GĐ2 (campaign)", sum(1 for r in UC_ROWS if r[3] == "GĐ2")),
        ("UC đủ 7 mục (ưu tiên)", sum(1 for r in UC_ROWS if r[5] == "Đủ 7 mục")),
        ("UC có WBS-REC-00", "có" if any(r[0] == "UC-BP-REC-00" for r in UC_ROWS) else "thiếu"),
        ("Màn chấm công ATT-FID", 46),
        ("LIVE trình duyệt", browser_counts.get("LIVE", 0)),
        ("PARTIAL trình duyệt", browser_counts.get("PARTIAL", 0)),
        ("STUB_UI trình duyệt", browser_counts.get("STUB_UI", 0)),
        ("GĐ2-HOLD trình duyệt", browser_counts.get("GĐ2-HOLD", 0)),
        ("Chức năng sâu MISSING (18)", 18),
        ("Hàng sheet 02b (18 + 3 alias)", len(ATT_DEEP_ROWS)),
        ("Dòng khoảng trống tóm tắt", len(GAP_ROWS)),
        ("Phiên bản", VERSION),
        ("SRS đồng bộ", "0.7"),
        ("Inventory", "0.3.3"),
        ("Ma trận khoảng trống", "v1.1"),
        ("Trình duyệt sâu", "ATT-DEEP-QA-01 (đọc-only)"),
        ("Cấm wording", "Không ghi unfinished-PAY / họp lương chưa hoàn tất"),
        ("READY_FOR_TECHSPEC", "KHÔNG — matrix NOT_READY"),
        ("uat_done / Attendance CLOSED", "false / false"),
    ]
    ws.cell(2, 1, "Chỉ số")
    ws.cell(2, 2, "Giá trị")
    style_header(ws, 2, 2, fill="64748B")
    for i, (k, v) in enumerate(stats, 3):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)
    style_body(ws, 3, 2 + len(stats), 2)
    set_widths(ws, [55, 40])

    try:
        wb.save(OUT)
        return OUT
    except PermissionError:
        wb.save(OUT_FALLBACK)
        return OUT_FALLBACK


if __name__ == "__main__":
    path = build()
    print(f"OK: {path}")
    print(f"version: {VERSION}")
    print(f"UC: {len(UC_ROWS)}")
    print(f"ATT: {len(ATT_SURFACES)}")
    print(f"DEEP: {len(ATT_DEEP_ROWS)}")
    print(f"GAP: {len(GAP_ROWS)}")
    print(f"size_kb: {path.stat().st_size // 1024}")
