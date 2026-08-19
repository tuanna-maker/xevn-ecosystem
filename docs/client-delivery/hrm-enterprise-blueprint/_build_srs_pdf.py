# -*- coding: utf-8 -*-
"""Sinh PDF SRS khách từ Excel WBS (03 + 03b) — tiếng Việt, Unicode."""
from __future__ import annotations

from pathlib import Path

from fpdf import FPDF
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
XLSX_CANDIDATES = [
    ROOT / "WBS_HRM_ENTERPRISE_KHACH_MOI.xlsx",
    ROOT / "WBS_HRM_ENTERPRISE_KHACH.xlsx",
]
OUT = ROOT / "SRS_HRM_ENTERPRISE_KHACH.pdf"
FONT = Path(r"C:\Windows\Fonts\arial.ttf")
FONT_BOLD = Path(r"C:\Windows\Fonts\arialbd.ttf")


class SrsPdf(FPDF):
    def header(self) -> None:
        if self.page_no() == 1:
            return
        self.set_font("ArialVN", "", 9)
        self.set_text_color(100, 116, 139)
        self.cell(0, 8, "Đặc tả yêu cầu phần mềm — HRM doanh nghiệp (bản gửi khách)", align="L")
        self.ln(10)

    def footer(self) -> None:
        self.set_y(-12)
        self.set_font("ArialVN", "", 8)
        self.set_text_color(100, 116, 139)
        self.cell(0, 8, f"Trang {self.page_no()}/{{nb}} · Không khẳng định phần mềm đã nghiệm thu", align="C")


def pick_xlsx() -> Path:
    for p in XLSX_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError("Không thấy file Excel WBS khách")


def load_data(xlsx: Path):
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    uc_rows = []
    ws = wb["03_Tinh_huong_nghiep_vu"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # expect col: STT, Module, WBS, Ten, Muc dich, Ai, Dien bien tom tat, Quy tac, ...
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[3]:
            continue
        uc_rows.append({
            "module": row[1] or "",
            "wbs": row[2] or "",
            "ten": row[3] or "",
            "muc_dich": row[4] or "",
            "ai": row[5] or "",
            "quy_tac": row[7] or "",
            "kiem_tra": row[8] or "",
            "pass": row[9] or "",
            "fail": row[10] or "",
            "edge": row[11] or "",
            "uc": row[16] or "",
            "br": row[17] or "",
        })

    steps_by_uc: dict[str, list[dict]] = {}
    ws2 = wb["03b_Dien_bien_chi_tiet"]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or not row[14]:
            continue
        uc = str(row[14])
        steps_by_uc.setdefault(uc, []).append({
            "buoc": row[4],
            "loai": row[5],
            "ai": row[6],
            "user": row[7],
            "sys": row[8],
            "dk": row[9],
            "ok": row[10],
            "fail": row[11],
            "next": row[12],
        })
    wb.close()
    return uc_rows, steps_by_uc


def _clean(text: str) -> str:
    t = str(text or "—").replace("\r\n", "\n").replace("\r", "\n")
    # break ultra-long tokens that can choke line-break
    out = []
    for part in t.split(" "):
        if len(part) > 80:
            for i in range(0, len(part), 60):
                out.append(part[i : i + 60])
        else:
            out.append(part)
    return " ".join(out)


def _mc(pdf: SrsPdf, text: str, h: float) -> None:
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(pdf.epw, h, _clean(text))


def h1(pdf: SrsPdf, text: str) -> None:
    pdf.set_font("ArialVN", "B", 14)
    pdf.set_text_color(30, 64, 175)
    _mc(pdf, text, 8)
    pdf.ln(2)


def h2(pdf: SrsPdf, text: str) -> None:
    pdf.set_font("ArialVN", "B", 12)
    pdf.set_text_color(15, 23, 42)
    _mc(pdf, text, 7)
    pdf.ln(1)


def h3(pdf: SrsPdf, text: str) -> None:
    if pdf.get_y() > 250:
        pdf.add_page()
    pdf.set_font("ArialVN", "B", 11)
    pdf.set_text_color(15, 118, 110)
    _mc(pdf, text, 6)
    pdf.ln(1)


def body(pdf: SrsPdf, text: str, size: int = 10) -> None:
    pdf.set_font("ArialVN", "", size)
    pdf.set_text_color(15, 23, 42)
    _mc(pdf, text or "—", 5.2)
    pdf.ln(1)


def label_para(pdf: SrsPdf, label: str, text: str) -> None:
    pdf.set_font("ArialVN", "B", 10)
    pdf.set_text_color(51, 65, 85)
    _mc(pdf, label, 5)
    body(pdf, str(text) if text else "—", 10)


def build() -> Path:
    if not FONT.exists():
        raise FileNotFoundError(f"Thiếu font {FONT}")
    xlsx = pick_xlsx()
    uc_rows, steps_by_uc = load_data(xlsx)

    pdf = SrsPdf(format="A4")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_font("ArialVN", "", str(FONT))
    pdf.add_font("ArialVN", "B", str(FONT_BOLD if FONT_BOLD.exists() else FONT))

    # Cover
    pdf.add_page()
    pdf.ln(30)
    pdf.set_font("ArialVN", "B", 20)
    pdf.set_text_color(30, 64, 175)
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(pdf.epw, 10, "ĐẶC TẢ YÊU CẦU PHẦN MỀM\nHệ thống quản lý nhân sự doanh nghiệp", align="C")
    pdf.ln(8)
    pdf.set_font("ArialVN", "", 12)
    pdf.set_text_color(15, 23, 42)
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(
        pdf.epw,
        7,
        "Bản gửi khách chốt logic nghiệp vụ trên giấy\n"
        "Bốn khối: Tuyển dụng · Nhân sự · Chấm công & Nghỉ phép · Tiền lương & Phúc lợi\n"
        f"Số tình huống: {len(uc_rows)} · Nguồn: bảng WBS/Excel diễn biến chi tiết\n"
        "Phiên bản gói: 0.4 (đồng bộ inventory / ma trận)",
        align="C",
    )
    pdf.ln(12)
    body(
        pdf,
        "Tài liệu này mô tả use case và quy tắc nghiệp vụ để Ban dự án / đối tác đọc và chốt trước khi mở đặc tả kỹ thuật và phát triển. "
        "Không khẳng định phần mềm đã nghiệm thu hay đã triển khai xong.",
        10,
    )

    # Ch1
    pdf.add_page()
    h1(pdf, "1. Giới thiệu")
    h2(pdf, "1.1. Mục đích")
    body(
        pdf,
        "Chuẩn hóa vòng đời nhân sự từ định biên → tuyển → hồ sơ → chấm công/phép → bảng công chốt → lương, "
        "với các quy tắc ngoại lệ (phép xuyên cuối tuần/lễ, giữ chỗ quỹ phép, gộp lương giữa tháng, tách dữ liệu lương thưởng mật) "
        "thống nhất giữa các đơn vị.",
    )
    h2(pdf, "1.2. Phạm vi")
    body(
        pdf,
        "Trong phạm vi (logic giấy): bốn khối độc lập và ranh giới liên kết đã thống nhất. "
        "Ngoài phạm vi tài liệu này: đặc tả kỹ thuật chi tiết, thiết kế dữ liệu vật lý, hợp đồng tích hợp — giữ HOLD đến khi khách xác nhận.",
    )
    h2(pdf, "1.3. Ràng buộc cốt lõi")
    body(
        pdf,
        "1) Tiền lương chỉ đọc bảng công đã ký chốt — không lấy giờ thẳng từ tăng ca hay đơn phép để tính lương.\n"
        "2) Tuyển dụng không nối trực tiếp sang tiền lương.\n"
        "3) Thông tin lương, thuế, bảo hiểm, ngân hàng chỉ người đúng quyền (chuyên viên lương thưởng & phúc lợi) được xem.\n"
        "4) Các quyết định còn mở (công thức lương, trong/ngoài định biên, đơn vị trừ phép…) ghi rõ «chờ chốt» — chưa khẳng định khách đã ký.",
    )

    # Ch2
    h1(pdf, "2. Bốn khối và liên kết")
    body(
        pdf,
        "Tuyển dụng — «Bao giờ có người làm?»\n"
        "Nhân sự — «Hồ sơ đủ giấy tờ chưa? Ai xem dữ liệu mật?»\n"
        "Chấm công & Nghỉ phép — «Giờ công đúng ca? Phép trừ đúng ngày làm?»\n"
        "Tiền lương & Phúc lợi — «Công thức ai cấu hình? Đổi lương giữa kỳ có một phiếu thực nhận?»\n\n"
        "Liên kết: Nhận việc → hồ sơ nhân sự (không nhập lại) → hồ sơ Hoạt động mở phép/ca → bảng công chốt → lương. "
        "Cấm tuyển dụng trao đổi trực tiếp với tiền lương.",
    )

    # Ch3 — UC
    h1(pdf, "3. Yêu cầu chức năng (từng tình huống)")
    body(
        pdf,
        "Mỗi tình huống gồm: mục đích, người tham gia, quy tắc, điều kiện đạt/không đạt, và diễn biến từng bước "
        "(rút từ bảng Excel diễn biến chi tiết — đủ để Ban nghiệp vụ, phân tích giải pháp, phát triển web/mobile và kiểm thử cùng hiểu).",
    )

    current_mod = None
    for i, uc in enumerate(uc_rows, 1):
        if uc["module"] != current_mod:
            current_mod = uc["module"]
            h2(pdf, f"Module: {current_mod}")

        code = uc["uc"] or f"UC-{i}"
        h3(pdf, f"{i}. {code} — {uc['ten']}")
        label_para(pdf, "Mục đích", uc["muc_dich"])
        label_para(pdf, "Người tham gia", uc["ai"])
        if uc["wbs"]:
            label_para(pdf, "Hạng mục WBS", uc["wbs"])
        label_para(pdf, "Quy tắc nghiệp vụ", uc["quy_tac"])
        label_para(pdf, "Được chấp nhận khi", uc["pass"])
        label_para(pdf, "Không chấp nhận khi", uc["fail"])
        if uc["edge"]:
            label_para(pdf, "Tình huống đặc biệt", uc["edge"])

        steps = steps_by_uc.get(str(code), [])
        if steps:
            pdf.set_font("ArialVN", "B", 10)
            pdf.set_text_color(51, 65, 85)
            pdf.multi_cell(0, 5, "Diễn biến từng bước")
            for s in steps:
                buoc = s.get("buoc")
                title = f"Bước {buoc}" if buoc not in (None, "T") else "Kết thúc thành công"
                if s.get("loai"):
                    title = f"{title} · {s['loai']}"
                pdf.set_font("ArialVN", "B", 9)
                pdf.set_text_color(15, 23, 42)
                _mc(pdf, title, 5)
                bits = [
                    f"Ai: {s.get('ai') or '—'}",
                    f"Người dùng làm: {s.get('user') or '—'}",
                    f"Hệ thống xử lý: {s.get('sys') or '—'}",
                    f"Điều kiện: {s.get('dk') or '—'}",
                    f"Thành công: {s.get('ok') or '—'}",
                    f"Thất bại: {s.get('fail') or '—'}",
                ]
                if s.get("next"):
                    bits.append(f"Tiếp theo: {s['next']}")
                body(pdf, "\n".join(bits), 9)
        pdf.ln(2)

    # Ch4 decisions hint
    h1(pdf, "4. Việc cần chốt với khách (tóm tắt)")
    body(
        pdf,
        "Chi tiết và cột ghi ý kiến nằm trên sheet «Cần chốt với khách» trong file Excel WBS. "
        "Các nhóm chính: trong/ngoài định biên; công thức lương (ai cấu hình / kiểm soát phát hành); "
        "đơn vị trừ phép (nửa ngày hay giờ); cấp phép theo thâm niên/chức vụ; tạm hoãn bảo hiểm khi ốm dài; "
        "phạm vi module tài sản giai đoạn này.",
    )

    h1(pdf, "5. Ghi chú giai đoạn kỹ thuật")
    body(
        pdf,
        "Đặc tả kỹ thuật, thiết kế dữ liệu vật lý và hợp đồng tích hợp chi tiết giữ HOLD — chỉ mở sau khi khách xác nhận logic trên giấy "
        "(WBS/Excel + tài liệu này).",
    )

    pdf.output(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"OK: {path}")
    print(f"size_kb: {path.stat().st_size // 1024}")
