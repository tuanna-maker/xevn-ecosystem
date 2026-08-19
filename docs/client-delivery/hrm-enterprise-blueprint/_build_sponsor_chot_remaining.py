# -*- coding: utf-8 -*-
"""Excel residual sau khi sponsor đã điền SPONSOR_CHOT_FILL.xlsx (v1.0 · 34 dòng).

Output: SPONSOR_CHOT_REMAINING.xlsx
  - 00_Da_chot_tu_phieu_truoc (chỉ đọc)
  - 01_Con_can_chot (cột vàng)
  - 02_18_MISSING (cột vàng)
  - 03_UC_Lich_EXPAND (cột vàng)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "SPONSOR_CHOT_REMAINING.xlsx"
OUT_ALT = ROOT / "SPONSOR_CHOT_REMAINING_ALT.xlsx"
ANSWERS = Path("docs/qa/evidence/_tmp-sponsor-chot-fill-read.json")
VERSION = "1.0"

THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
FILL_HDR = PatternFill("solid", fgColor="1E40AF")
FILL_Y = PatternFill("solid", fgColor="FEF3C7")
FILL_RO = PatternFill("solid", fgColor="F8FAFC")
FILL_OK = PatternFill("solid", fgColor="DCFCE7")
FILL_TITLE = PatternFill("solid", fgColor="DBEAFE")
FONT_HDR = Font(bold=True, color="FFFFFF", size=11)
FONT_T = Font(bold=True, color="1E3A8A", size=13)
WRAP = Alignment(wrap_text=True, vertical="center")

# inv, menu, meaning, proposed
MISSING_18 = [
    ("S03", "Tổng quan → Chấm công ngay", "Nút mở chấm từ tổng quan", "Giữ / gộp vào vào-ra"),
    ("S04", "Tổng quan → Tùy chỉnh layout", "Bố cục đang khóa", "GĐ2 hay mở MVP"),
    ("S07", "Biểu đồ tròn loại nghỉ", "Phân tích loại nghỉ", "MVP / GĐ2 / OUT"),
    ("S15", "Thẻ QR nhân viên", "Xem/tải/in QR", "PROP-03e · MVP?"),
    ("S16", "Phóng to QR", "Dialog phóng to", "Theo S15"),
    ("S25", "Xóa bảng chấm", "Xác nhận xóa kỳ", "Ai được xóa?"),
    ("S28", "Xóa bản ghi chấm", "Xác nhận xóa dòng", "Ai được xóa?"),
    ("S29", "Xuất từ sổ bản ghi", "Path xuất ≠ báo cáo", "Gộp #30 / riêng"),
    ("S32", "Chi tiết ô tuần", "Dialog ô tuần", "MVP?"),
    ("S33", "Icon tuần no-op", "Bút/cài/tải trống", "Ẩn / làm / GĐ2"),
    ("S39", "Sao chép ca", "Nút copy ca", "MVP / GĐ2"),
    ("S43", "Panel quỹ phép", "Số dư theo loại", "PROP-05b · bắt buộc?"),
    ("S65", "Import NV chấm", "Dialog nhập khẩu", "MVP?"),
    ("S66", "Icon lọc/tải NV", "No-op", "Ẩn / làm"),
    ("S70", "Reset/Xem trước quy tắc", "No-op", "GĐ2?"),
    ("S71", "Gợi ý phương thức chấm", "No-op", "MVP / GĐ2"),
    ("S74", "Danh sách điểm GPS", "List vùng chấm", "PROP-03d · MVP?"),
    ("S75", "Thêm điểm GPS", "name/lat/lon/radius", "PROP-03d · MVP?"),
]

LICH_UC = [
    ("UC-BP-ATT-01", "Quy tắc ca theo bộ phận", "ATT"),
    ("UC-BP-ATT-03", "Điểm danh đa nguồn", "ATT"),
    ("UC-BP-ATT-03b", "Lịch lễ / Tết dương+âm", "ATT"),
    ("UC-BP-ATT-04", "Cấp phát phép + 5 loại", "ATT"),
    ("UC-BP-ATT-04b", "Ứng phép", "ATT"),
    ("UC-BP-ATT-05", "Phép chuyển kỳ", "ATT"),
    ("UC-BP-ATT-06", "Phép bù OT", "ATT"),
    ("UC-BP-ATT-07", "Nghỉ ốm + BH", "ATT"),
    ("UC-BP-ATT-12", "Mở quỹ phép khi Hoạt động", "ATT"),
    ("UC-BP-REC-00", "Thư viện JD master", "REC"),
    ("UC-BP-REC-04", "Quét kho CV nội bộ", "REC"),
    ("UC-BP-REC-05", "Lịch sử trạng thái UV", "REC"),
    ("UC-BP-REC-06", "Thư tuyển + đánh giá PV", "REC"),
    ("UC-BP-REC-07", "Offer → hồ sơ NS", "REC"),
    ("UC-BP-CORE-02b", "Metadata field hồ sơ", "CORE"),
    ("UC-BP-CORE-03", "Checklist giấy tờ", "CORE"),
    ("UC-BP-CORE-04", "OCR giấy tờ", "CORE"),
    ("UC-BP-CORE-05", "Tài sản cấp phát", "CORE"),
    ("UC-BP-CORE-06", "Thu hồi tài sản nghỉ việc", "CORE"),
    ("UC-BP-CORE-07", "Kích hoạt Hoạt động", "CORE"),
    ("UC-BP-CORE-09", "Hợp đồng LĐ Word", "CORE"),
    ("UC-BP-CORE-10", "BHXH lifecycle", "CORE"),
    ("UC-BP-PAY-03", "Giảm trừ gia cảnh", "PAY"),
    ("UC-BP-PAY-05", "Trần BH trên kỳ", "PAY"),
    ("UC-BP-PAY-06", "Tính lương kỳ", "PAY"),
    ("UC-BP-PAY-07", "Tất toán nghỉ việc", "PAY"),
    ("UC-BP-PAY-08", "Phiếu lương", "PAY"),
    ("UC-BP-PAY-09", "Phân nhóm bảng lương", "PAY"),
]

REMAIN_Q = [
    (
        "R-FY-01",
        "Năm tài chính phép",
        "Anh chọn «Năm tài chính» nhưng chưa ghi tháng bắt đầu. Tháng nào là tháng 1 của năm phép?",
        "Team cần số để cấu hình menu",
        "Tháng 1|Tháng 4|Tháng 7|Tháng 10|Khác (ghi Ghi chú)",
    ),
    (
        "R-LV-ADV-01",
        "Ứng phép — tham số cấu hình",
        "Trần ứng tối đa (% quỹ năm / số ngày cố định)? Có bắt buộc duyệt đặc biệt không?",
        "Anh yêu cầu cấu hình giống MISA/tham chiếu — cần số mặc định GĐ1",
        "% quỹ năm (ghi % ở Ghi chú)|Số ngày cố định (ghi số)|Cấu hình theo loại phép|Khác",
    ),
    (
        "R-LV-ADV-02",
        "Ứng phép — trừ kỳ sau",
        "Khi ứng: trừ ngay quỹ tương lai hay trừ khi cấp phép năm mới?",
        "—",
        "Trừ ngay quỹ tương lai|Trừ khi cấp năm mới|Cấu hình được|Khác",
    ),
    (
        "R-SICK-01",
        "Nghỉ ốm — thứ tự quỹ",
        "Anh: tách BHXH/CTY + check còn phép → hết phép = không lương. Thứ tự trừ mặc định?",
        "Cần rule máy",
        "Phép năm trước → BHXH → CTY → không lương|BHXH trước → phép → không lương|Cấu hình thứ tự|Khác",
    ),
    (
        "R-SIGN-01",
        "Ký chốt bảng công — thứ tự",
        "Anh: NV + QL trực tiếp + HR. Thứ tự bắt buộc?",
        "—",
        "NV → QL → HR|QL → NV → HR|Song song NV+QL rồi HR|Cấu hình workflow XBOS|Khác",
    ),
    (
        "R-PROP-03d",
        "UC địa điểm GPS (03d)",
        "Có ADD vào SRS giấy + làm MVP (web+mobile) không?",
        "Anh muốn code P0 GPS/vân tay/khuôn mặt trên mobile",
        "IN MVP giấy + code|Chỉ mobile trước|GĐ2|OUT",
    ),
    (
        "R-PROP-03e",
        "UC thẻ QR NV (03e)",
        "Thẻ QR xem/tải/in — MVP?",
        "—",
        "IN MVP giấy + code|Chỉ mobile|GĐ2|OUT",
    ),
    (
        "R-PROP-05b",
        "UC panel quỹ phép (05b)",
        "Panel số dư theo loại phép khi nộp đơn — MVP?",
        "—",
        "IN MVP giấy + code|GĐ2|OUT",
    ),
    (
        "R-FACE-01",
        "Face ID — phạm vi MVP",
        "Anh đưa Face vào MVP (trước họp = GĐ2). Phạm vi?",
        "Ghi chú mobile: GPS / vân tay / khuôn mặt ưu tiên app",
        "Mobile only MVP|Web+Mobile|Mobile demo trước web|Khác",
    ),
    (
        "R-DEMO-01",
        "Demo ngày mai — phạm vi thật",
        "Anh: demo ngày mai + toàn bộ UC. Thực tế stack chưa đủ mọi UC LIVE. Demo cái gì?",
        "Tránh overpromise",
        "Chỉ màn LIVE hiện có|LIVE + mock luồng giấy|Full script giấy không claim product|Khác (ghi list)",
    ),
    (
        "R-PAY-DD-01",
        "Công thức kéo-thả GĐ1",
        "Anh chọn GĐ1 kéo-thả (team từng đề xuất form trước). Xác nhận?",
        "Đồng ý 2 bước soạn→phát hành vẫn giữ",
        "Kéo-thả GĐ1 đủ|Form GĐ1 + kéo-thả GĐ2|Khác",
    ),
    (
        "R-OCR-01",
        "OCR giấy tờ (CORE-04)",
        "OCR prefill hồ sơ — đưa MVP hay GĐ2?",
        "—",
        "MVP|GĐ2|OUT",
    ),
    (
        "R-CAMPAIGN-01",
        "Chiến dịch tuyển đa kênh (REC-03)",
        "Họp khóa GĐ2. Giữ GĐ2 hay kéo vào MVP?",
        "—",
        "Giữ GĐ2|Kéo MVP|OUT",
    ),
    (
        "R-PDF-01",
        "PDF luồng UC",
        "Sau khi đọc PDF SRS/UC flow — anh chốt «đủ để làm» hay cần bổ sung UC nào?",
        "Điền sau khi đọc PDF kèm theo",
        "Đủ|Cần bổ sung (liệt kê mã UC ở Ghi chú)|Chưa đọc",
    ),
]


def _hdr(ws, row, headers, fill=FILL_HDR):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = FONT_HDR
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title(ws, text, merge_to="G"):
    ws["A1"] = text
    ws["A1"].font = FONT_T
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells(f"A1:{merge_to}1")
    ws.row_dimensions[1].height = 36


def sheet_done(wb, answers: list[dict]):
    ws = wb.create_sheet("00_Da_chot_tu_phieu_truoc", 0)
    _title(
        ws,
        f"ĐÃ CHỐT từ SPONSOR_CHOT_FILL.xlsx — chỉ đọc (v{VERSION}). "
        "Không sửa sheet này; nếu đổi ý → ghi ở sheet 01.",
        "E",
    )
    _hdr(ws, 3, ["STT", "Mã", "Quyết định đã ghi", "Ghi chú đã ghi", "Nhóm"])
    for i, a in enumerate(answers, 1):
        r = 3 + i
        vals = [i, a.get("code"), a.get("decision"), a.get("note") or "", a.get("group")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.fill = FILL_OK
        ws.row_dimensions[r].height = 40
    _widths(ws, [5, 18, 55, 45, 22])
    ws.freeze_panes = "A4"


def sheet_remain(wb):
    ws = wb.create_sheet("01_Con_can_chot", 1)
    _title(
        ws,
        "CÒN CẦN CHỐT — chỉ điền cột VÀNG (F, G, H). Nguồn: phân tích phiếu đã điền + PROP/LICH chưa có trên v1.0.",
        "H",
    )
    ws["A2"] = (
        "PDF luồng UC kèm theo: SRS_HRM_ENTERPRISE_KHACH.pdf + (nếu có) SPONSOR_UC_FLOW_CHOT.pdf. "
        "Đọc PDF rồi điền R-PDF-01."
    )
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 32

    headers = [
        "STT",
        "Mã",
        "Chủ đề",
        "Câu hỏi cần chốt",
        "Vì sao còn mở",
        "QUYẾT ĐỊNH ← ĐIỀN",
        "Ghi chú",
        "Ngày (dd/MM/yyyy)",
    ]
    _hdr(ws, 3, headers)

    dv_cache: dict[str, DataValidation] = {}
    for i, (ma, chu_de, cau, vi_sao, lua) in enumerate(REMAIN_Q, 1):
        r = 3 + i
        q_full = cau + "\n\nLựa chọn gợi ý: " + lua.replace("|", " · ")
        vals = [i, ma, chu_de, q_full, vi_sao, "", "", ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.fill = FILL_Y if c >= 6 else FILL_RO
        ws.row_dimensions[r].height = 64
        if "|" in lua:
            items = [x.replace(",", " -") for x in lua.split("|")]
            key = lua
            if key not in dv_cache:
                dv = DataValidation(type="list", formula1='"' + ",".join(items) + '"', allow_blank=True)
                ws.add_data_validation(dv)
                dv_cache[key] = dv
            dv_cache[key].add(ws.cell(r, 6))

    _widths(ws, [5, 14, 22, 50, 36, 32, 28, 12])
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{3 + len(REMAIN_Q)}"


def sheet_missing(wb):
    ws = wb.create_sheet("02_18_MISSING", 2)
    _title(
        ws,
        "18 bề mặt chấm công sâu — chọn IN MVP giấy · GĐ2 · OUT · DEFER (cột vàng E).",
        "F",
    )
    _hdr(ws, 3, ["STT", "Mã surface", "Menu / chỗ trên app", "Ý nghĩa", "QUYẾT ĐỊNH ← ĐIỀN", "Ghi chú"])
    dv = DataValidation(
        type="list",
        formula1='"IN MVP giấy,GĐ2,OUT,DEFER"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    for i, (inv, menu, meaning, hint) in enumerate(MISSING_18, 1):
        r = 3 + i
        for c, v in enumerate([i, inv, menu, meaning + f" ({hint})", "", ""], 1):
            cell = ws.cell(r, c, v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.fill = FILL_Y if c >= 5 else FILL_RO
        dv.add(ws.cell(r, 5))
        ws.row_dimensions[r].height = 36
    _widths(ws, [5, 12, 36, 40, 18, 28])
    ws.freeze_panes = "A4"


def sheet_lich(wb):
    ws = wb.create_sheet("03_UC_Lich_EXPAND", 3)
    _title(
        ws,
        "UC còn «Lịch» (SRS khung) — chọn EXPAND (viết đủ diễn biến) · GĐ2 · OUT · WAIVER. "
        "Chi tiết luồng: xem PDF SRS / UC flow.",
        "F",
    )
    _hdr(ws, 3, ["STT", "Mã UC", "Tên ngắn", "Khối", "QUYẾT ĐỊNH ← ĐIỀN", "Ghi chú"])
    dv = DataValidation(
        type="list",
        formula1='"EXPAND,GĐ2,OUT,WAIVER"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    for i, (ma, ten, khoi) in enumerate(LICH_UC, 1):
        r = 3 + i
        for c, v in enumerate([i, ma, ten, khoi, "", ""], 1):
            cell = ws.cell(r, c, v)
            cell.border = THIN
            cell.alignment = WRAP
            cell.fill = FILL_Y if c >= 5 else FILL_RO
        dv.add(ws.cell(r, 5))
        ws.row_dimensions[r].height = 28
    _widths(ws, [5, 18, 36, 10, 16, 32])
    ws.freeze_panes = "A4"


def sheet_guide(wb):
    ws = wb.create_sheet("Huong_dan", 4)
    ws["A1"] = "Hướng dẫn"
    ws["A1"].font = FONT_T
    lines = [
        "1. Sheet 00 = những gì anh đã chốt ở SPONSOR_CHOT_FILL.xlsx — chỉ đọc.",
        "2. Sheet 01 = câu còn thiếu số/thứ tự/phạm vi — điền cột vàng.",
        "3. Sheet 02 = 18 surface sâu chấm công — mỗi dòng một quyết định.",
        "4. Sheet 03 = UC còn «Lịch» — EXPAND hay để sau.",
        "5. Đọc PDF: SRS_HRM_ENTERPRISE_KHACH.pdf (luồng + sơ đồ) trước khi chốt sheet 03.",
        "6. Không điền KHACH_MOI. Save → báo PM.",
        "7. Demo ngày mai: trả lời R-DEMO-01 để team không claim sai.",
    ]
    for i, t in enumerate(lines, 3):
        ws.cell(i, 1, t)
    ws.column_dimensions["A"].width = 110


def build() -> Path:
    answers = []
    if ANSWERS.exists():
        answers = json.loads(ANSWERS.read_text(encoding="utf-8")).get("rows", [])
    wb = Workbook()
    # remove default
    default = wb.active
    wb.remove(default)
    sheet_done(wb, answers)
    sheet_remain(wb)
    sheet_missing(wb)
    sheet_lich(wb)
    sheet_guide(wb)
    try:
        wb.save(OUT)
        return OUT
    except PermissionError:
        wb.save(OUT_ALT)
        return OUT_ALT


if __name__ == "__main__":
    path = build()
    print(f"OK {path} kb={path.stat().st_size // 1024} remain_q={len(REMAIN_Q)} missing={len(MISSING_18)} lich={len(LICH_UC)}")
