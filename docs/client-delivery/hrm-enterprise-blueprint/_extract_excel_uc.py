# -*- coding: utf-8 -*-
"""Extract sheet 03 + 03b from WBS Excel for SRS enrichment."""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "WBS_HRM_ENTERPRISE_KHACH_MOI.xlsx"
OUT = ROOT / "_excel_uc_extract.json"


def cell(v):
    if v is None:
        return ""
    return str(v).strip()


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["03_Tinh_huong_nghiep_vu"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [cell(h) for h in rows[0]]
    situations = []
    for r in rows[1:]:
        if not any(r):
            continue
        d = {headers[i]: cell(r[i]) if i < len(r) else "" for i in range(len(headers))}
        situations.append(d)

    wsb = wb["03b_Dien_bien_chi_tiet"]
    brows = list(wsb.iter_rows(values_only=True))
    bheaders = [cell(h) for h in brows[0]]
    steps_by_uc: dict[str, list[dict]] = defaultdict(list)
    for r in brows[1:]:
        if not any(r):
            continue
        d = {bheaders[i]: cell(r[i]) if i < len(r) else "" for i in range(len(bheaders))}
        uc = d.get("Ma_UC") or d.get("Mã UC") or d.get("Ma UC") or ""
        if not uc:
            # try first col that looks like UC-
            for v in d.values():
                if v.startswith("UC-BP-"):
                    uc = v
                    break
        if not uc:
            continue
        steps_by_uc[uc].append(d)

    payload = {
        "headers_03": headers,
        "headers_03b": bheaders,
        "situations": situations,
        "steps_by_uc": steps_by_uc,
        "uc_count_03": len(situations),
        "uc_with_steps": len(steps_by_uc),
        "step_total": sum(len(v) for v in steps_by_uc.values()),
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print("wrote", OUT)
    print("03 rows", len(situations))
    print("03b uc", len(steps_by_uc), "steps", payload["step_total"])
    print("headers_03", headers)
    print("headers_03b", bheaders)
    if situations:
        print("sample_03 keys", list(situations[0].keys()))
        print("sample_03", {k: situations[0][k][:80] for k in list(situations[0])[:8]})
    sample_uc = next(iter(steps_by_uc))
    print("sample_uc", sample_uc, "n", len(steps_by_uc[sample_uc]))
    print("sample_step", steps_by_uc[sample_uc][0])


if __name__ == "__main__":
    main()
