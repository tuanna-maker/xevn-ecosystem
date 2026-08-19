# -*- coding: utf-8 -*-
"""Phiếu chốt sponsor — Excel điền được (cột vàng).

Output:
  - SPONSOR_CHOT_FILL.xlsx (file riêng — khuyến nghị mở cái này)
  - Sheet 00_Chot_Sponsor trong WBS_HRM_ENTERPRISE_UC_CHOT.xlsx (cùng gói workshop)
Không ghi vào WBS_HRM_ENTERPRISE_KHACH_MOI.xlsx.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
OUT_STANDALONE = ROOT / "SPONSOR_CHOT_FILL.xlsx"
OUT_UC_CHOT = ROOT / "WBS_HRM_ENTERPRISE_UC_CHOT.xlsx"
VERSION = "1.1"

# nhóm, mã, câu hỏi, gợi ý team, lựa chọn (dropdown text), placeholder quyết định
ROWS: list[tuple[str, str, str, str, str]] = [
    (
        "0. Chữ ký D7",
        "D7-1",
        "Đã chốt logic trên giấy (SRS v0.7 + WBS UC chốt) với khách / anh Nam chưa?",
        "Team chưa claim đã ký",
        "Chưa|Đã chốt nội bộ|Đã ký khách",
    ),
    (
        "0. Chữ ký D7",
        "D7-2",
        "Khi nào được mở lại code/demo trên sản phẩm?",
        "HOLD đến xác nhận giấy",
        "Giữ HOLD|Mở sau ngày … (ghi cột Ghi chú)|Mở ngay (ghi phạm vi ở Ghi chú)",
    ),
    (
        "0. Chữ ký D7",
        "D7-3",
        "Nếu mở code — phạm vi nào?",
        "—",
        "Chỉ sửa lỗi crash|Remaster UI|UC mới theo WBS|Khác (ghi Ghi chú)",
    ),
    (
        "1. Lương",
        "Q-PAY-FORMULA",
        "Ai soạn / ai phát hành công thức lương? Có cần 2 bước (nháp → hiệu lực)?",
        "Khuyến nghị: C&B soạn · kỹ thuật/đồng ký phát hành — xem DECISION_PACKET_Q_PAY_FORMULA.md",
        "Đồng ý 2 bước|C&B tự phát hành|Chỉ kỹ thuật DB|Khác (ghi Ghi chú)",
    ),
    (
        "1. Lương",
        "Q-PAY-F-2",
        "GĐ1 có cần kéo-thả công thức trên UI không?",
        "Team: GĐ1 form + xem trước; kéo-thả = GĐ2",
        "GĐ1 form|GĐ1 kéo-thả|Hoãn",
    ),
    (
        "1. Lương",
        "Q-PAY-F-3",
        "Biến số OT / phép vào lương lấy từ đâu?",
        "Họp đã khóa: chỉ bảng công đã chốt",
        "Xác nhận đúng (chỉ bảng công chốt)|Được lấy thêm nguồn (ghi Ghi chú)",
    ),
    (
        "2. Tuyển dụng",
        "Q-REC-HEADCOUNT",
        "Yêu cầu tuyển ngoài định biên xử lý thế nào?",
        "Họp: có luồng ngoài ĐB + BOD",
        "Cấm ngoài ĐB|Cho ngoài ĐB + duyệt BOD|Cho + duyệt khác (ghi Ghi chú)",
    ),
    (
        "2. Tuyển dụng",
        "Q-REC-HC-2",
        "Ai là nguồn sự thật số định biên theo phòng?",
        "Họp: kế hoạch HC / đề xuất",
        "HR tập đoàn|CEO công ty thành viên|Trưởng phòng + HR|Khác (ghi Ghi chú)",
    ),
    (
        "3. Nghỉ phép",
        "Q-LEAVE-ACCRUAL",
        "5 loại phép (năm · thâm niên · bù OT · chuyển kỳ · ứng) — cộng dồn/hết hạn theo lịch nào?",
        "Cần chốt trước khi viết FR sâu quy tắc nghỉ",
        "Năm dương lịch|Năm tài chính (ghi tháng bắt đầu ở Ghi chú)|Theo hợp đồng|Khác (ghi Ghi chú)",
    ),
    (
        "3. Nghỉ phép",
        "Q-LEAVE-UNIT",
        "Đơn vị tính phép mặc định?",
        "—",
        "Ngày|Giờ|Cả hai theo loại phép|Khác (ghi Ghi chú)",
    ),
    (
        "3. Nghỉ phép",
        "Q-LEAVE-3",
        "Ứng phép: trần tối đa / trừ kỳ sau thế nào?",
        "Trả lời tự do ở cột Quyết định hoặc Ghi chú",
        "(tự viết)",
    ),
    (
        "3. Nghỉ phép",
        "Q-LEAVE-4",
        "Nghỉ ốm có tách quỹ BHXH vs quỹ công ty không?",
        "Liên Q-SI-SUSPEND",
        "Có tách|Một quỹ|Khác (ghi Ghi chú)",
    ),
    (
        "4. BH / Hồ sơ / Tài sản",
        "Q-SI-SUSPEND",
        "Nghỉ không lương / tạm hoãn HĐ — BHXH tạm dừng xử lý trong HRM hay ngoài hệ thống?",
        "SRS còn mỏng",
        "Trong HRM|Ngoài (chỉ ghi chú)|GĐ2|Khác (ghi Ghi chú)",
    ),
    (
        "4. BH / Hồ sơ / Tài sản",
        "Q-ASSET-MODULE",
        "Module tài sản trên hồ sơ NV: MVP CRUD đầy đủ hay chỉ xem?",
        "SRS còn khung",
        "Chỉ xem|CRUD MVP|GĐ2|OUT",
    ),
    (
        "4. BH / Hồ sơ / Tài sản",
        "Q-XBOT-PROFILE",
        "Field động hồ sơ: XBOS tập đoàn hay HRM từng công ty định nghĩa?",
        "Team đề xuất: XBOS publish → HRM pull",
        "XBOS SoT|HRM SoT|Hybrid (ghi Ghi chú)",
    ),
    (
        "5. Chấm công",
        "Q-ATT-SIGN",
        "Ai được ký chốt bảng công?",
        "Có FR ký chốt",
        "QL trực tiếp|HR|Cả hai theo cấp|Khác (ghi Ghi chú)",
    ),
    (
        "5. Chấm công",
        "Q-ATT-SUMMARY",
        "Tab Tổng hợp = cùng dữ liệu bản ghi hay báo cáo API riêng?",
        "Product hiện = records (cần chốt trung thực)",
        "Chấp nhận = records GĐ1|Bắt buộc API riêng|Khác (ghi Ghi chú)",
    ),
    (
        "5. Chấm công",
        "Q-ATT-HOLIDAY",
        "Lịch lễ âm — cấu hình theo năm ở đâu?",
        "Chưa có màn riêng",
        "Trong ATT Settings|Catalog tập đoàn|GĐ2|Khác (ghi Ghi chú)",
    ),
    (
        "5. Chấm công",
        "Q-ATT-FACE",
        "Face ID check-in?",
        "Team: GĐ2 đã khóa họp",
        "Xác nhận GĐ2|Đưa vào MVP|OUT",
    ),
    (
        "5. Chấm công",
        "Q-ATT-MISSING",
        "18 bề mặt sâu (GPS site, thẻ QR NV, quỹ phép…) — giấy trước hay cho code P0?",
        "D7: giấy trước",
        "Chỉ giấy/WBS|Cho code P0 (liệt kê ở Ghi chú)|Hoãn hết",
    ),
    (
        "6. Gói gửi khách",
        "S1",
        "Gửi SRS PDF v0.7 + WBS UC chốt cho review logic?",
        "Team: READY gửi review",
        "Gửi|Chưa|Sửa trước (ghi Ghi chú)",
    ),
    (
        "6. Gói gửi khách",
        "S2",
        "Gap matrix nội bộ — gửi nguyên văn khách?",
        "Team: không gửi nguyên văn",
        "Giữ nội bộ|Tóm tắt 1 trang gửi khách",
    ),
    (
        "6. Gói gửi khách",
        "S3",
        "TechSpec/API/DB depth?",
        "HOLD đến khi chốt Q-* + giấy",
        "Giữ HOLD|Mở sớm (ghi rủi ro ở Ghi chú)",
    ),
    (
        "7. Master data",
        "MD-S1",
        "Loại phép: chỉ khóa 5 loại họp hay HR được thêm loại mới?",
        "Chi tiết REF/CFG: MASTER_DATA_CONFIG_CLASSIFICATION.md",
        "Chỉ 5 loại khóa|HR được thêm|Khác (ghi Ghi chú)",
    ),
    (
        "7. Master data",
        "MD-S2",
        "SoT ca làm việc: HRM work_shifts thắng; XBOS shifts chỉ tham chiếu?",
        "Team đề xuất đồng ý ADR",
        "Đồng ý|Khác (ghi Ghi chú)",
    ),
    (
        "7. Master data",
        "MD-S3",
        "Ai được thêm/sửa/xóa điểm GPS chấm công?",
        "—",
        "HR công ty thành viên|IT|QL vận hành|Khác (ghi Ghi chú)",
    ),
    (
        "7. Master data",
        "MD-S4",
        "Công thức lương — biến/phụ cấp lưu đâu khi thêm trên UI?",
        "Trùng Q-PAY-FORMULA — điền thêm nếu khác packet",
        "(xem Q-PAY-FORMULA)|Khác (ghi Ghi chú)",
    ),
    (
        "7. Master data",
        "MD-S5",
        "Số định biên lấy từ đâu khi tạo yêu cầu tuyển?",
        "—",
        "Kế hoạch HC|Org chart|Khác (ghi Ghi chú)",
    ),
    (
        "8. UC đề xuất (PROP)",
        "PROP-03d",
        "UC địa điểm GPS / bán kính (ADD SRS sau nếu chọn)?",
        "W3 propose-only — bắt buộc nếu giữ chấm GPS GĐ1",
        "IN MVP giấy|GĐ2|OUT|DEFER",
    ),
    (
        "8. UC đề xuất (PROP)",
        "PROP-03e",
        "UC thẻ QR nhân viên xem/tải/in?",
        "Bắt buộc nếu giữ kênh QR GĐ1",
        "IN MVP giấy|GĐ2|OUT|DEFER",
    ),
    (
        "8. UC đề xuất (PROP)",
        "PROP-05b",
        "UC panel quỹ phép (số dư · giữ chỗ · chuyển kỳ)?",
        "Siết/tách từ quỹ phép hiện có",
        "IN MVP giấy|GĐ2|OUT|DEFER",
    ),
    (
        "8. UC đề xuất (PROP)",
        "PM-HOLIDAY",
        "Màn cấu hình lịch lễ âm theo năm?",
        "PRODUCT_MISSING spine",
        "IN MVP giấy|GĐ2|OUT|DEFER",
    ),
    (
        "8. UC đề xuất (PROP)",
        "PM-OCR",
        "OCR / nhận diện chứng từ nghỉ (nếu có trong họp)?",
        "Chốt phạm vi giấy",
        "IN MVP giấy|GĐ2|OUT|DEFER",
    ),
    (
        "8. UC đề xuất (PROP)",
        "PM-SPLIT",
        "UI tách kỳ lương / chốt công tách kỳ?",
        "PRODUCT_MISSING / policy",
        "IN MVP giấy|GĐ2|OUT|DEFER",
    ),
    (
        "8. UC đề xuất (PROP)",
        "ATT-STUB",
        "Các màn STUB settings/roster (#17–18 · #37–46): giữ honesty GĐ1 hay bắt buộc LIVE trước go-live?",
        "Browser đã stamp STUB_UI",
        "Giữ stub GĐ1|Bắt buộc LIVE trước go-live|GĐ2|Khác (ghi Ghi chú)",
    ),
    (
        "9. Chính sách Lịch (SRS_THIN)",
        "LICH-ATT",
        "Nhóm UC «Lịch» chấm công/phép còn khung (~29): xử lý thế nào trước TechSpec?",
        "W3 SRS completeness FAIL — expand hoặc waiver",
        "EXPAND SRS trước TechSpec|Waiver GĐ1 (ghi phạm vi ở Ghi chú)|GĐ2|OUT",
    ),
    (
        "9. Chính sách Lịch (SRS_THIN)",
        "LICH-LEAVE",
        "Diễn biến sâu 5 loại phép + accrual: viết FR đủ 7 mục ngay hay sau chốt Q-LEAVE-*?",
        "Liên Q-LEAVE-ACCRUAL / UNIT",
        "Viết FR ngay sau phiếu này|Sau khi có Q-LEAVE|GĐ2|Waiver (ghi Ghi chú)",
    ),
    (
        "10. UI / thương hiệu",
        "UI-1",
        "Sau chốt giấy có mở remaster giao diện (Precision Motion) không?",
        "Chi tiết: SPONSOR_UI_BRAND_OPEN_QUESTIONS.md",
        "Có|Không|Chỉ login/shell|Tư vấn thêm trước",
    ),
    (
        "10. UI / thương hiệu",
        "UI-2",
        "Ưu tiên remaster màn nào trước?",
        "—",
        "Chấm công popup|Nhân sự|Lương|Login|Khác (ghi Ghi chú)",
    ),
    (
        "11. Xác nhận",
        "SIGN-1",
        "Họ và tên người chốt",
        "Điền cột Quyết định",
        "(tự viết)",
    ),
    (
        "11. Xác nhận",
        "SIGN-2",
        "Vai trò",
        "Điền cột Quyết định",
        "(tự viết)",
    ),
    (
        "11. Xác nhận",
        "SIGN-3",
        "Ngày chốt (dd/MM/yyyy)",
        "Điền cột Quyết định",
        "(tự viết)",
    ),
    (
        "11. Xác nhận",
        "SIGN-4",
        "Đã điền đủ P0 (D7 + Q-PAY-FORMULA + Q-REC-HEADCOUNT + Q-LEAVE-ACCRUAL + Q-LEAVE-UNIT + PROP-03d/03e/05b + LICH-ATT)?",
        "Chi tiết từng UC Lịch còn ở W3_PAPER_PACKET §5.1 / sheet 02b",
        "Đủ để PM cập nhật matrix|Chưa đủ (ghi còn thiếu ở Ghi chú)",
    ),
]

THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
FILL_HDR = PatternFill("solid", fgColor="1E40AF")
FILL_HINT = PatternFill("solid", fgColor="FEF3C7")  # vàng — điền đây
FILL_RO = PatternFill("solid", fgColor="F8FAFC")
FILL_TITLE = PatternFill("solid", fgColor="DBEAFE")
FONT_HDR = Font(bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(bold=True, color="1E3A8A", size=14)
WRAP = Alignment(wrap_text=True, vertical="center")


def _write_sheet(ws) -> None:
    ws["A1"] = (
        "PHIẾU CHỐT SPONSOR — chỉ điền cột G (vàng) và H (ghi chú). "
        f"Phiên bản {VERSION} · Không sửa cột A–F trừ khi team nhờ."
    )
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 36

    ws["A2"] = (
        "Cách dùng: mở bằng Excel / LibreOffice → gõ vào ô vàng → Save. "
        "File này cũng nằm trong WBS_HRM_ENTERPRISE_UC_CHOT.xlsx (sheet 00_Chot_Sponsor). "
        "KHÔNG điền vào WBS_HRM_ENTERPRISE_KHACH_MOI.xlsx."
    )
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 40

    headers = [
        "STT",
        "Nhóm",
        "Mã",
        "Câu hỏi",
        "Gợi ý team (chỉ đọc)",
        "Các lựa chọn gợi ý",
        "QUYẾT ĐỊNH CỦA ANH ← ĐIỀN ĐÂY",
        "Ghi chú thêm",
        "Ngày điền (dd/MM/yyyy)",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = FONT_HDR
        cell.fill = FILL_HDR
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN
    ws.row_dimensions[3].height = 32

    # Collect unique dropdown lists
    dropdown_map: dict[str, str] = {}
    for i, (nhom, ma, cau, goi_y, lua_chon) in enumerate(ROWS, 1):
        r = 3 + i
        values = [i, nhom, ma, cau, goi_y, lua_chon.replace("|", " · "), "", "", ""]
        for c, v in enumerate(values, 1):
            cell = ws.cell(r, c, v)
            cell.border = THIN
            cell.alignment = WRAP
            if c in (7, 8, 9):
                cell.fill = FILL_HINT
            else:
                cell.fill = FILL_RO
        ws.row_dimensions[r].height = 48
        if lua_chon != "(tự viết)" and "|" in lua_chon:
            # Excel DV list max ~255 chars for inline — use formula if long
            key = lua_chon
            if key not in dropdown_map:
                dropdown_map[key] = lua_chon

    # Apply data validations (one DV per unique list)
    for lua_chon, formula_list in dropdown_map.items():
        # openpyxl: comma-separated for list type, escape commas in items by using quotes carefully
        items = lua_chon.split("|")
        # Replace commas in items to avoid DV parse issues
        safe = [x.replace(",", " -") for x in items]
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(safe) + '"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error = "Chọn trong danh sách hoặc xóa rồi gõ tự do ở Ghi chú"
        dv.errorTitle = "Lựa chọn"
        dv.prompt = "Chọn hoặc gõ ở cột Ghi chú nếu chọn Khác"
        dv.promptTitle = "Quyết định"
        ws.add_data_validation(dv)
        for i, row in enumerate(ROWS):
            if row[4] == lua_chon:
                dv.add(ws.cell(4 + i, 7))

    widths = [5, 18, 16, 48, 40, 36, 36, 28, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{3 + len(ROWS)}"


def build_standalone() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Chot_Sponsor"
    _write_sheet(ws)
    # Guide sheet
    g = wb.create_sheet("Huong_dan")
    g["A1"] = "Hướng dẫn nhanh"
    g["A1"].font = FONT_TITLE
    lines = [
        "1. Chỉ điền các ô màu VÀNG (cột G, H, I).",
        "2. Cột G có dropdown — bấm mũi tên chọn; nếu «Khác» thì giải thích ở cột H.",
        "3. Ưu tiên điền đủ: D7-1..3 · Q-PAY-FORMULA · Q-REC-HEADCOUNT · Q-LEAVE-ACCRUAL · Q-LEAVE-UNIT · PROP-03d/03e/05b · LICH-ATT · SIGN-*.",
        "4. 18 dòng MISSING: điền thêm sheet 02b trong WBS_HRM_ENTERPRISE_UC_CHOT.xlsx (hoặc bảng §4.1 trong W3_PAPER_PACKET).",
        "5. File riêng: SPONSOR_CHOT_FILL.xlsx — hoặc sheet 00_Chot_Sponsor trong UC_CHOT.",
        "6. Không dùng WBS_HRM_ENTERPRISE_KHACH_MOI.xlsx để trả lời phiếu này.",
        "7. Sau khi Save → báo PM / gửi lại file.",
        "8. UI brand: SPONSOR_UI_BRAND_OPEN_QUESTIONS.md · Master data: MASTER_DATA_CONFIG_CLASSIFICATION.md.",
    ]
    for i, t in enumerate(lines, 3):
        g.cell(i, 1, t)
        g.cell(i, 1).alignment = WRAP
    g.column_dimensions["A"].width = 100
    try:
        wb.save(OUT_STANDALONE)
        return OUT_STANDALONE
    except PermissionError:
        alt = ROOT / "SPONSOR_CHOT_FILL_v1.1.xlsx"
        wb.save(alt)
        return alt


def merge_into_uc_chot() -> Path | None:
    if not OUT_UC_CHOT.exists():
        return None
    wb = load_workbook(OUT_UC_CHOT)
    if "00_Chot_Sponsor" in wb.sheetnames:
        del wb["00_Chot_Sponsor"]
    ws = wb.create_sheet("00_Chot_Sponsor", 0)
    _write_sheet(ws)
    # ensure Huong_dan only in standalone; optional tiny note on sheet title enough
    try:
        wb.save(OUT_UC_CHOT)
        return OUT_UC_CHOT
    except PermissionError:
        alt = ROOT / "WBS_HRM_ENTERPRISE_UC_CHOT_WITH_CHOT.xlsx"
        wb.save(alt)
        return alt


if __name__ == "__main__":
    p1 = build_standalone()
    p2 = merge_into_uc_chot()
    print(f"standalone: {p1} ({p1.stat().st_size // 1024} KB)")
    print(f"uc_chot: {p2}")
    print(f"rows: {len(ROWS)}")
    print(f"version: {VERSION}")
