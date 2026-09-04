import openpyxl
import json
import os
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\OneDrive\Tài liệu\Vibe Coding\projects\xevn-ecosystem\docs\từ khách hàng\Gửi P.CNTT\1. Điều phối hàng hóa\2025.07.30 Bảng lương BP ĐPHH.xlsx', data_only=True)
sheet = wb['BCC Điều phối']

# Group rows by MNV (Mã nhân viên)
emp_grouped = {}

for r in range(8, sheet.max_row + 1):
    mnv = sheet.cell(r, 2).value
    ma_phu = sheet.cell(r, 3).value
    name = sheet.cell(r, 4).value
    ngay_vao = sheet.cell(r, 6).value
    ngay_thu_viec = sheet.cell(r, 7).value
    status_str = sheet.cell(r, 8).value
    ma_vp = sheet.cell(r, 9).value
    ma_buu_cuc = sheet.cell(r, 10).value

    mnv_clean = str(mnv or '').strip()
    name_clean = str(name or '').strip()

    if not mnv_clean and not name_clean:
        continue

    # Days 1 to 31
    days_shifts = []
    for c in range(11, 42):
        val = sheet.cell(r, c).value
        days_shifts.append(str(val) if val is not None else '')

    row_data = {
        'row': r,
        'ma_phu': str(ma_phu or '').strip(),
        'ma_vp': str(ma_vp or '').strip(),
        'ma_buu_cuc': str(ma_buu_cuc or '').strip(),
        'status': str(status_str or '').strip(),
        'days': days_shifts,
        'tong_cong': sheet.cell(r, 42).value,
        'tong_gio': sheet.cell(r, 43).value
    }

    if mnv_clean not in emp_grouped:
        emp_grouped[mnv_clean] = {
            'mnv': mnv_clean,
            'name': name_clean,
            'status': str(status_str or '').strip(),
            'ngay_vao': str(ngay_vao or '').strip()[:10],
            'ngay_thu_viec': str(ngay_thu_viec or '').strip()[:10],
            'rows': []
        }
    emp_grouped[mnv_clean]['rows'].append(row_data)

# Remove stdout print to avoid Windows encoding error
# Done parsing

# Now build 31-day aggregated attendance for each employee
emp_attendance = []

for mnv, emp in emp_grouped.items():
    daily_records = []
    total_official_hours = 0
    total_probation_hours = 0
    total_ot_150 = 0
    total_work_days = 0
    total_leave_days = 0

    for d in range(1, 32):
        day_str = f"{d:02d}"
        date_str = f"2026-08-{day_str}"
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        is_sunday = dt.weekday() == 6

        # Check shift hours across all rows for this employee on day d
        active_hours = 0
        active_location = ''
        shift_raw = ''

        for r_item in emp['rows']:
            val = r_item['days'][d - 1].strip()
            if val and val != '0':
                shift_raw = val
                active_location = r_item['ma_vp'] or r_item['ma_buu_cuc']
                try:
                    active_hours = float(val)
                except ValueError:
                    if val.upper() in ['F', 'P', 'PHEP']:
                        shift_raw = 'F'
                break

        status = 'present'
        check_in = None
        check_out = None

        if shift_raw == 'F':
            status = 'leave'
            total_leave_days += 1
            check_in = f"{date_str}T08:00:00.000Z"
            check_out = f"{date_str}T17:00:00.000Z"
        elif active_hours > 0:
            status = 'present'
            total_work_days += 1
            check_in = f"{date_str}T08:00:00.000Z"
            end_h = 8 + int(active_hours)
            end_m = int(round((active_hours % 1) * 60))
            check_out = f"{date_str}T{end_h:02d}:{end_m:02d}:00.000Z"

            if emp['status'] == 'TV':
                total_probation_hours += min(active_hours, 8)
            else:
                total_official_hours += min(active_hours, 8)

            if active_hours > 8:
                total_ot_150 += (active_hours - 8)
        else:
            if is_sunday:
                status = 'off'
                check_in = f"{date_str}T08:00:00.000Z"
                check_out = f"{date_str}T17:00:00.000Z"
            else:
                # Default working weekday in DPHH roster if standard
                status = 'present'
                total_work_days += 1
                check_in = f"{date_str}T08:00:00.000Z"
                check_out = f"{date_str}T17:00:00.000Z"
                if emp['status'] == 'TV':
                    total_probation_hours += 8
                else:
                    total_official_hours += 8

        daily_records.append({
            'day': d,
            'date': date_str,
            'status': status,
            'check_in': check_in,
            'check_out': check_out,
            'active_hours': active_hours or (8 if status == 'present' else 0),
            'location': active_location
        })

    emp_attendance.append({
        'mnv': mnv,
        'name': emp['name'],
        'status': emp['status'],
        'total_work_days': total_work_days,
        'total_official_hours': total_official_hours,
        'total_probation_hours': total_probation_hours,
        'total_ot_150': total_ot_150,
        'total_leave_days': total_leave_days,
        'daily_records': daily_records
    })

os.makedirs('scratch', exist_ok=True)
with open('scratch/dphh_exact_attendance_data.json', 'w', encoding='utf-8') as f:
    json.dump(emp_attendance, f, ensure_ascii=False, indent=2)

# Finish saving
