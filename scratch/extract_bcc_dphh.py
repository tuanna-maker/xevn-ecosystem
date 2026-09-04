import openpyxl
import json
import os

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\OneDrive\Tài liệu\Vibe Coding\projects\xevn-ecosystem\docs\từ khách hàng\Gửi P.CNTT\1. Điều phối hàng hóa\2025.07.30 Bảng lương BP ĐPHH.xlsx', data_only=True)
sheet = wb['BCC Điều phối']

emp_rows = []
for r in range(8, sheet.max_row + 1):
    mnv = sheet.cell(r, 2).value
    ma_phu = sheet.cell(r, 3).value
    name = sheet.cell(r, 4).value
    status = sheet.cell(r, 8).value
    if mnv or name:
        days_data = [sheet.cell(r, c).value for c in range(11, 42)]
        emp_rows.append({
            'row': r,
            'mnv': str(mnv or '').strip(),
            'ma_phu': str(ma_phu or '').strip(),
            'name': str(name or '').strip(),
            'status': str(status or '').strip(),
            'days': [str(v) if v is not None else '' for v in days_data]
        })

os.makedirs('scratch', exist_ok=True)
with open('scratch/dphh_excel_bcc_rows.json', 'w', encoding='utf-8') as f:
    json.dump(emp_rows, f, ensure_ascii=False, indent=2)

print(f"Successfully saved {len(emp_rows)} rows to scratch/dphh_excel_bcc_rows.json")
