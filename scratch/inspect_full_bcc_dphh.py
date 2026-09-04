import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\OneDrive\Tài liệu\Vibe Coding\projects\xevn-ecosystem\docs\từ khách hàng\Gửi P.CNTT\1. Điều phối hàng hóa\2025.07.30 Bảng lương BP ĐPHH.xlsx', data_only=True)
sheet = wb['BCC Điều phối']

# Remove stdout print to avoid Windows encoding error

headers = []
for c in range(1, sheet.max_column + 1):
    h3 = str(sheet.cell(3, c).value or '').strip()
    h4 = str(sheet.cell(4, c).value or '').strip()
    h5 = str(sheet.cell(5, c).value or '').strip()
    h6 = str(sheet.cell(6, c).value or '').strip()
    h7 = str(sheet.cell(7, c).value or '').strip()
    headers.append({
        'col': c,
        'h3': h3,
        'h4': h4,
        'h5': h5,
        'h6': h6,
        'h7': h7
    })

with open('scratch/bcc_dieu_phoi_all_cols.json', 'w', encoding='utf-8') as f:
    json.dump(headers, f, ensure_ascii=False, indent=2)

# Inspect employee XE01480 (Phạm Việt Đức) and XE01683 (Vũ Anh Việt)
sample_rows = []
for r in range(8, 25):
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    sample_rows.append({
        'row': r,
        'stt': row_vals[0],
        'mnv': row_vals[1],
        'ma_phu': row_vals[2],
        'name': row_vals[3],
        'buu_cuc': row_vals[9],
        'days_1_31': row_vals[10:41],
        'tong_cong': row_vals[41],
        'tong_gio': row_vals[42],
        'cong_chuan': row_vals[43],
        'extra_cols': row_vals[44:]
    })

with open('scratch/bcc_dieu_phoi_samples.json', 'w', encoding='utf-8') as f:
    json.dump(sample_rows, f, ensure_ascii=False, indent=2)

print('Done inspecting!')
