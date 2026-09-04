import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel():
    wb = openpyxl.Workbook()
    
    # Header styles
    navy_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    dark_green_fill = PatternFill(start_color="276A3C", end_color="276A3C", fill_type="solid")
    white_bold_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # ----------------------------------------------------
    # Sheet 1: System & UAT Accounts
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Tài Khoản UAT & Demo"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1["A1"] = "DANH SÁCH TÀI KHOẢN DEMO & QUẢN LÝ TRÊN APP MOBILE"
    ws1["A1"].font = title_font
    ws1["A2"] = "Lưu ý: Có thể sử dụng Email HOẶC Số điện thoại để đăng nhập vào Mobile App"
    ws1["A2"].font = Font(name="Calibri", size=11, italic=True, color="595959")
    
    headers1 = ["STT", "Mã NV", "Họ và Tên", "Email Đăng Nhập", "Số Điện Thoại", "Mật Khẩu", "Vai Trò / Phân Quyền", "Ghi Chú"]
    ws1.append([]) # A3 empty
    ws1.append(headers1) # A4
    
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=col_num)
        cell.fill = navy_header_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    uat_accounts = [
        (1, "HLD-0001", "Nguyễn Văn An", "uat.nv0001@xe.vn", "0901234561", "xevn-uat-2026", "Nhân viên (Staff)", "Tài khoản nhân viên UAT mẫu"),
        (2, "TRS-0002", "UAT Quản lý", "uat.nv0002@xe.vn", "0901234562", "xevn-uat-2026", "Quản lý (Manager)", "Tài khoản Trưởng phòng có quyền Duyệt công / Phép"),
        (3, "HLD-0003", "UAT NV 0003", "uat.nv0003@xe.vn", "0901234563", "xevn-uat-2026", "Nhân viên (Staff)", "Tài khoản nhân viên UAT mẫu 3"),
        (4, "ADMIN01", "CEO Tập đoàn", "ceotnhhxe@xe.vn", "0901234560", "Portal@123", "Ban Giám Đốc (CEO)", "Tài khoản Ban Giám Đốc xem báo cáo toàn hệ thống")
    ]
    
    for row_idx, data in enumerate(uat_accounts, 5):
        for col_idx, val in enumerate(data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in (1, 2, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # ----------------------------------------------------
    # Sheet 2: 43 DPHH Accounts
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Phòng Điều Phối Hàng Hóa")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "DANH SÁCH TÀI KHOẢN ĐĂNG NHẬP MOBILE — PHÒNG ĐIỀU PHỐI HÀNG HÓA (43 NHÂN SỰ)"
    ws2["A1"].font = title_font
    ws2["A2"] = "Mật khẩu mặc định cho toàn bộ 43 nhân sự Điều Phối Hàng Hóa: xevn-pilot (Đăng nhập bằng Email hoặc SĐT)"
    ws2["A2"].font = Font(name="Calibri", size=11, italic=True, color="595959")
    
    headers2 = ["STT", "Mã NV", "Họ và Tên Nhân Sự", "Email Đăng Nhập Mobile", "Số Điện Thoại Đăng Nhập", "Mật Khẩu Đăng Nhập", "Phòng Ban", "Trạng Thái Data"]
    ws2.append([])
    ws2.append(headers2)
    
    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col_num)
        cell.fill = dark_green_fill
        cell.font = white_bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    dphh_list = [
        ("XE00036", "Đoàn Tiến Đạt", "xe00036@seed.xevn.local", "0980000036"),
        ("XE00047", "Trịnh Thị Vui", "xe00047@seed.xevn.local", "0980000047"),
        ("XE00061", "Nguyễn Thùy Dương", "xe00061@seed.xevn.local", "0980000061"),
        ("XE00073", "Nguyễn Trung Tú", "xe00073@seed.xevn.local", "0980000073"),
        ("XE00087", "Vũ Mạnh Đạt", "xe00087@seed.xevn.local", "0980000087"),
        ("XE00102", "Đặng Hải Đang", "xe00102@seed.xevn.local", "0980000102"),
        ("XE00106", "Vũ Thị Thu Hằng", "xe00106@seed.xevn.local", "0980000106"),
        ("XE00131", "Bùi Đức Mạnh", "xe00131@seed.xevn.local", "0980000131"),
        ("XE00136", "Nguyễn Công Việt", "xe00136@seed.xevn.local", "0980000136"),
        ("XE00182", "Phạm Đình Minh", "xe00182@seed.xevn.local", "0980000182"),
        ("XE00218", "Nguyễn Tùng Dương", "xe00218@seed.xevn.local", "0980000218"),
        ("XE00279", "Vũ Văn Nam", "xe00279@seed.xevn.local", "0980000279"),
        ("XE00331", "Trần Văn Đức", "xe00331@seed.xevn.local", "0980000331"),
        ("XE00332", "Trần Thanh Tuyền", "xe00332@seed.xevn.local", "0980000332"),
        ("XE00358", "Lê Hoàng Linh", "xe00358@seed.xevn.local", "0980000358"),
        ("XE00564", "Trương Quang Minh", "xe00564@seed.xevn.local", "0980000564"),
        ("XE00640", "Ngô Hồng Phong", "xe00640@seed.xevn.local", "0980000640"),
        ("XE00708", "Đỗ Huy Hoàng", "xe00708@seed.xevn.local", "0980000708"),
        ("XE00903", "Nguyễn Xuân Khôi", "xe00903@seed.xevn.local", "0980000903"),
        ("XE00963", "Nông Viết Phương", "xe00963@seed.xevn.local", "0980000963"),
        ("XE00968", "Lê Chí Công", "xe00968@seed.xevn.local", "0980000968"),
        ("XE00988", "Trần Bá Lợi", "xe00988@seed.xevn.local", "0980000988"),
        ("XE01113", "Vũ Duy Hoàng", "xe01113@seed.xevn.local", "0980001113"),
        ("XE01132", "Lê Hoài Nam", "xe01132@seed.xevn.local", "0980001132"),
        ("XE01134", "Phạm Quang Huy", "xe01134@seed.xevn.local", "0980001134"),
        ("XE01173", "Đào Trung Thiện", "xe01173@seed.xevn.local", "0980001173"),
        ("XE01174", "Vũ Hồng Sơn", "xe01174@seed.xevn.local", "0980001174"),
        ("XE01203", "Hoàng Quốc Hưng", "xe01203@seed.xevn.local", "0980001203"),
        ("XE01239", "Ngô Hoàng Anh", "xe01239@seed.xevn.local", "0980001239"),
        ("XE01324", "Nguyễn Trung Hiếu", "xe01324@seed.xevn.local", "0980001324"),
        ("XE01339", "Phạm Đức Trung", "xe01339@seed.xevn.local", "0980001339"),
        ("XE01480", "Phạm Việt Đức", "xe01480@seed.xevn.local", "0980001480"),
        ("XE01484", "Trần Quang Khải", "xe01484@seed.xevn.local", "0980001484"),
        ("XE01495", "Trần Đức Mạnh", "xe01495@seed.xevn.local", "0980001495"),
        ("XE01561", "Lê Văn Thuần", "xe01561@seed.xevn.local", "0980001561"),
        ("XE01571", "Nguyễn Hoàng Anh", "xe01571@seed.xevn.local", "0980001571"),
        ("XE01579", "Nguyễn Đức Long", "xe01579@seed.xevn.local", "0980001579"),
        ("XE01640", "Vũ Hải Sơn", "xe01640@seed.xevn.local", "0980001640"),
        ("XE01683", "Vũ Anh Việt", "xe01683@seed.xevn.local", "0980001683"),
        ("XE01686", "Nguyễn Trọng Nghĩa", "xe01686@seed.xevn.local", "0980001686"),
        ("XE01702", "Nguyễn Mạnh Tuấn", "xe01702@seed.xevn.local", "0980001702"),
        ("XE01751", "Phạm Quang Chính", "xe01751@seed.xevn.local", "0980001751"),
        ("XE01753", "Trần Huy Tú", "xe01753@seed.xevn.local", "0980001753")
    ]
    
    for idx, (code, name, email, phone) in enumerate(dphh_list, 1):
        row_data = (
            idx,
            code,
            name,
            email,
            phone,
            "xevn-pilot",
            "Phòng Điều Phối Hàng Hóa",
            "Đã chấm công chuẩn Excel 08/2026"
        )
        row_num = idx + 4
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in (1, 2, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for sheet in [ws1, ws2]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    target_dir = r"docs\từ khách hàng"
    os.makedirs(target_dir, exist_ok=True)
    file_path = os.path.join(target_dir, "Danh_Sach_Tai_Khoan_Dang_Nhap_Mobile.xlsx")
    wb.save(file_path)
    print(f"Successfully generated Excel file at: {file_path}")

if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    generate_excel()
