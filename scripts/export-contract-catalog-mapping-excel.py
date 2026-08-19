import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.Workbook()

# Styles
font_title = Font(name="Calibri", size=14, bold=True, color="1F4E78")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_regular = Font(name="Calibri", size=10)
font_code = Font(name="Consolas", size=10, bold=True, color="003366")
font_italic = Font(name="Calibri", size=10, italic=True, color="555555")

fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

thin_border_side = Side(border_style="thin", color="D9D9D9")
border_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Headers matching Form "Thêm mới điều khoản"
headers_clause = [
    "STT",
    "Mã điều khoản (code)",
    "Tiêu đề điều khoản (tiêu đề thuần)",
    "Thứ tự hiển thị khi in (Tự động đánh số Điều)",
    "Nhóm điều khoản (Tên nhóm hiển thị trên UI)",
    "Mã nhóm kỹ thuật (dùng cấu hình API / System Code)",
    "Gói áp dụng (Gói nghề)",
    "Bắt buộc (mandatory)",
    "Nội dung điều khoản đầy đủ (body_vi - Copy vào ô Textarea)"
]

# -------------------------------------------------------------
# SHEET 1: ĐIỀU KHOẢN KHỐI VĂN PHÒNG
# -------------------------------------------------------------
ws1 = wb.active
ws1.title = "Điều Khoản Khối Văn Phòng"
ws1.views.sheetView[0].showGridLines = True

ws1["A1"] = "THƯ VIỆN ĐIỀU KHOẢN HỢP ĐỒNG KHỐI VĂN PHÒNG (TÊN NHÓM TIẾNG VIỆT RÕ RÀNG)"
ws1["A1"].font = font_title

ws1.append([])
ws1.append(headers_clause)

for col_num, header in enumerate(headers_clause, 1):
    cell = ws1.cell(row=3, column=col_num)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

data_vp = [
    (
        1,
        "THOI_HAN_CONG_VIEC_VP",
        "Thời hạn và công việc hợp đồng",
        "Tự động gắn prefix 'Điều 1.' khi đưa vào Mẫu HĐ (vị trí 1)",
        "Căn cứ pháp lý / Thời hạn",
        "LEGAL_BASIS",
        "IT / văn phòng (IT_OFFICE)",
        "Có",
        """- Loại hợp đồng lao động: {{contract_type}}
- Từ ngày {{effective_date}} đến ngày {{expiration_date}}
- Địa điểm làm việc: {{work_location}} hoặc các địa điểm kinh doanh khác của Công ty theo yêu cầu công việc.
- Chức danh chuyên môn: {{position_name}}
- Chức vụ (nếu có): {{position_title}}
- Công việc phải làm: Thực hiện các công việc theo Bản mô tả công việc (JD) của chức danh {{position_name}} và theo sự phân công, điều động của Quản lý trực tiếp và Ban Giám đốc Công ty."""
    ),
    (
        2,
        "CHE_DO_LAM_VIEC_VP",
        "Chế độ làm việc",
        "Tự động gắn prefix 'Điều 2.' khi đưa vào Mẫu HĐ (vị trí 2)",
        "Thời giờ làm việc",
        "WORKING_HOURS",
        "IT / văn phòng (IT_OFFICE)",
        "Có",
        """- Thời gian làm việc: 06 ngày/tuần, 48 giờ/tuần (từ Thứ Hai đến Thứ Bảy). Thời gian làm việc cụ thể trong ngày áp dụng theo Quy chế/Nội quy lao động của Công ty.
- Dụng cụ làm việc: Được cấp phát trang thiết bị, dụng cụ làm việc theo chính sách hiện hành của Công ty và nhu cầu thực tế của vị trí công việc."""
    ),
    (
        3,
        "QUYEN_LOI_NGLD_VP",
        "Quyền lợi của Người lao động",
        "Tự động gắn prefix 'Điều 3.1.' khi đưa vào Mẫu HĐ (vị trí 3)",
        "Lương / đãi ngộ & Quyền lợi",
        "COMPENSATION",
        "IT / văn phòng (IT_OFFICE)",
        "Có",
        """- Phương tiện đi lại làm việc: Tự túc.
- Mức lương chính hoặc tiền công: {{base_salary}} VNĐ/tháng.
- Hình thức trả lương: Theo tháng, thanh toán bằng Tiền mặt hoặc Chuyển khoản qua Ngân hàng.
- Phụ cấp gồm: Được hưởng các khoản phụ cấp theo Chính sách về tiền lương và đãi ngộ hiện hành của Công ty.
- Kỳ hạn trả lương: Được trả lương định kỳ vào ngày 10 và ngày 25 của tháng kế tiếp hàng tháng.
- Tiền thưởng: Được hưởng tiền thưởng hàng năm dựa trên kết quả sản xuất kinh doanh của Công ty và mức độ hoàn thành nhiệm vụ của cá nhân theo Quy chế thưởng.
- Chế độ nâng lương: Thực hiện theo Quy chế nâng lương định kỳ hoặc đột xuất của Công ty.
- Trang bị bảo hộ lao động: Theo quy định về bảo hộ lao động hiện hành phù hợp với tính chất công việc.
- Chế độ nghỉ ngơi: Hưởng chế độ nghỉ hàng tuần, phép năm, nghỉ lễ tết theo quy định của Luật Lao động và Quy chế nội bộ của Công ty.
- Bảo hiểm Xã hội, Bảo hiểm Y tế, Bảo hiểm Thất nghiệp: Được trích đóng theo quy định của Luật Bảo hiểm Xã hội và Luật Bảo hiểm Y tế hiện hành.
- Chế độ đào tạo: Được tham gia các khóa đào tạo nâng cao trình độ chuyên môn nghiệp vụ theo chính sách đào tạo của Công ty.
- Những thỏa thuận khác: Người lao động có trách nhiệm giữ bí mật tất cả những thông tin liên quan đến việc sản xuất, kinh doanh, tổ chức và các thông tin bảo mật khác của Công ty mà Anh/Chị biết được trong quá trình làm việc. Anh/Chị không được tiết lộ thông tin đó với bất kỳ ai trong hay ngoài Công ty nếu không có sự chấp thuận trước bằng văn bản của Giám đốc Công ty. Trách nhiệm giữ bí mật thông tin này vẫn sẽ tiếp tục kể cả sau khi Anh/Chị nghỉ việc."""
    ),
    (
        4,
        "NGHIA_VU_NGLD_VP",
        "Nghĩa vụ của Người lao động",
        "Tự động gắn prefix 'Điều 3.2.' khi đưa vào Mẫu HĐ (vị trí 4)",
        "Công việc / nghĩa vụ",
        "JOB_DUTIES",
        "IT / văn phòng (IT_OFFICE)",
        "Có",
        """- Hoàn thành những công việc đã cam kết trong Hợp đồng lao động và trong Bản mô tả công việc (JD).
- Chấp hành việc thuyên chuyển, điều động công tác, lệnh điều hành sản xuất kinh doanh do Người sử dụng lao động quyết định.
- Tuân thủ nội dung kỷ luật lao động, an toàn lao động, an ninh trật tự, phòng cháy chữa cháy, các quy trình tiêu chuẩn và chuẩn mực văn hóa ứng xử của Công ty tại từng thời điểm.
- Báo cáo kịp thời với người có trách nhiệm khi phát hiện nguy cơ gây tai nạn lao động, độc hại hoặc sự cố nguy hiểm; tham gia cấp cứu và khắc phục hậu quả khi có lệnh của Người sử dụng lao động.
- Bồi thường thiệt hại vật chất đối với những công cụ lao động hoặc tài sản được giao cho cá nhân sử dụng mà để xảy ra mất mát, hư hỏng do lỗi chủ quan theo Luật Lao động.
- Thông báo kịp thời cho Phòng Nhân sự Công ty những thay đổi thông tin cá nhân như Căn cước công danh, địa chỉ thường trú, tạm trú, số điện thoại liên lạc.
- Thực hiện nghĩa vụ báo trước theo quy định của pháp luật nếu đơn phương chấm dứt hợp đồng lao động và thực hiện trách nhiệm bàn giao công việc/tài sản theo đúng quy trình của Công ty.
- Bồi thường chi phí đào tạo (nếu có) khi đơn phương chấm dứt hợp đồng lao động trái pháp luật."""
    ),
    (
        5,
        "NGHIA_VU_QUYEN_HAN_NSDLD",
        "Nghĩa vụ và quyền hạn của Người sử dụng lao động",
        "Tự động gắn prefix 'Điều 4.' khi đưa vào Mẫu HĐ (vị trí 5)",
        "Công việc / nghĩa vụ",
        "JOB_DUTIES",
        "Chung (GENERAL)",
        "Có",
        """1. Nghĩa vụ:
- Bảo đảm việc làm và thực hiện đầy đủ những điều đã cam kết trong Hợp đồng lao động này.
- Thanh toán đầy đủ, đúng thời hạn các chế độ và quyền lợi cho Người lao động theo Hợp đồng lao động, Phụ lục hợp đồng và Thỏa ước lao động tập thể (nếu có).
- Thực hiện nghĩa vụ báo trước theo quy định của pháp luật khi đơn phương chấm dứt hợp đồng lao động.

2. Quyền hạn:
- Điều hành Người lao động hoàn thành công việc theo Hợp đồng lao động (bố trí, điều chuyển, tạm ngừng việc theo nhu cầu sản xuất kinh doanh).
- Tạm hoãn, chấm dứt Hợp đồng lao động, kỷ luật Người lao động theo quy định của pháp luật và Nội quy lao động của Công ty.
- Xử lý kỷ luật lao động và yêu cầu Người lao động bồi thường trách nhiệm vật chất nếu vi phạm quy định của Công ty và pháp luật."""
    ),
    (
        6,
        "DIEU_KHOAN_THI_HANH",
        "Điều khoản thi hành",
        "Tự động gắn prefix 'Điều 5.' khi đưa vào Mẫu HĐ (vị trí 6)",
        "Giải quyết tranh chấp / Khác",
        "DISPUTE_LAW",
        "Chung (GENERAL)",
        "Có",
        """- Những vấn đề về lao động không ghi trong Hợp đồng lao động này thì áp dụng theo quy định của Bộ luật Lao động và các quy định nội bộ của Công ty.
- Hợp đồng này được làm thành 02 (hai) bản có giá trị pháp lý như nhau, mỗi bên giữ 01 (một) bản và có hiệu lực kể từ ngày ký."""
    )
]

for row_data in data_vp:
    ws1.append(row_data)

for r in range(4, 4 + len(data_vp)):
    for c in range(1, 10):
        cell = ws1.cell(row=r, column=c)
        cell.border = border_thin
        if c in (1, 2, 4, 5, 6, 7, 8):
            cell.alignment = align_center
        else:
            cell.alignment = align_left
        if c in (2, 6):
            cell.font = font_code
        elif c == 4:
            cell.font = font_italic
        else:
            cell.font = font_regular
        if r % 2 == 1:
            cell.fill = fill_zebra

# -------------------------------------------------------------
# SHEET 2: ĐIỀU KHOẢN KHỐI LÁI XE
# -------------------------------------------------------------
ws2 = wb.create_sheet(title="Điều Khoản Khối Lái Xe")
ws2.views.sheetView[0].showGridLines = True

ws2["A1"] = "THƯ VIỆN ĐIỀU KHOẢN HỢP ĐỒNG KHỐI LÁI XE (TÊN NHÓM TIẾNG VIỆT RÕ RÀNG)"
ws2["A1"].font = font_title

ws2.append([])
ws2.append(headers_clause)

for col_num, header in enumerate(headers_clause, 1):
    cell = ws2.cell(row=3, column=col_num)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

data_lx = [
    (
        1,
        "THOI_HAN_CONG_VIEC_LX",
        "Thời hạn và công việc hợp đồng (Lái xe)",
        "Tự động gắn prefix 'Điều 1.' khi đưa vào Mẫu HĐ (vị trí 1)",
        "Căn cứ pháp lý / Thời hạn",
        "LEGAL_BASIS",
        "Lái xe (DRIVER)",
        "Có",
        """- Loại hợp đồng lao động: {{contract_type}}
- Từ ngày {{effective_date}} đến ngày {{expiration_date}}
- Địa điểm làm việc: {{work_location}} (IC8 Phù Ninh, Phú Thọ / VP Nam Định / VP Việt Trì...) hoặc các luồng tuyến vận tải khác của Công ty.
- Chức danh chuyên môn: Nhân viên lái xe
- Giấy phép lái xe: Số {{driver_license_no}}, Hạng {{driver_license_class}}, Ngày cấp: {{driver_license_issue_date}}, Nơi cấp: {{driver_license_issuer}}.
- Công việc phải làm: Vận hành phương tiện vận tải, đưa đón hành khách/hàng hóa theo luồng tuyến, ca chạy và sự phân công của Bộ phận Điều hành xe và Trưởng chi nhánh."""
    ),
    (
        2,
        "CHE_DO_LAM_VIEC_LX",
        "Chế độ làm việc (Lái xe)",
        "Tự động gắn prefix 'Điều 2.' khi đưa vào Mẫu HĐ (vị trí 2)",
        "Thời giờ làm việc",
        "WORKING_HOURS",
        "Lái xe (DRIVER)",
        "Có",
        """- Thời gian làm việc: 06 ngày/tuần, 48 giờ/tuần (hoặc phân bổ theo ca chạy luồng tuyến vận tải do Điều hành sắp xếp).
- Cấp phát dụng cụ & trang thiết bị: Được bàn giao phương tiện xe, trang thiết bị an toàn, thẻ nghiệp vụ và các công cụ phục vụ công việc lái xe theo chính sách Công ty."""
    ),
    (
        3,
        "QUYEN_LOI_NGLD_LX",
        "Quyền lợi của Người lao động (Lái xe)",
        "Tự động gắn prefix 'Điều 3.1.' khi đưa vào Mẫu HĐ (vị trí 3)",
        "Lương / đãi ngộ & Quyền lợi",
        "COMPENSATION",
        "Lái xe (DRIVER)",
        "Có",
        """- Phương tiện đi lại làm việc: Tự túc.
- Mức lương chính hoặc tiền công: {{base_salary}} VNĐ/tháng.
- Hình thức trả lương: Trả theo tháng qua Chuyển khoản Ngân hàng hoặc Tiền mặt.
- Phụ cấp & Thưởng: Được hưởng Phụ cấp sắc điện, Phụ cấp tuyến, Phụ cấp ăn ca và các khoản thưởng hiệu suất an toàn theo Chính sách tiền lương Khối Lái xe.
- Kỳ hạn trả lương: Thanh toán 2 kỳ/tháng (vào ngày 10 và ngày 25 hàng tháng).
- Chế độ Bảo hiểm: Trích đóng BHXH, BHYT, BHTN theo quy định pháp luật đối với HĐLĐ chính thức.
- Trang bị bảo hộ & Chế độ nghỉ ngơi: Cấp phát đồng phục lái xe, hưởng chế độ nghỉ phép năm, lễ tết theo luật định."""
    ),
    (
        4,
        "NGHIA_VU_NGLD_LX",
        "Nghĩa vụ đặc thù của Lái xe",
        "Tự động gắn prefix 'Điều 3.2.' khi đưa vào Mẫu HĐ (vị trí 4)",
        "Phương tiện / An toàn lái xe",
        "DRIVER_VEHICLE",
        "Lái xe (DRIVER)",
        "Có",
        """- Hoàn thành đầy đủ công việc lái xe được giao theo lệnh điều xe và bản mô tả công việc.
- Nghiêm túc chấp hành Luật Giao thông đường bộ, tuân thủ tuyệt đối quy định an toàn giao thông, quy trình vận hành xe và chuẩn mực văn hóa ứng xử với hành khách.
- Kiểm tra an toàn kỹ thuật phương tiện trước và sau ca chạy; báo cáo kịp thời với Điều hành/Quản lý khi phát hiện nguy cơ sự cố, hỏng hóc hoặc nguy cơ tai nạn.
- Chịu trách nhiệm bồi thường thiệt hại vật chất đối với phương tiện, công cụ lao động hoặc tài sản được giao nếu để xảy ra mất mát, hư hỏng, tai nạn do lỗi chủ quan của lái xe.
- Thông báo kịp thời cho Phòng Nhân sự trước 30 ngày khi Giấy phép lái xe (GPLX) hoặc Căn cước công dân chuẩn bị hết hạn giá trị sử dụng.
- Bồi thường chi phí đào tạo và nghĩa vụ ký quỹ (nếu có) khi đơn phương chấm dứt hợp đồng lao động trái pháp luật."""
    ),
    (
        5,
        "NGHIA_VU_QUYEN_HAN_NSDLD_LX",
        "Nghĩa vụ và quyền hạn của NSDLĐ (Lái xe)",
        "Tự động gắn prefix 'Điều 4.' khi đưa vào Mẫu HĐ (vị trí 5)",
        "Công việc / nghĩa vụ",
        "JOB_DUTIES",
        "Lái xe (DRIVER)",
        "Có",
        """1. Nghĩa vụ:
- Bảo đảm việc làm, điều xe hợp lý và thanh toán đầy đủ các chế độ lương, phụ cấp đúng thời hạn.
- Thực hiện đầy đủ nghĩa vụ báo trước khi chấm dứt hợp đồng theo quy định pháp luật.

2. Quyền hạn:
- Điều hành, bố trí ca chạy, luồng tuyến và điều chuyển tạm thời vị trí công tác theo nhu cầu vận tải của Công ty.
- Trừ lương/khấu trừ ký quỹ hoặc tạm hoãn HĐLĐ, xử lý kỷ luật nếu lái xe vi phạm Luật Giao thông, gây tai nạn hoặc vi phạm nghiêm trọng nội quy an toàn vận tải."""
    ),
    (
        6,
        "DIEU_KHOAN_THI_HANH_LX",
        "Điều khoản thi hành (Lái xe)",
        "Tự động gắn prefix 'Điều 5.' khi đưa vào Mẫu HĐ (vị trí 6)",
        "Giải quyết tranh chấp / Khác",
        "DISPUTE_LAW",
        "Lái xe (DRIVER)",
        "Có",
        """- Những vấn đề không ghi trong Hợp đồng này được áp dụng theo quy định của Bộ luật Lao động và Quy chế quản lý đội xe của Công ty.
- Hợp đồng lập thành 02 (hai) bản có giá trị pháp lý như nhau, mỗi bên giữ 01 (một) bản và có hiệu lực từ ngày ký."""
    )
]

for row_data in data_lx:
    ws2.append(row_data)

for r in range(4, 4 + len(data_lx)):
    for c in range(1, 10):
        cell = ws2.cell(row=r, column=c)
        cell.border = border_thin
        if c in (1, 2, 4, 5, 6, 7, 8):
            cell.alignment = align_center
        else:
            cell.alignment = align_left
        if c in (2, 6):
            cell.font = font_code
        elif c == 4:
            cell.font = font_italic
        else:
            cell.font = font_regular
        if r % 2 == 1:
            cell.fill = fill_zebra

# -------------------------------------------------------------
# SHEET 3: DANH MỤC NHÓM ĐIỀU KHOẢN (TRA CỨU TRÊN MÀN HÌNH)
# -------------------------------------------------------------
ws3 = wb.create_sheet(title="Danh Mục Nhóm Điều Khoản")
ws3.views.sheetView[0].showGridLines = True

ws3["A1"] = "DANH MỤC BẢNG MÃ & TÊN HÌNH THỨC NHÓM ĐIỀU KHOẢN TRÊN GIAO DIỆN PHẦN MỀM"
ws3["A1"].font = font_title

headers_cat = ["STT", "Tên Nhóm Hiển Thị Tiếng Việt (Chọn trên Menu UI)", "Mã Nhóm Kỹ Thuật (System Code)", "Mô Tả Ứng Dụng Trong Hợp Đồng"]
ws3.append([])
ws3.append(headers_cat)

for col_num, header in enumerate(headers_cat, 1):
    cell = ws3.cell(row=3, column=col_num)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

groups_catalog = [
    (1, "Căn cứ pháp lý", "LEGAL_BASIS", "Nhóm chứa căn cứ Bộ luật lao động, Luật doanh nghiệp, các quyết định thành lập"),
    (2, "Bên A / Bên B", "PARTIES", "Thông tin thông số đại diện người sử dụng lao động và thông tin cá nhân người lao động"),
    (3, "Công việc / nghĩa vụ", "JOB_DUTIES", "Mô tả công việc phải làm, chức danh chuyên môn, nhiệm vụ và nghĩa vụ cơ bản"),
    (4, "Thời hạn / thử việc", "TERM_PROBATION", "Loại hợp đồng, ngày bắt đầu/kết thúc, thời gian thử việc"),
    (5, "Lương / đãi ngộ", "COMPENSATION", "Mức lương chính, phụ cấp, thưởng, kỳ hạn trả lương, ngân hàng"),
    (6, "Nâng bậc / nâng lương", "GRADE_RAISE", "Quy chế xét nâng lương định kỳ hoặc đột xuất theo hiệu suất"),
    (7, "Thời giờ làm việc", "WORKING_HOURS", "Số ngày/tuần, số giờ/tuần, ca chạy luồng tuyến, thời giờ nghỉ ngơi"),
    (8, "Bảo hộ lao động", "PPE", "Trang bị bảo hộ, đồng phục, dụng cụ làm việc"),
    (9, "BHXH / BHYT / BHTN", "SOCIAL_INSURANCE", "Trích đóng các khoản bảo hiểm bắt buộc theo luật định"),
    (10, "Đào tạo", "TRAINING", "Quyền lợi đào tạo nâng cao tay nghề và cam kết bồi thường chi phí đào tạo"),
    (11, "Bảo mật", "NDA_TRADE_SECRET", "Cam kết bảo mật thông tin kinh doanh, bí mật công nghệ của công ty"),
    (12, "Sở hữu trí tuệ", "IP_WORK_PRODUCT", "Quyền sở hữu đối với các sản phẩm sáng tạo, sáng kiến trong quá trình làm việc"),
    (13, "Thiết bị CNTT", "IT_EQUIPMENT", "Bàn giao máy tính, tài khoản email, trang thiết bị công nghệ văn phòng"),
    (14, "Phương tiện / GPLX", "DRIVER_VEHICLE", "Quy định đặc thù về Giấy phép lái xe, hạng bằng, phương tiện vận tải"),
    (15, "An toàn / cấm rượu bia", "DRIVER_SAFETY_ALCOHOL", "Quy định nồng độ cồn bằng 0, an toàn giao thông đường bộ đối với lái xe"),
    (16, "Trách nhiệm TNGT", "DRIVER_LIABILITY", "Trách nhiệm vật chất bồi thường sự cố tai nạn giao thông do lỗi chủ quan"),
    (17, "Chấm dứt HĐ", "TERMINATION_GENERAL", "Quy định nghĩa vụ báo trước, nghĩa vụ bàn giao và chấm dứt HĐLĐ"),
    (18, "Giải quyết tranh chấp", "DISPUTE_LAW", "Thỏa thuận cơ quan tài chính/tòa án giải quyết khi xảy ra tranh chấp HĐ")
]

for g_row in groups_catalog:
    ws3.append(g_row)

for r in range(4, 4 + len(groups_catalog)):
    for c in range(1, 5):
        cell = ws3.cell(row=r, column=c)
        cell.border = border_thin
        if c in (1, 3):
            cell.alignment = align_center
        else:
            cell.alignment = align_left
        if c == 3:
            cell.font = font_code
        else:
            cell.font = font_regular
        if r % 2 == 1:
            cell.fill = fill_zebra

# Set column widths
for ws in [ws1, ws2]:
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 90

ws3.column_dimensions["A"].width = 6
ws3.column_dimensions["B"].width = 35
ws3.column_dimensions["C"].width = 28
ws3.column_dimensions["D"].width = 70

output_file_name = r"docs/từ khách hàng/Bóc_Tách_Điều_Khoản_Hợp_Đồng_XE_FULL_TEN_NHOM.xlsx"
try:
    wb.save(output_file_name)
    print(f"EXCEL_REBUILT_SUCCESS: {output_file_name}")
except Exception as e:
    alt_file = r"docs/từ khách hàng/Bóc_Tách_Điều_Khoản_Hợp_Đồng_XE_FULL_TEN_NHOM_v2.xlsx"
    wb.save(alt_file)
    print(f"EXCEL_REBUILT_SUCCESS: {alt_file}")
