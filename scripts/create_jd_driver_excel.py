import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: JD_Templates (Template chính)
# ============================================================
ws1 = wb.active
ws1.title = "JD_Templates"

# Header style
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = [
    "code", "title", "position_code", "position_name",
    "job_description", "requirements", "notes", "status", "layout_version"
]

for col_idx, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Data rows
jd_data = [
    {
        "code": "JD-DRIVER-001",
        "title": "Lái xe tải trọng 1.5-3.5 tấn",
        "position_code": "DRV_TRUCK_1_5_3_5",
        "position_name": "Lái xe tải trọng 1.5-3.5 tấn",
        "job_description": (
            "Mô tả công việc:\n"
            "- Lái xe vận chuyển hàng hóa theo tuyến đường chỉ định\n"
            "- Kiểm tra xe trước/đi sau chuyến (lốp, dầu, đèn, phanh)\n"
            "- Báo cáo sự cố kịp thời cho quản lý đội xe\n"
            "- Tuân thủ quy trình an toàn giao thông, giới hạn tốc độ\n"
            "- Ghi nhật ký hành trình, nhiên liệu, bảo dưỡng"
        ),
        "requirements": (
            "Yêu cầu bắt buộc:\n"
            "- Bằng lái xe hạng B2, C (tùy tải trọng)\n"
            "- Giấy phép lái xe còn hạn sử dụng\n"
            "- Kinh nghiệm lái xe tải ≥ 2 năm\n"
            "- Sạch tiền án, tiền sự, sức khỏe tốt\n"
            "- Có thẻ BHXH, BHYT\n\n"
            "Yêu cầu ưu tiên:\n"
            "- Có kinh nghiệm chạy đường dài Hà Nội - Bắc Giang/Quảng Ninh\n"
            "- Biết sửa chữa cơ bản xe (thay lốp, kiểm tra dầu)\n"
            "- Ưu tiên đã từng làm cho công ty logistics"
        ),
        "notes": (
            "- Thử việc 1 tháng (lương 100%)\n"
            "- Có chế độ ăn trưa tại công ty\n"
            "- Xe được bảo dưỡng định kỳ hàng tháng"
        ),
        "status": "active",
        "layout_version": 2
    },
    {
        "code": "JD-DRIVER-002",
        "title": "Lái xe container 20/40ft",
        "position_code": "DRV_CONTAINER",
        "position_name": "Lái xe container",
        "job_description": (
            "Mô tả công việc:\n"
            "- Lái xe đầu kéo container 20ft/40ft cảng Hải Phòng/Cái Mép - các ICD/kho nội địa\n"
            "- Thao tác gác/khóa container chuẩn ISO\n"
            "- Điều phối với cảng, bãi đỗ, kho bãi về lịch nhận/trả container\n"
            "- Tuân thủ quy trình VGM (Verified Gross Mass)"
        ),
        "requirements": (
            "Yêu cầu bắt buộc:\n"
            "- Bằng lái xe hạng C, E\n"
            "- Kinh nghiệm lái container ≥ 3 năm\n"
            "- Biết thao tác gác container, khóa twistlock\n"
            "- Có thẻ ICC (International Carrier Card) hoặc đang làm hồ sơ\n"
            "- Không vi phạm an toàn giao thông 12 tháng gần nhất"
        ),
        "notes": (
            "- Lương cước theo chuyến + thưởng an toàn\n"
            "- Hỗ trợ ăn nghỉ khi chờ bến\n"
            "- Được cấp quần áo mưa, đũa an toàn"
        ),
        "status": "active",
        "layout_version": 2
    },
    {
        "code": "JD-DRIVER-003",
        "title": "Lái xe buýt tuyến cố định",
        "position_code": "DRV_BUS_FIXED",
        "position_name": "Lái xe buýt tuyến cố định",
        "job_description": (
            "Mô tả công việc:\n"
            "- Lái xe buýt theo tuyến đường, giờ chạy cố định\n"
            "- Đón/trả khách tại các trạm quy định\n"
            "- Kiểm tra xe hàng ngày (phanh, đèn, cửa, còi)\n"
            "- Báo cáo sự cố, tai nạn ngay lập tức\n"
            "- Thu vé/hỗ trợ kiểm soát vé (nếu tuyến không có nhân viên thu)"
        ),
        "requirements": (
            "Yêu cầu bắt buộc:\n"
            "- Bằng lái xe hạng D (xe chở người > 30 chỗ)\n"
            "- Giấy khám sức khỏe định kỳ 6 tháng (theo QCVN)\n"
            "- Kinh nghiệm lái xe khách ≥ 2 năm\n"
            "- Không tiền án, vi phạm ATGT nghiêm trọng\n"
            "- Thái độ phục vụ khách hàng tốt"
        ),
        "notes": (
            "- Lương cứng + thưởng an toàn + thưởng Tết\n"
            "- Đóng BHXH/BHYT/BHTN ngày vào làm\n"
            "- Có chế độ nghỉ năm, nghỉ bệnh theo luật\n"
            "- Xe mới, điều hòa, GPS giám sát"
        ),
        "status": "active",
        "layout_version": 2
    }
]

# Write data rows
wrap_alignment = Alignment(wrap_text=True, vertical="top")
for row_idx, jd in enumerate(jd_data, 2):
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=jd[header])
        cell.alignment = wrap_alignment
        cell.border = thin_border

# Set column widths
col_widths = [18, 30, 22, 30, 55, 55, 45, 12, 16]
for i, width in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# Freeze header row
ws1.freeze_panes = "A2"

# ============================================================
# SHEET 2: JD_Dynamic_Fields (Thông tin chi tiết bổ sung cho từng mẫu JD)
# ============================================================
ws2 = wb.create_sheet("JD_Dynamic_Fields")

# Header đầy đủ: mã trường kỹ thuật + tên hiển thị + kiểu dữ liệu + danh sách lựa chọn + giá trị mẫu + hướng dẫn
headers2 = [
    "Mã JD mẫu",
    "Mã trường (key)",       # field_code - key kỹ thuật cho dev/API
    "Tên hiển thị",           # tên tiếng Việt cho người dùng
    "Kiểu dữ liệu",           # text, number, select, date, boolean
    "Danh sách lựa chọn",     # cho select: các giá trị cho phép (VD: B2|C|D|E)
    "Giá trị mẫu",            # ví dụ giá trị
    "Ghi chú / Hướng dẫn điền"
]
for col_idx, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dữ liệu: mỗi dòng = 1 thông tin chi tiết của 1 mẫu JD
# field_code = key kỹ thuật (snake_case) dùng trong values_json
# field_type = text|number|select|date|boolean
# field_options = danh sách giá trị cho select, phân cách bằng |
dynamic_fields = [
    # JD-DRIVER-001: Lái xe tải trọng 1.5-3.5 tấn
    {"template_code": "JD-DRIVER-001", "field_code": "license_type", "field_key": "Loại bằng lái xe", "field_type": "select", "field_options": "B2|C|D|E", "field_value": "B2, C", "field_note": "Chọn loại bằng phù hợp tải trọng xe. Ví dụ: B2 (xe tải ≤3.500kg), C (xe tải >3.500kg), D (xe khách >30 chỗ), E (đầu kéo rơ-moóc). Nếu xe 1.5-3.5 tấn thì bằng B2 đủ, nhưng khuyến nghị C để linh hoạt."},
    {"template_code": "JD-DRIVER-001", "field_code": "experience_years", "field_key": "Số năm kinh nghiệm", "field_type": "number", "field_options": "", "field_value": "2", "field_note": "Số năm tối thiểu đã lái xe tải. Điền số nguyên (ví dụ: 2, 3, 5)."},
    {"template_code": "JD-DRIVER-001", "field_code": "route_area", "field_key": "Khu vực chạy", "field_type": "text", "field_options": "", "field_value": "Nội tỉnh Hà Nội - Bắc Giang/Quảng Ninh", "field_note": "Mô tả tuyến đường/hành trình chính. Ví dụ: 'Nội thành Hà Nội', 'Hà Nội - Hải Phòng', 'Tuyến Bắc - Nam'."},
    {"template_code": "JD-DRIVER-001", "field_code": "shift_pattern", "field_key": "Ca làm việc / Đòn bẩy", "field_type": "text", "field_options": "", "field_value": "2 ngày/1 nghỉ", "field_note": "Cách sắp xếp ca. Ví dụ: '2 ngày làm 1 ngày nghỉ', '1 ca/ngày (sáng)', 'Theo đơn hàng', 'Đơn hàng fixed'."},
    {"template_code": "JD-DRIVER-001", "field_code": "vehicle_type", "field_key": "Loại xe điều khiển", "field_type": "text", "field_options": "", "field_value": "Thùng/Furgon 1.5-3.5 tấn", "field_note": "Dòng xe cụ thể. Ví dụ: 'Xe thùng 2 tấn', 'Furgon 3.5 tấn', 'Xe ben 5 tấn'."},
    {"template_code": "JD-DRIVER-001", "field_code": "salary_range", "field_key": "Mức lương tham khảo", "field_type": "text", "field_options": "", "field_value": "12-15 triệu/tháng", "field_note": "Khoảng lương gross/tháng. Có thể ghi: '12-15 triệu', 'Theo cước 400k/chuyến', 'Thương lượng'."},
    {"template_code": "JD-DRIVER-001", "field_code": "benefits", "field_key": "Quyền lợi được hưởng", "field_type": "text", "field_options": "", "field_value": "BHXH/BHYT/BHTN, ăn trưa, thưởng Tết, xe bảo dưỡng định kỳ", "field_note": "Liệt kê phúc lợi. Ví dụ: 'Đầy đủ BHXH/BHYT/BHTN', 'Ăn trưa tại công ty', 'Thưởng Tết 1 tháng lương', 'Xe được bảo dưỡng hằng tháng'."},
    {"template_code": "JD-DRIVER-001", "field_code": "probation_months", "field_key": "Thời gian thử việc (tháng)", "field_type": "number", "field_options": "", "field_value": "1", "field_note": "Số tháng thử việc theo Luật Lao động. Thường là 1-2 tháng. Điền số nguyên."},

    # JD-DRIVER-002: Lái xe container 20/40ft
    {"template_code": "JD-DRIVER-002", "field_code": "license_type", "field_key": "Loại bằng lái xe", "field_type": "select", "field_options": "C|E", "field_value": "C, E", "field_note": "Container yêu cầu bằng C (xe tải nặng) hoặc E (đầu kéo rơ-moóc). Bắt buộc có bằng E cho đầu kéo container."},
    {"template_code": "JD-DRIVER-002", "field_code": "experience_years", "field_key": "Số năm kinh nghiệm", "field_type": "number", "field_options": "", "field_value": "3", "field_note": "Kinh nghiệm lái container/đầu kéo. Thường yêu cầu ≥3 năm do tính đặc thù an toàn cảng."},
    {"template_code": "JD-DRIVER-002", "field_code": "route_area", "field_key": "Khu vực chạy", "field_type": "text", "field_options": "", "field_value": "Cảng Hải Phòng/Cái Mép - ICD nội địa", "field_note": "Tuyến cảng - ICD/kho. Ví dụ: 'Cảng Hải Phòng - ICD Sóng Thần', 'Cái Mép - ICD Long Bình'."},
    {"template_code": "JD-DRIVER-002", "field_code": "shift_pattern", "field_key": "Ca làm việc / Đòn bẩy", "field_type": "text", "field_options": "", "field_value": "Theo lịch bến cảng", "field_note": "Ca phụ thuộc lịch nhận/trả container tại cảng. Có thể làm đêm, cuối tuần. Ghi: 'Theo lịch bến', '24/7 sẵn sàng'."},
    {"template_code": "JD-DRIVER-002", "field_code": "vehicle_type", "field_key": "Loại xe điều khiển", "field_type": "text", "field_options": "", "field_value": "Đầu kéo container 20/40ft", "field_note": "Chỉ định đầu kéo (tractor head) kéo container. Ví dụ: 'Đầu kéo 4x2', 'Đầu kéo 6x4', 'Scania/Volvo/Howo'."},
    {"template_code": "JD-DRIVER-002", "field_code": "salary_range", "field_key": "Mức lương tham khảo", "field_type": "text", "field_options": "", "field_value": "15-20 triệu/tháng (theo cước)", "field_note": "Container thường tính theo cước/chuyến. Ghi rõ: '15-20 triệu', '450k/chuyến Cái Mép - ICD', 'Có thưởng an toàn'."},
    {"template_code": "JD-DRIVER-002", "field_code": "benefits", "field_key": "Quyền lợi được hưởng", "field_type": "text", "field_options": "", "field_value": "BHXH/BHYT/BHTN, ăn nghỉ chờ bến, quần áo PPE, thưởng an toàn", "field_note": "Đặc thù container: có hỗ trợ ăn/nghỉ khi chờ bến, cấp PPE (mũ, áo phản quang, đũa an toàn), thưởng 0 tai nạn."},
    {"template_code": "JD-DRIVER-002", "field_code": "icc_card", "field_key": "Thẻ ICC (Quốc tế)", "field_type": "select", "field_options": "Đã có|Đang làm hồ sơ|Yêu cầu công ty hỗ trợ|Không cần", "field_value": "Yêu cầu", "field_note": "Thẻ ICC (International Carrier Card) bắt buộc cho xe ra cảng quốc tế. Chọn: 'Đã có', 'Đang làm hồ sơ', 'Yêu cầu công ty hỗ trợ', 'Không cần'."},

    # JD-DRIVER-003: Lái xe buýt tuyến cố định
    {"template_code": "JD-DRIVER-003", "field_code": "license_type", "field_key": "Loại bằng lái xe", "field_type": "select", "field_options": "D", "field_value": "D", "field_note": "Bắt buộc bằng D (xe chở người >30 chỗ). Không chấp nhận bằng B2/C."},
    {"template_code": "JD-DRIVER-003", "field_code": "experience_years", "field_key": "Số năm kinh nghiệm", "field_type": "number", "field_options": "", "field_value": "2", "field_note": "Kinh nghiệm lái xe khách. Ưu tiên có kinh nghiệm tuyến cố định."},
    {"template_code": "JD-DRIVER-003", "field_code": "route_area", "field_key": "Khu vực chạy", "field_type": "text", "field_options": "", "field_value": "Tuyến cố định nội thành", "field_note": "Tuyến đường cụ thể. Ví dụ: 'Tuyến 01: Bến xe Giáp Bát - Bến xe Mỹ Đình', 'Tuyến nội huyện'."},
    {"template_code": "JD-DRIVER-003", "field_code": "shift_pattern", "field_key": "Ca làm việc / Đòn bẩy", "field_type": "select", "field_options": "2 ca/ngày (sáng/chiều)|1 ca/ngày (sáng)|1 ca/ngày (chiều)|Ca đêm|Linh hoạt", "field_value": "2 ca/ngày (sáng/chiều)", "field_note": "Xe buýt thường chạy 2 ca: sáng (5h-11h) + chiều (13h-19h). Chọn: '2 ca/ngày (sáng/chiều)', '1 ca/ngày (sáng)', '1 ca/ngày (chiều)', 'Ca đêm', 'Linh hoạt'."},
    {"template_code": "JD-DRIVER-003", "field_code": "vehicle_type", "field_key": "Loại xe điều khiển", "field_type": "text", "field_options": "", "field_value": "Buýt 30-45 chỗ", "field_note": "Số chỗ ngồi. Ví dụ: 'Buýt 32 chỗ', 'Buýt 45 chỗ (Samco/Thaco/Hyundai)'."},
    {"template_code": "JD-DRIVER-003", "field_code": "salary_range", "field_key": "Mức lương tham khảo", "field_type": "text", "field_options": "", "field_value": "10-13 triệu/tháng", "field_note": "Lương cứng + thưởng. Ghi: '10-13 triệu (lương cứng)', 'Có thưởng an toàn 500k/tháng', 'Thưởng Tết 1.5 tháng'."},
    {"template_code": "JD-DRIVER-003", "field_code": "benefits", "field_key": "Quyền lợi được hưởng", "field_type": "text", "field_options": "", "field_value": "BHXH/BHYT/BHTN đầy đủ, lương cứng, thưởng an toàn, nghỉ năm", "field_note": "Phúc lợi xe khách: lương ổn định, thưởng an toàn hàng tháng, nghỉ năm theo luật, xe mới có điều hòa/GPS."},
    {"template_code": "JD-DRIVER-003", "field_code": "health_check_interval_months", "field_key": "Khoảng cách khám sức khỏe (tháng)", "field_type": "number", "field_options": "", "field_value": "6", "field_note": "Theo QCVN, lái xe khách phải khám sức khỏe định kỳ 6 tháng/lần. Ghi: '6 tháng (theo quy định)', 'Công ty chi phí'."},
]

for row_idx, row_data in enumerate(dynamic_fields, 2):
    # Map Vietnamese headers to dict keys
    key_map = {
        "Mã JD mẫu": "template_code",
        "Mã trường (key)": "field_code",
        "Tên hiển thị": "field_key",
        "Kiểu dữ liệu": "field_type",
        "Danh sách lựa chọn": "field_options",
        "Giá trị mẫu": "field_value",
        "Ghi chú / Hướng dẫn điền": "field_note",
    }
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=row_data[key_map[header]])
        cell.alignment = wrap_alignment
        cell.border = thin_border

col_widths2 = [22, 22, 30, 18, 40, 40, 70]
for i, width in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

ws2.freeze_panes = "A2"

# ============================================================
# SHEET 3: Position_Catalog_Required (Catalog job_titles cần có trước)
# ============================================================
ws3 = wb.create_sheet("Position_Catalog_Required")

headers3 = ["code", "name_vi", "source", "description"]
for col_idx, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

position_catalog = [
    {"code": "DRV_TRUCK_1_5_3_5", "name_vi": "Lái xe tải trọng 1.5-3.5 tấn", "source": "hrm_native", "description": "Lái xe tải nhẹ/trung bình, bằng B2/C"},
    {"code": "DRV_CONTAINER", "name_vi": "Lái xe container", "source": "hrm_native", "description": "Lái xe đầu kéo container cảng/ICD, bằng C/E, có thẻ ICC"},
    {"code": "DRV_BUS_FIXED", "name_vi": "Lái xe buýt tuyến cố định", "source": "hrm_native", "description": "Lái xe khách tuyến cố định, bằng D, khám sức khỏe 6 tháng"},
]

for row_idx, pos in enumerate(position_catalog, 2):
    for col_idx, header in enumerate(headers3, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=pos[header])
        cell.alignment = wrap_alignment
        cell.border = thin_border

col_widths3 = [25, 35, 18, 55]
for i, width in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = width

ws3.freeze_panes = "A2"

# ============================================================
# SHEET 4: Hướng dẫn sử dụng (dành cho người điều hành/HR)
# ============================================================
ws4 = wb.create_sheet("Hướng_dẫn")

guide_content = [
    ["HƯỚNG DẪN TẠO MẪU MÔ TẢ CÔNG VIỆC (JD) LÁI XE TRONG XEVN HRM", ""],
    ["", ""],
    ["1. CHUẨN BỊ DANH MỤC CHỨC DANH (BẮT BUỘC TRƯỚC KHI TẠO JD)", ""],
    ["", "Vào Cài đặt → Danh mục nghiệp vụ → tab \"Chức danh\" (job_titles)"],
    ["", "Tạo 3 chức danh sau (xem sheet Position_Catalog_Required):"],
    ["", "  - Mã: DRV_TRUCK_1_5_3_5 | Tên: Lái xe tải trọng 1.5-3.5 tấn | Nguồn: HRM"],
    ["", "  - Mã: DRV_CONTAINER     | Tên: Lái xe container           | Nguồn: HRM"],
    ["", "  - Mã: DRV_BUS_FIXED     | Tên: Lái xe buýt tuyến cố định | Nguồn: HRM"],
    ["", "Lưu ý: Chức danh phải tồn tại trước, mới chọn được khi tạo JD."],
    ["", ""],
    ["2. TẠO MẪU JD TRÊN GIAO DIỆN (UI)", ""],
    ["", "Vào Cài đặt → Thư viện JD master (tab jd-master-library)"],
    ["", "Bấm nút \"Thêm mới\" → Điền thông tin theo sheet JD_Templates:"],
    ["", "  - Mã mẫu JD: JD-DRIVER-001, JD-DRIVER-002, JD-DRIVER-003 (tự đặt, không trùng)"],
    ["", "  - Tiêu đề: Tên vị trí hiển thị cho ứng viên xem (VD: Lái xe tải trọng 1.5-3.5 tấn)"],
    ["", "  - Chức danh: BẮT BUỘC chọn từ danh sách thả xuống (các mã đã tạo ở bước 1)"],
    ["", "  - Mô tả công việc: Nội dung chi tiết công việc hàng ngày"],
    ["", "  - Yêu cầu ứng viên: Điều kiện bắt buộc & ưu tiên"],
    ["", "  - Ghi chú: Thông tin bổ sung (thử việc, phúc lợi, v.v.)"],
    ["", "  - Trạng thái: \"Đang dùng\" (active) để xuất bản, \"Nháp\" (draft) để lưu tạm"],
    ["", "Bấm \"Lưu nháp\" hoặc \"Xuất bản\""],
    ["", ""],
    ["3. THÔNG TIN CHI TIẾT BỔ SUNG (Sheet JD_Dynamic_Fields)", ""],
    ["", "Sau khi tạo JD xong, bạn có thể bổ sung các thông tin chi tiết ở sheet JD_Dynamic_Fields:"],
    ["", "Mỗi dòng = 1 thông tin của 1 mẫu JD. Các cột:"],
    ["", "  - Mã JD mẫu: Mã JD vừa tạo (VD: JD-DRIVER-001)"],
    ["", "  - Tên thông tin: Chọn loại thông tin (VD: Loại bằng lái xe, Mức lương tham khảo, ...)"],
    ["", "  - Giá trị mẫu: Điền giá trị cụ thể (VD: B2, 12-15 triệu/tháng, ...)"],
    ["", "  - Ghi chú/Hướng dẫn: Xem cột này để biết ý nghĩa, cách điền, ví dụ"],
    ["", ""],
    ["4. DANH SÁCH CÁC LOẠI THÔNG TIN CHI TIẾT CÓ THỂ ĐIỀN", ""],
    ["", "Dưới đây là các loại thông tin thường dùng cho JD lái xe (xem sheet JD_Dynamic_Fields):"],
    ["", "  - Loại bằng lái xe: B2 / C / D / E (theo quy định pháp luật)"],
    ["", "  - Số năm kinh nghiệm: Số nguyên (VD: 2, 3, 5)"],
    ["", "  - Khu vực chạy: Mô tả tuyến đường (VD: Nội thành Hà Nội, Cảng Hải Phòng - ICD Sóng Thần)"],
    ["", "  - Ca làm việc / Đòn bẩy: Cách sắp xếp ca (VD: 2 ngày/1 nghỉ, 2 ca/ngày, Theo lịch bến)"],
    ["", "  - Loại xe điều khiển: Dòng xe cụ thể (VD: Thùng 2 tấn, Đầu kéo 4x2, Buýt 45 chỗ)"],
    ["", "  - Mức lương tham khảo: Khoảng lương hoặc cách tính (VD: 12-15 triệu/tháng, 450k/chuyến)"],
    ["", "  - Quyền lợi được hưởng: Phúc lợi (VD: BHXH/BHYT/BHTN, ăn trưa, thưởng Tết, PPE)"],
    ["", "  - Thời gian thử việc (tháng): 1-2 tháng (theo Luật Lao động)"],
    ["", "  - Thẻ ICC (Quốc tế): Cho lái container ra cảng quốc tế (Đã có/Đang làm/Yêu cầu hỗ trợ)"],
    ["", "  - Khoảng cách khám sức khỏe (tháng): 6 tháng (bắt buộc cho lái xe khách theo QCVN)"],
    ["", ""],
    ["5. LƯU Ý QUAN TRỌNG", ""],
    ["", "- Chức danh (position_code) BẮT BUỘC phải có trong danh mục Chức danh trước khi tạo JD"],
    ["", "- Hệ thống sẽ từ chối nếu chọn chức danh không tồn tại"],
    ["", "- Trạng thái chỉ được chọn: Nháp (draft) | Đang dùng (active) | Ngưng (retired)"],
    ["", "- Thông tin chi tiết (sheet JD_Dynamic_Fields) là tùy chọn, giúp lọc/tìm kiếm JD chính xác hơn"],
    ["", "- Mỗi mẫu JD có thể có tập thông tin chi tiết khác nhau tùy đặc thù vị trí"],
]

for row_idx, (col1, col2) in enumerate(guide_content, 1):
    cell1 = ws4.cell(row=row_idx, column=1, value=col1)
    cell2 = ws4.cell(row=row_idx, column=2, value=col2)
    if row_idx == 1:
        cell1.font = Font(bold=True, size=14, color="2F5496")
    if col1 and not col2 and row_idx > 1:
        cell1.font = Font(bold=True, size=11)

ws4.column_dimensions['A'].width = 80
ws4.column_dimensions['B'].width = 20

# Save - use timestamp in filename to avoid OneDrive lock issues
from datetime import datetime
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
final_path = rf"C:\Users\ADMIN\OneDrive\Tài liệu\Vibe Coding\projects\xevn-ecosystem\docs\brand-new-documents-20270801\JD_Driver_Template_{timestamp}.xlsx"
wb.save(final_path)
print(f"Excel created: {final_path}")